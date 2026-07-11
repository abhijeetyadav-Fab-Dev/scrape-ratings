"""
Universal Hotel Data Scraper
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

A flexible, extensible scraping framework for pulling data from
multiple hotel extranet sources (Booking.com, MMT, etc.).

Architecture:
  Source Plugin (defines available fields & login/extract logic)
  -> ScrapeJob (config: which source + which fields)
  -> ScrapeJobRunner (executes the job via Playwright)
  -> CSV output

How to add a new source:
  1. Subclass ExtranetSource
  2. Define available_fields, source_name, login_url
  3. Implement login(), navigate_to_section(), extract_data()
  4. Register it in EXTRANET_SOURCES
"""

import sys, os, csv, json, time, re, threading, subprocess, pickle, sqlite3
from pathlib import Path
from abc import ABC, abstractmethod
from datetime import datetime, timedelta

# Set Playwright browser path to the user's local ms-playwright folder if running as a frozen executable
if getattr(sys, 'frozen', False):
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = str(Path.home() / "AppData" / "Local" / "ms-playwright")

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QCheckBox, QGroupBox, QTextEdit, QProgressBar, QFileDialog,
    QComboBox, QScrollArea, QFrame, QTabWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from playwright.sync_api import sync_playwright

# ────────────────────────────────────────────────────────────
#  Config & constants
# ────────────────────────────────────────────────────────────

COOKIES_DIR = Path.home() / ".scrape-ratings"
COOKIES_DIR.mkdir(exist_ok=True)

BOOKING_EXTRANET_COOKIES = COOKIES_DIR / "booking_extranet_cookies.pkl"
MMT_EXTRANET_COOKIES = COOKIES_DIR / "mmt_extranet_cookies.pkl"
GOIBIBO_EXTRANET_COOKIES = COOKIES_DIR / "goibibo_extranet_cookies.pkl"
AGODA_EXTRANET_COOKIES = COOKIES_DIR / "agoda_extranet_cookies.pkl"
EXPEDIA_EXTRANET_COOKIES = COOKIES_DIR / "expedia_extranet_cookies.pkl"


# Shared Chrome debug port (different from ratings scraper's 9222)
EXTRANET_DEBUG_PORT = 9223

def clean_row_keys(row: dict) -> dict:
    cleaned = {}
    prefixes = [
        "promo_", "res_", "dash_", "rate_", "prop_", "boost_", "inb_", "rev_", "fin_", "anl_",
        "mmt_rev_", "mmt_settlement_", "goi_rev_", "goi_settlement_", "agd_rev_", "agd_prop_",
        "exp_rev_", "exp_prop_", "htl_rev_", "htl_prop_", "exp_ci_", "htl_ci_", "goi_rpt_",
        "mmt_", "goi_", "agd_", "exp_", "htl_", "promo_mmt_"
    ]
    for k, v in row.items():
        if k in ["hotel_id", "hotel_name", "sub_tab", "_source", "_error"]:
            cleaned[k] = v
            continue
        cleaned_key = k
        for pref in prefixes:
            if k.startswith(pref):
                cleaned_key = k[len(pref):]
                break
        cleaned_key = cleaned_key.replace("_", " ")
        cleaned[cleaned_key] = v
    return cleaned

# ────────────────────────────────────────────────────────────
#  ExtranetSource — abstract base for all data-source plugins
# ────────────────────────────────────────────────────────────

class ExtranetSource(ABC):
    """Override these to define a new extranet data source."""

    source_key = ""


    @property
    @abstractmethod
    def source_name(self) -> str:
        """Human-readable name, e.g. 'Booking.com Extranet'"""
        ...

    @property
    @abstractmethod
    def login_url(self) -> str:
        """URL where the user logs in."""
        ...

    @property
    @abstractmethod
    def available_fields(self) -> list[dict]:
        """List of field groups & fields the user can select.
        Each entry: { "group": "Reservations", "fields": [
            {"key": "guest_name",    "label": "Guest Name"},
            {"key": "check_in",      "label": "Check-in Date"},
            {"key": "check_out",     "label": "Check-out Date"},
        ]}
        """
        ...

    @property
    def cookies_path(self) -> Path:
        """Path where session cookies are saved/loaded."""
        raise NotImplementedError

    @property
    def multi_tab(self) -> bool:
        """If True, the worker opens a new tab for each section.
        Useful for SPAs (like MMT) where navigating between sections
        on a single page causes session or rendering issues.
        """
        return False

    @abstractmethod
    def login(self, page) -> None:
        """Navigate to login_url and guide the user through login.
        This runs in headed mode — user enters credentials manually.
        After login, the cookies are saved automatically.
        """
        ...

    @abstractmethod
    def navigate_to_section(self, page, section_key: str) -> None:
        """Navigate to the page section that contains the requested data.
        section_key comes from the field's 'key' (first part before underscore).
        """
        ...

    @abstractmethod
    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows of data from the current page.
        selected_fields is the list of field dicts the user chose.
        Returns a list of dicts (one per row) with keys matching field['key'].
        """
        ...

    def _append_scraped_property_data(self, hotel_id: str, hotel_name: str, rows: list[dict], status: str = "Completed"):
        """Shared helper to write a property's scraped rows to Excel and log to SQLite in real-time."""
        session_id = getattr(self, "session_id", None)
        out_path = getattr(self, "output_path", None)
        
        # 1. Update SQLite
        if session_id:
            try:
                ScrapeHistoryManager.add_scraped_property(session_id, hotel_id, hotel_name, status, len(rows))
            except Exception:
                pass
                
        # 2. Append to file in real-time
        if out_path and rows:
            try:
                # Clean row keys first
                cleaned_rows = [clean_row_keys(r) for r in rows]

                # Construct ordered keys from job or dynamic discovery
                field_keys = []
                if getattr(self, "job", None):
                    field_keys = [clean_row_keys({f["key"]: ""}).popitem()[0] for f in self.job.selected_fields]
                else:
                    all_keys = set()
                    for r in cleaned_rows:
                        all_keys.update(r.keys())
                    for k in ["hotel_id", "hotel_name", "sub_tab", "_source", "_error"]:
                        all_keys.discard(k)
                    field_keys = sorted(list(all_keys))

                # Deduplicate columns to prevent duplicate hotel_id/hotel_name headers
                o_keys = []
                for k in field_keys:
                    if k not in o_keys:
                        o_keys.append(k)
                for k in ["hotel_id", "hotel_name", "sub_tab", "_source", "_error"]:
                    if k not in o_keys:
                        o_keys.append(k)

                if out_path.lower().endswith(".xlsx"):
                    from openpyxl import Workbook, load_workbook
                    if os.path.exists(out_path):
                        wb = load_workbook(out_path)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(o_keys)
                    
                    for r in cleaned_rows:
                        row_data = [r.get(k, "") for k in o_keys]
                        ws.append(row_data)
                    wb.save(out_path)
                else:
                    file_exists = os.path.exists(out_path) and os.path.getsize(out_path) > 0
                    with open(out_path, "a", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=o_keys, extrasaction="ignore")
                        if not file_exists:
                            writer.writeheader()
                        writer.writerows(cleaned_rows)
            except Exception:
                pass

    # ── Shared helper methods for section-aware extraction ──

    def _try_selectors(self, page, selectors: list[str]) -> str:
        """Try each CSS selector and return first non-empty inner_text."""
        for sel in selectors:
            try:
                el = page.query_selector(sel)
                if el:
                    text = el.inner_text().strip()
                    if text:
                        return text
            except Exception:
                continue
        return ""

    def _extract_value_by_label(self, page, label_text: str) -> str:
        """Find a form field by its associated label text.
        Uses a single page.evaluate call to handle all patterns:
        for= attribute, next sibling, parent's next sibling.
        """
        try:
            return page.evaluate(
                """(text) => {
                    const labels = document.querySelectorAll('label');
                    for (const label of labels) {
                        if (!label.textContent.includes(text)) continue;
                        const forAttr = label.getAttribute('for');
                        if (forAttr) {
                            const el = document.getElementById(forAttr);
                            if (el) {
                                const val = el.value || el.textContent || '';
                                if (val.trim()) return val.trim();
                            }
                        }
                        const sibling = label.nextElementSibling;
                        if (sibling) {
                            const txt = sibling.textContent || '';
                            if (txt.trim()) return txt.trim();
                        }
                        const parent = label.parentElement;
                        if (parent && parent.nextElementSibling) {
                            const txt = parent.nextElementSibling.textContent || '';
                            if (txt.trim()) return txt.trim();
                        }
                    }
                    return '';
                }""",
                label_text
            )
        except Exception:
            pass
        return ""

    def _extract_table_fields(self, page, field_keys: list[str],
                               known_columns: dict[str, str] = None) -> list[dict]:
        """Extract fields from data tables with column header matching.
        known_columns maps lowercase column header patterns -> field keys.
        """
        # First, check if the page is an error page
        is_error, error_context = self._is_error_page(page)
        if is_error:
            return [{"_error": error_context, "_source": "error_detected"}]

        rows = []
        try:
            tables = page.query_selector_all("table")
            for table in tables:
                header_els = table.query_selector_all("th")
                headers = [h.inner_text().strip().lower() for h in header_els]
                if not headers:
                    continue
                col_map = {}
                if known_columns:
                    for i, hdr in enumerate(headers):
                        for pattern, fk in known_columns.items():
                            if pattern in hdr and fk in field_keys:
                                col_map[i] = fk
                body_rows = table.query_selector_all("tbody tr, tr:not(:has(th))")
                for tr in body_rows:
                    cells = tr.query_selector_all("td")
                    row = {}
                    if col_map:
                        for i, cell in enumerate(cells):
                            if i in col_map:
                                row[col_map[i]] = cell.inner_text().strip()
                    else:
                        for i, cell in enumerate(cells):
                            if i < len(headers):
                                key = f"raw_{headers[i].replace(' ', '_')}"
                                row[key] = cell.inner_text().strip()
                    if row:
                        rows.append(row)
        except Exception:
            pass
        return rows

    def _extract_card_fields(self, page, field_keys: list[str],
                              card_selector: str, field_mappings: dict[str, str]) -> list[dict]:
        """Parse repeated card elements with child selectors per field.
        field_mappings: field_key -> CSS selector relative to each card.
        """
        rows = []
        try:
            cards = page.query_selector_all(card_selector)
            for card in cards:
                row = {}
                for fk, sel in field_mappings.items():
                    if fk in field_keys:
                        try:
                            el = card.query_selector(sel)
                            if el:
                                row[fk] = el.inner_text().strip()
                        except Exception:
                            pass
                if row:
                    rows.append(row)
        except Exception:
            pass
        return rows

    def _extract_metric_cards(self, page, field_keys: list[str],
                               label_map: dict[str, str],
                               card_selector: str = "") -> list[dict]:
        """Parse metric/KPI cards to extract value-label pairs.
        label_map: keyword in card text -> field key.
        """
        rows = []
        row = {}
        try:
            sel = card_selector or "[class*='metric'], [class*='kpi'], [class*='stat'], [class*='card'], [class*='widget']"
            cards = page.query_selector_all(sel)
            for card in cards:
                card_text = card.inner_text()
                lines = [l.strip() for l in card_text.split("\n") if l.strip()]
                text_lower = card_text.lower()
                for keyword, fk in label_map.items():
                    if keyword in text_lower and fk in field_keys and not row.get(fk):
                        for line in lines:
                            if keyword not in line.lower():
                                row[fk] = line
                                break
                        if not row.get(fk):
                            row[fk] = card_text.strip()[:300]
        except Exception:
            pass
        if row:
            rows.append(row)
        return rows


    def _is_error_page(self, page) -> tuple[bool, str]:
        """Detect if the page is showing an error, login page, or is not a valid data page.
        Returns (is_error, error_context).
        """
        error_indicators = [
            "sorry, this page does not exist",
            "sorry, this page isn't working",
            "page not found",
            "404 not found",
            "access denied",
            "you are not authorized",
            "please sign in",
            "sign in to continue",
            "session expired",
        ]
        try:
            url = page.url
            title = (page.evaluate("document.title") or "").lower()
            body_text = ""
            body = page.query_selector("body")
            if body:
                body_text = body.inner_text()[:2000].lower()

            for indicator in error_indicators:
                if indicator in body_text or indicator in title:
                    return True, f"Error page: '{indicator}' at {url}"

            if body_text and len(body_text.strip()) < 150:
                return True, f"Page too short ({len(body_text.strip())} chars) at {url}"
        except Exception:
            pass
        return False, ""

    def _generic_fallback(self, page) -> list[dict]:
        """Final fallback: try tables -> list items -> body text.
        Checks for error pages first and returns meaningful context instead of raw error text."""
        # First, check if the page is an error page
        is_error, error_context = self._is_error_page(page)
        if is_error:
            return [{"_error": error_context, "_source": "error_detected"}]

        rows = []
        try:
            tables = page.query_selector_all("table")
            if tables:
                for table in tables:
                    headers = [h.inner_text().strip().lower() for h in
                               table.query_selector_all("th")]
                    if not headers:
                        continue
                    body_rows = table.query_selector_all("tbody tr")
                    for tr in body_rows:
                        cells = tr.query_selector_all("td")
                        row = {}
                        for i, cell in enumerate(cells):
                            if i < len(headers):
                                key = f"raw_{headers[i].replace(' ', '_')}"
                                row[key] = cell.inner_text().strip()
                        if row:
                            rows.append(row)
        except Exception:
            pass
        if not rows:
            try:
                items = page.query_selector_all(
                    "[data-booking-id], .booking-item, .reservation-card, "
                    ".booking-row, [class*=booking], [class*=order], "
                    ".MuiTableRow-root, [data-testid*=booking]"
                )
                for item in items:
                    rows.append({"raw_data": item.inner_text().strip()})
            except Exception:
                pass
        if not rows:
            try:
                body = page.query_selector("body")
                if body:
                    body_text = body.inner_text()[:5000]
                    if len(body_text.strip()) >= 150:
                        rows.append({"raw_page_text": body_text})
                    else:
                        rows.append({"_error": f"No meaningful data (page has {len(body_text.strip())} chars)",
                                     "_source": "empty_page"})
            except Exception:
                pass
        return rows

    def _navigate_to_sub_tab(self, page, label: str) -> bool:
        """Find and click the sub-tab link or menu item matching the label text."""
        try:
            page.wait_for_timeout(1500) # let page settle
            
            # Lowercase label for matching
            lower_label = label.lower().strip()
            
            # Let's search and click using JS evaluate to handle shadow DOM, icons, and hidden elements
            clicked = page.evaluate("""(labelText) => {
                const lower = labelText.toLowerCase().trim();
                
                // Helper to score how well an element matches the label
                function getMatchScore(el) {
                    const text = el.textContent.toLowerCase().trim();
                    if (text === lower) return 3; // exact match
                    if (text.includes(lower)) {
                        // If it's a menu item or link, higher score
                        const tag = el.tagName.toLowerCase();
                        if (tag === 'a' || tag === 'button' || el.getAttribute('role') === 'tab' || el.classList.contains('bui-tab__link')) {
                            return 2;
                        }
                        return 1;
                    }
                    return 0;
                }
                
                let bestElement = null;
                let bestScore = 0;
                
                const candidates = document.querySelectorAll('a, button, li, [role="tab"], .bui-tab__link, span, div');
                for (const el of candidates) {
                    // Filter out elements that are not visible or too small
                    if (el.offsetWidth === 0 && el.offsetHeight === 0) continue;
                    
                    const score = getMatchScore(el);
                    if (score > bestScore) {
                        bestScore = score;
                        bestElement = el;
                    }
                }
                
                if (bestElement && bestScore >= 2) {
                    bestElement.click();
                    return true;
                }
                return false;
            }""", label)
            
            if clicked:
                page.wait_for_timeout(2000) # Wait for page load/settle
                return True
        except Exception:
            pass
        return False


# ────────────────────────────────────────────────────────────
#  Booking.com Extranet Source
# ────────────────────────────────────────────────────────────

class BookingExtranetSource(ExtranetSource):
    source_name = "Booking.com Extranet"
    login_url = "https://admin.booking.com/"

    SUB_TAB_FIELDS = {
        "dash_home", "rate_calendar", "rate_open_close", "rate_copy_future", "rate_plans", 
        "rate_value_adds", "rate_connectivity_errors", "rate_country_rates", "rate_mobile_rates",
        "promo_choose_new", "promo_simulate_max", "promo_active", "res_list", "prop_page_score", 
        "prop_general_info", "prop_vat_tax", "prop_photos", "prop_policies", "prop_res_policies", 
        "prop_facilities", "prop_room_details", "prop_room_amenities", "prop_descriptions", 
        "prop_messaging", "prop_sustainability", "boost_opportunity_center", "boost_genius_program", 
        "boost_preferred_program", "boost_long_stays", "boost_visibility_booster", "boost_smart_flex", 
        "boost_sponsored_listings", "inb_reservation_messages", "inb_booking_messages", 
        "inb_guest_qa", "rev_guest_reviews", "rev_guest_experience", "fin_payout_info", 
        "fin_invoices_docs", "fin_res_statement", "fin_overview", "fin_help", "fin_settings", 
        "anl_dashboard", "anl_demand", "anl_pace", "anl_sales_stats", "anl_booker_insights", 
        "anl_book_window", "anl_cancellation_char", "anl_comparable", "anl_genius_report", 
        "anl_ranking", "anl_performance"
    }

    SUB_TAB_URLS = {
        "dash_home": "home.html",
        "rate_calendar": "rates_availability.html",
        "rate_open_close": "rates_availability.html",
        "rate_copy_future": "rates_availability.html",
        "rate_plans": "rate_plans.html",
        "rate_value_adds": "value_adds.html",
        "rate_connectivity_errors": "connectivity_errors.html",
        "rate_country_rates": "country_rates.html",
        "rate_mobile_rates": "mobile_rates.html",
        "promo_choose_new": "promotions/list.html",
        "promo_simulate_max": "promotions/list.html",
        "promo_active": "promotions/list.html",
        "res_list": "search_reservations.html",
        "prop_page_score": "content_score.html",
        "prop_general_info": "contacts.html",
        "prop_vat_tax": "vat_tax.html",
        "prop_photos": "photos.html",
        "prop_policies": "policies.html",
        "prop_res_policies": "reservation_policies.html",
        "prop_facilities": "facilities.html",
        "prop_room_details": "room_details.html",
        "prop_room_amenities": "room_amenities.html",
        "prop_descriptions": "descriptions.html",
        "prop_messaging": "messaging_settings.html",
        "prop_sustainability": "sustainability.html",
        "boost_opportunity_center": "opportunities.html",
        "boost_genius_program": "genius.html",
        "boost_preferred_program": "preferred.html",
        "boost_long_stays": "long_stays.html",
        "boost_visibility_booster": "visibility_booster.html",
        "boost_smart_flex": "smart_flex.html",
        "boost_sponsored_listings": "sponsored_listings.html",
        "inb_reservation_messages": "messaging/inbox.html",
        "inb_booking_messages": "messaging/inbox.html",
        "inb_guest_qa": "messaging/inbox.html",
        "rev_guest_reviews": "reviews.html",
        "rev_guest_experience": "reviews.html",
        "fin_payout_info": "finance_payout.html",
        "fin_invoices_docs": "finance_invoices.html",
        "fin_res_statement": "finance_reservations.html",
        "fin_overview": "finance_overview.html",
        "fin_help": "finance_help.html",
        "fin_settings": "finance_settings.html",
        "anl_dashboard": "statistics/index.html",
        "anl_demand": "statistics/demand.html",
        "anl_pace": "statistics/pace.html",
        "anl_sales_stats": "statistics/sales.html",
        "anl_booker_insights": "statistics/booker_insights.html",
        "anl_book_window": "statistics/book_window.html",
        "anl_cancellation_char": "statistics/cancellation_char.html",
        "anl_comparable": "statistics/comparable.html",
        "anl_genius_report": "statistics/genius_report.html",
        "anl_ranking": "statistics/ranking.html",
        "anl_performance": "statistics/index.html",
    }

    @property
    def cookies_path(self):
        return BOOKING_EXTRANET_COOKIES

    @property
    def available_fields(self):
        return [
            {
                "group": "Home",
                "section": "dashboard",
                "fields": [
                    {"key": "dash_home", "label": "Home Dashboard (Sub-Tab)"},
                    {"key": "dash_occupancy", "label": "Occupancy"},
                    {"key": "dash_revenue_ytd", "label": "Revenue YTD"},
                    {"key": "dash_avg_daily_rate", "label": "Average Daily Rate"},
                    {"key": "dash_revpar", "label": "RevPAR"},
                    {"key": "dash_bookings_today", "label": "Bookings Today"},
                    {"key": "dash_check_ins_today", "label": "Check-ins Today"},
                    {"key": "dash_check_outs_today", "label": "Check-outs Today"},
                    {"key": "dash_net_revenue", "label": "Net Revenue"},
                    {"key": "dash_commission_total", "label": "Total Commission"},
                ]
            },
            {
                "group": "Rates & availability",
                "section": "rates",
                "fields": [
                    {"key": "rate_calendar", "label": "Calendar (Sub-Tab)"},
                    {"key": "rate_open_close", "label": "Open/close rooms (Sub-Tab)"},
                    {"key": "rate_copy_future", "label": "Copy rates to future dates (Sub-Tab)"},
                    {"key": "rate_plans", "label": "Rate plans (Sub-Tab)"},
                    {"key": "rate_value_adds", "label": "Value adds (Sub-Tab)"},
                    {"key": "rate_connectivity_errors", "label": "Connectivity errors (Sub-Tab)"},
                    {"key": "rate_country_rates", "label": "Country rates (Sub-Tab)"},
                    {"key": "rate_mobile_rates", "label": "Mobile rates (Sub-Tab)"},
                ]
            },
            {
                "group": "Promotions",
                "section": "promotions",
                "fields": [
                    {"key": "promo_choose_new", "label": "Choose new promotion (Sub-Tab)"},
                    {"key": "promo_simulate_max", "label": "Simulate max discount (Sub-Tab)"},
                    {"key": "promo_active", "label": "Your active promotions (Sub-Tab)"},
                    {"key": "promo_Name", "label": "Offer Name"},
                    {"key": "promo_Discount", "label": "Discount % / Amount"},
                    {"key": "promo_Bookable_period", "label": "Valid From"},
                    {"key": "promo_Stay_dates", "label": "Valid To"},
                    {"key": "promo_Bookings", "label": "Bookings"},
                    {"key": "promo_Room_nights", "label": "Room Nights"},
                    {"key": "promo_Average_daily_rate", "label": "Average Daily Rate"},
                    {"key": "promo_Revenue", "label": "Revenue"},
                    {"key": "promo_Cancelled_room_nights", "label": "Cancelled Nights"},
                    {"key": "promo_Status", "label": "Status"},
                ]
            },
            {
                "group": "Reservations",
                "section": "reservations",
                "fields": [
                    {"key": "res_list", "label": "Reservations list (Sub-Tab)"},
                    {"key": "res_guest_name", "label": "Guest Name"},
                    {"key": "res_check_in", "label": "Check-in Date"},
                    {"key": "res_check_out", "label": "Check-out Date"},
                    {"key": "res_room_type", "label": "Room Type"},
                    {"key": "res_status", "label": "Booking Status"},
                    {"key": "res_rate_plan", "label": "Rate Plan"},
                ]
            },
            {
                "group": "Property",
                "section": "property",
                "fields": [
                    {"key": "prop_page_score", "label": "Property Page Score (Sub-Tab)"},
                    {"key": "prop_general_info", "label": "General info & property status (Sub-Tab)"},
                    {"key": "prop_vat_tax", "label": "VAT/Tax/Charges (Sub-Tab)"},
                    {"key": "prop_photos", "label": "Photos (Sub-Tab)"},
                    {"key": "prop_policies", "label": "Property policies (Sub-Tab)"},
                    {"key": "prop_res_policies", "label": "Reservation policies (Sub-Tab)"},
                    {"key": "prop_facilities", "label": "Facilities & services (Sub-Tab)"},
                    {"key": "prop_room_details", "label": "Room details (Sub-Tab)"},
                    {"key": "prop_room_amenities", "label": "Room amenities (Sub-Tab)"},
                    {"key": "prop_descriptions", "label": "View Your Descriptions (Sub-Tab)"},
                    {"key": "prop_messaging", "label": "Messaging Preferences (Sub-Tab)"},
                    {"key": "prop_sustainability", "label": "Sustainability (Sub-Tab)"},
                    {"key": "prop_name", "label": "Property Name"},
                    {"key": "prop_description", "label": "Description"},
                    {"key": "prop_amenities", "label": "Amenities"},
                    {"key": "prop_room_types", "label": "Room Types"},
                    {"key": "prop_facilities_data", "label": "Facilities"},
                    {"key": "prop_policies_data", "label": "Policies"},
                    {"key": "prop_house_rules", "label": "House Rules"},
                    {"key": "prop_photos_data", "label": "Photos"},
                ]
            },
            {
                "group": "Boost performance",
                "section": "boost",
                "fields": [
                    {"key": "boost_opportunity_center", "label": "Opportunity Center (Sub-Tab)"},
                    {"key": "boost_genius_program", "label": "Genius Partner Program (Sub-Tab)"},
                    {"key": "boost_preferred_program", "label": "Preferred Partner Program (Sub-Tab)"},
                    {"key": "boost_long_stays", "label": "Long stays toolkit (Sub-Tab)"},
                    {"key": "boost_visibility_booster", "label": "Visibility Booster (Sub-Tab)"},
                    {"key": "boost_smart_flex", "label": "Smart Flex Reservations program (Sub-Tab)"},
                    {"key": "boost_sponsored_listings", "label": "Sponsored Listings (Sub-Tab)"},
                    {"key": "boost_visibility_score", "label": "Visibility Score"},
                    {"key": "boost_preferred_status", "label": "Preferred Status"},
                    {"key": "boost_genius_tier", "label": "Genius Tier"},
                    {"key": "boost_conversion_rate", "label": "Conversion Rate"},
                    {"key": "boost_competitor_rank", "label": "Competitor Rank"},
                    {"key": "boost_search_views", "label": "Search Views"},
                    {"key": "boost_property_views", "label": "Property Views"},
                    {"key": "boost_bookings", "label": "Bookings"},
                    {"key": "boost_cancellations", "label": "Cancellations"},
                ]
            },
            {
                "group": "Inbox",
                "section": "inbox",
                "fields": [
                    {"key": "inb_reservation_messages", "label": "Reservation messages (Sub-Tab)"},
                    {"key": "inb_booking_messages", "label": "Booking.com Messages (Sub-Tab)"},
                    {"key": "inb_guest_qa", "label": "Guest Q&A (Sub-Tab)"},
                    {"key": "inb_guest_name", "label": "Guest Name"},
                    {"key": "inb_subject", "label": "Subject"},
                    {"key": "inb_message", "label": "Message"},
                    {"key": "inb_date", "label": "Date"},
                    {"key": "inb_status", "label": "Status"},
                ]
            },
            {
                "group": "Guest Reviews",
                "section": "reviews",
                "fields": [
                    {"key": "rev_guest_reviews", "label": "Guest Reviews (Sub-Tab)"},
                    {"key": "rev_guest_experience", "label": "Guest experience (Sub-Tab)"},
                    {"key": "rev_guest_name", "label": "Guest Name"},
                    {"key": "rev_score", "label": "Rating / Score"},
                    {"key": "rev_comment", "label": "Review Comment"},
                    {"key": "rev_date", "label": "Review Date"},
                    {"key": "rev_response", "label": "Review Response"},
                ]
            },
            {
                "group": "Finance",
                "section": "financial",
                "fields": [
                    {"key": "fin_payout_info", "label": "Payout info (Sub-Tab)"},
                    {"key": "fin_invoices_docs", "label": "Invoices and documents (Sub-Tab)"},
                    {"key": "fin_res_statement", "label": "Reservations statement (Sub-Tab)"},
                    {"key": "fin_overview", "label": "Financial Overview (Sub-Tab)"},
                    {"key": "fin_help", "label": "Finance Help (Sub-Tab)"},
                    {"key": "fin_settings", "label": "Finance settings (Sub-Tab)"},
                ]
            },
            {
                "group": "Analytics",
                "section": "analytics",
                "fields": [
                    {"key": "anl_dashboard", "label": "Analytics dashboard (Sub-Tab)"},
                    {"key": "anl_demand", "label": "Demand for BangaloreNew (Sub-Tab)"},
                    {"key": "anl_pace", "label": "Your pace of bookings (Sub-Tab)"},
                    {"key": "anl_sales_stats", "label": "Sales Statistics (Sub-Tab)"},
                    {"key": "anl_booker_insights", "label": "Booker insights (Sub-Tab)"},
                    {"key": "anl_book_window", "label": "Book Window Info (Sub-Tab)"},
                    {"key": "anl_cancellation_char", "label": "Cancellation Characteristics (Sub-Tab)"},
                    {"key": "anl_comparable", "label": "Comparable properties (Sub-Tab)"},
                    {"key": "anl_genius_report", "label": "Genius Report (Sub-Tab)"},
                    {"key": "anl_ranking", "label": "Ranking Dashboard (Sub-Tab)"},
                    {"key": "anl_performance", "label": "Performance dashboard (Sub-Tab)"},
                    {"key": "anl_sales_revenue", "label": "Sales Revenue"},
                    {"key": "anl_room_nights", "label": "Room Nights"},
                    {"key": "anl_adr", "label": "ADR"},
                    {"key": "anl_cancellations", "label": "Cancellations"},
                    {"key": "anl_page_views", "label": "Page Views"},
                    {"key": "anl_click_through", "label": "CTR"},
                    {"key": "anl_booking_demand", "label": "Booking Demand"},
                    {"key": "anl_market_share", "label": "Market Share"},
                    {"key": "anl_competitor_pricing", "label": "Competitor Pricing"},
                    {"key": "anl_booking_window", "label": "Booking Window"},
                    {"key": "anl_country", "label": "Country"},
                    {"key": "anl_device", "label": "Device"},
                    {"key": "anl_traveler", "label": "Traveler"},
                ]
            },
        ]

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _fast_regex_extract(self, html: str, url: str, field_keys: list[str], section: str) -> list[dict]:
        """
        Fast Regex parsing for Booking.com Extranet.
        Matches exact HTML structure from the endpoints.
        """
        row = {}
        if section == "reviews":
            if "guest reviews" in html.lower():
                ms = re.search(r'bui-review-score__badge">\s*([\d.]+)\s*<', html)
                mc = re.search(r'based on\s*([\d,]+)\s*reviews?', html, re.IGNORECASE)
                if "rev_score" in field_keys:
                    row["rev_score"] = ms.group(1) if ms else ""
                if "rev_count" in field_keys or True: # Add count if not requested just in case
                    row["rev_count"] = mc.group(1).replace(",", "") if mc else "0"
                if row:
                    return [row]
        
        elif section == "boost":
            if "boost_genius_program" in field_keys and "GeniusBaseProgrammeConfig" in html:
                m = re.search(r'"status":"(ACTIVE|INACTIVE)","productConfig":\{"__typename":"GeniusBaseProgrammeConfig","isEligible":(?:true|false),"eligibilityStatus":"(\w+)"', html)
                mp = re.search(r'"isPriceCompetitive":(true|false)', html)
                if m:
                    status = m.group(1)
                    elig = m.group(2)
                    gs = "Enrolled" if status == "ACTIVE" else ("Eligible" if elig == "ELIGIBLE" else "Not Eligible")
                    row["boost_genius_program"] = gs
                if mp:
                    row["genius_competitive"] = "Competitive" if mp.group(1) == "true" else "Not Competitive"
                if row:
                    return [row]
            
            if "boost_preferred_program" in field_keys and "isPreferred" in html:
                is_pref = re.search(r'"isPreferred":\s*(\d+)', html)
                is_plus = re.search(r'"isPreferredPlus":\s*(\d+)', html)
                if is_pref:
                    ip = int(is_pref.group(1))
                    ipl = int(is_plus.group(1)) if is_plus else 0
                    if ip == 1:
                        status = "You're a member" + (" (Plus)" if ipl == 1 else "")
                    elif ip == 0:
                        status = "Not Enrolled"
                    else:
                        status = ""
                    row["boost_preferred_program"] = status
                    return [row]

        elif section == "dashboard":
            if "perf_score" in field_keys or "performanceScore" in html:
                m = re.search(r'"performanceScore":\s*\{"formattedScore":"([\d.]+%?)"', html)
                if m:
                    row["perf_score"] = m.group(1)
                    return [row]
        
        return []

    def navigate_to_section(self, page, section_key: str) -> None:
        worker = getattr(self, "worker", None)
        # Load from saved params file if present to check if we can bypass scanning
        params_path = COOKIES_DIR / f"{self.source_key}_params.json"
        if params_path.exists() and not getattr(self, "current_hotel_id", None):
            try:
                with open(params_path, "r") as f:
                    saved_params = json.load(f)
                self.current_hotel_id = saved_params.get("hotel_id")
                self.current_ses = saved_params.get("ses")
            except Exception:
                pass

        # Check if we are on the Group Homepage
        current_url = page.url.lower()
        if "/groups/home/" in current_url and not getattr(self, "current_hotel_id", None):
            if worker:
                worker.log_msg.emit("On Group Homepage. Waiting for portfolio list to load...")
            try:
                page.wait_for_selector("a[href*='hotel_id'], [data-hotel-id], table, .bui-table", timeout=5000)
            except Exception:
                pass
            if worker:
                worker.log_msg.emit("Fetching property list via GraphQL API...")
            properties = []
            
            # Fetch property list via GraphQL query GroupProperties
            try:
                js_query = """
                async () => {
                    try {
                        const propsQuery = {
                            query: `query GroupProperties {
                                propertyList {
                                    properties {
                                        hotelId
                                        hotelName
                                    }
                                }
                            }`
                        };
                        const r = await fetch('/dml/graphql', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify(propsQuery)
                        });
                        const d = await r.json();
                        if (d && d.data && d.data.propertyList && d.data.propertyList.properties) {
                            return d.data.propertyList.properties.map(p => ({
                                id: String(p.hotelId),
                                name: p.hotelName || ""
                            }));
                        }
                    } catch (e) {
                        return [];
                    }
                    return [];
                }
                """
                properties = page.evaluate(js_query)
            except Exception as e:
                if worker:
                    worker.log_msg.emit(f"GraphQL GroupProperties query failed: {e}. Falling back to page DOM scan.")
                properties = []

            if not properties:
                # Fallback: scan properties on the current Group page without paginating
                try:
                    properties = page.evaluate("""() => {
                        const props = [];
                        const elements = document.querySelectorAll('a, button, [data-hotel-id], [data-id]');
                        for (const el of elements) {
                            let hotelId = '';
                            let hotelName = '';
                            
                            if (el.getAttribute('data-hotel-id')) {
                                hotelId = el.getAttribute('data-hotel-id');
                            } else if (el.getAttribute('data-id') && /^\d+$/.test(el.getAttribute('data-id'))) {
                                hotelId = el.getAttribute('data-id');
                            } else {
                                const href = el.getAttribute('href') || '';
                                const match = href.match(/hotel_id=(\d+)/) || href.match(/\/hotel\/(\d+)/);
                                if (match) {
                                    hotelId = match[1];
                                }
                            }
                            
                            if (!hotelId || !/^\d+$/.test(hotelId)) continue;
                            
                            hotelName = el.textContent.trim();
                            if (hotelName.length < 4 || /manage|select|go|enter|edit|open/i.test(hotelName) || /^\d+$/.test(hotelName)) {
                                const row = el.closest('tr') || el.closest('[class*="row"]') || el.closest('[class*="item"]') || el.closest('li');
                                if (row) {
                                    const cells = row.querySelectorAll('td, div, span, a');
                                    for (const cell of cells) {
                                        if (cell === el) continue;
                                        const txt = cell.textContent.trim();
                                        if (txt.length >= 4 && !txt.includes('hotel_id') && !/^\d+$/.test(txt) && !/manage|select|go|enter|edit|open/i.test(txt)) {
                                            hotelName = txt;
                                            break;
                                        }
                                    }
                                }
                            }
                            
                            hotelName = hotelName.replace(/\s+/g, ' ').trim();
                            
                            if (!props.some(p => p.id === hotelId)) {
                                props.push({ id: hotelId, name: hotelName });
                            }
                        }
                        return props;
                    }""")
                except Exception:
                    pass

            if worker:
                worker.log_msg.emit(f"Discovered {len(properties)} properties in portfolio.")
            self._properties = properties
        else:
            self._properties = []

        # Extract session parameters from current page URL (ses, hotel_account_id, hotel_id)
        # First, wait up to 10 seconds for session parameters to appear in the URL (if the page is loading/redirecting)
        ses_match = None
        account_match = None
        hotel_match = None
        
        # Load from saved params file if present
        params_path = COOKIES_DIR / f"{self.source_key}_params.json"
        if params_path.exists():
            try:
                with open(params_path, "r") as f:
                    saved_params = json.load(f)
                if not getattr(self, "current_hotel_id", None):
                    self.current_hotel_id = saved_params.get("hotel_id")
                if not getattr(self, "current_ses", None):
                    self.current_ses = saved_params.get("ses")
            except Exception:
                pass

        # Only wait if we do not have cached parameters, or if we are actively on an admin page that hasn't loaded params in the URL yet
        has_cached_params = bool(getattr(self, "current_ses", None) and getattr(self, "current_hotel_id", None))
        if not has_cached_params or ("admin.booking.com" in page.url.lower() and "ses=" not in page.url.lower()):
            for _ in range(10):
                current_url = page.url
                ses_match = re.search(r'ses=([a-f0-9]+)', current_url)
                account_match = re.search(r'hotel_account_id=(\d+)', current_url)
                hotel_match = re.search(r'hotel_id=(\d+)', current_url)
                if ses_match or account_match or hotel_match:
                    break
                page.wait_for_timeout(1000)


        # If no ses params and we don't have cached session parameters, navigate to login to establish session
        if not ses_match and not account_match and not hotel_match and not getattr(self, "current_ses", None):
            # Check if we are already on a login or error page. If so, don't trigger redundant page loads
            if "login" not in page.url.lower():
                page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)
            current_url = page.url
            ses_match = re.search(r'ses=([a-f0-9]+)', current_url)
            account_match = re.search(r'hotel_account_id=(\d+)', current_url)
            hotel_match = re.search(r'hotel_id=(\d+)', current_url)

        base = "https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage"

        params = []
        
        # Use match if found, otherwise use self.current_ses
        ses = None
        if ses_match:
            ses = ses_match.group(1)
            self.current_ses = ses
        elif getattr(self, "current_ses", None):
            ses = self.current_ses
        if ses:
            params.append(f"ses={ses}")
            
        if account_match:
            params.append(f"hotel_account_id={account_match.group(1)}")
            
        # Override hotel_id if we scanned properties from group homepage
        properties = getattr(self, "_properties", [])
        hotel_id = None
        if properties:
            hotel_id = properties[0]['id']
            self.current_hotel_id = hotel_id
        elif hotel_match:
            hotel_id = hotel_match.group(1)
            self.current_hotel_id = hotel_id
        elif getattr(self, "current_hotel_id", None):
            hotel_id = self.current_hotel_id
            
        if hotel_id:
            params.append(f"hotel_id={hotel_id}")
            
        # Persist updated parameters to JSON if we have both
        if ses and hotel_id:
            try:
                params_path = COOKIES_DIR / f"{self.source_key}_params.json"
                with open(params_path, "w") as f:
                    json.dump({"hotel_id": hotel_id, "ses": ses}, f)
            except Exception:
                pass

        params.append("lang=en")
        param_str = "?" + "&".join(params) if params else ""


        section_map = {
            "dashboard":     f"{base}/home.html{param_str}",
            "reservations":  f"{base}/search_reservations.html{param_str}",
            "rates":         f"{base}/rates_availability.html{param_str}",
            "property":      f"{base}/content_score.html{param_str}",
            "boost":         f"{base}/opportunities.html{param_str}",
            "inbox":         f"{base}/messaging/inbox.html{param_str}",
            "reviews":       f"{base}/reviews.html{param_str}",
            "financial":     f"{base}/finance_overview.html{param_str}",
            "analytics":     f"{base}/statistics/index.html{param_str}",
            "promotions":    f"{base}/promotions/list.html{param_str}",
        }
        url = section_map.get(section_key, base)
        
        # Load section via fast fetch if possible
        use_fast_fetch = getattr(getattr(self, 'job', None), 'fast_mode', True)
        if use_fast_fetch:
            if worker:
                worker.log_msg.emit(f"  -> Background fetching section '{section_key}'...")
            html = self._fast_fetch_html(page, url)
            is_login_redirect = "sign-in" in html or "op_token" in html or len(html) < 200
            if html and not html.startswith("FETCH_ERROR") and not is_login_redirect:
                page.set_content(html, wait_until="commit")
            else:
                page.goto(url, timeout=30000, wait_until="domcontentloaded")
        else:
            page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    # ── Section-aware field extraction ──────────────────────

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which section we're on by field key prefix."""
        prefix_map = {
            "dash": "dashboard", "res": "reservations", "rate": "rates",
            "prop": "property", "boost": "boost", "inb": "inbox",
            "rev": "reviews", "fin": "financial", "anl": "analytics",
            "promo": "promotions",
        }
        for key in field_keys:
            prefix = key.split("_")[0]
            if prefix in prefix_map:
                return prefix_map[prefix]
        return "general"

    def _extract_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract Property Details fields from the Booking.com extranet property page."""
        row = {}

        # ── Property Name ────────────────────────────────────
        if "prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1",
                "h1[class*='name']", "h1[class*='title']",
                "input[name='name']", "input[name*='hotel_name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[data-name='hotel-name']", "[data-name='property-name']",
                "span[class*='bui-header__title']",
                ".bui-header__title",
                "div[class*='page-title']",
                "[class*='main-title']",
                "[class*='headline']",
            ])
            if not val:
                # Fallback: page <title> tag
                try:
                    val = page.evaluate("document.title").strip()
                    # Clean common suffixes
                    for suffix in [" - Booking.com", " - Booking.com Extranet", " | Booking.com", " - Manage your property"]:
                        if val.endswith(suffix):
                            val = val[:-len(suffix)].strip()
                            break
                except Exception:
                    pass
            row["prop_name"] = val or ""

        # ── Description ──────────────────────────────────────
        if "prop_description" in field_keys:
            val = self._try_selectors(page, [
                "textarea[name*='description']", "textarea[id*='description']",
                "[class*='description']", "[data-name='description']",
                "div[class*='desc']", "textarea[class*='desc']",
                ".editor-content", "[contenteditable='true']",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Description")
                if not val:
                    val = self._extract_value_by_label(page, "desc")
            row["prop_description"] = val or ""

        # ── Amenities ────────────────────────────────────────
        if "prop_amenities" in field_keys:
            val = self._try_selectors(page, [
                "[class*='amenity']", "[data-name='amenities']",
                ".amenities-list", ".facilities-list",
            ])
            if not val:
                # Collect text from checked checkboxes
                checked_items = page.evaluate("""() => {
                    const checked = document.querySelectorAll('input[type="checkbox"]:checked');
                    const texts = [];
                    for (const cb of checked) {
                        const label = cb.closest('label');
                        if (label) texts.push(label.textContent.trim());
                        else {
                            const parent = cb.parentElement;
                            if (parent) texts.push(parent.textContent.trim());
                        }
                    }
                    return texts.join('; ');
                }""")
                if checked_items and checked_items.strip():
                    val = checked_items.strip()
            row["prop_amenities"] = val or ""

        # ── Room Types ───────────────────────────────────────
        if "prop_room_types" in field_keys:
            val = self._try_selectors(page, [
                "[class*='room-type']", "[class*='roomtype']", "[data-name='room-types']",
                "[class*='room'] h3", "[class*='room'] h4",
                ".room-list", "[class*='unit-type']",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Room Types")
            row["prop_room_types"] = val or ""

        # ── Facilities & Services ────────────────────────────
        if "prop_facilities" in field_keys:
            val = self._try_selectors(page, [
                "[class*='facilities']", "[class*='facility']",
                "[data-name='facilities']", "[class*='services']",
                ".facilities-list", ".services-list",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Facilities")
            row["prop_facilities"] = val or ""

        # ── Policies ─────────────────────────────────────────
        if "prop_policies" in field_keys:
            val = self._try_selectors(page, [
                "[class*='policy']", "[data-name='policies']",
                "[class*='policy-section']",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Policies")
            row["prop_policies"] = val or ""

        # ── House Rules ──────────────────────────────────────
        if "prop_house_rules" in field_keys:
            val = self._try_selectors(page, [
                "[class*='house-rule']", "[class*='house_rule']",
                "[data-name='house-rules']", "[class*='rules']",
            ])
            if not val:
                val = self._extract_value_by_label(page, "House Rules")
            row["prop_house_rules"] = val or ""

        # ── Photos ───────────────────────────────────────────
        if "prop_photos" in field_keys:
            photo_urls = page.evaluate("""() => {
                const imgs = document.querySelectorAll('img[class*="photo"], img[class*="gallery"], [class*="photo"] img, [class*="gallery"] img');
                const urls = [];
                for (const img of imgs) {
                    const src = img.getAttribute('src') || img.getAttribute('data-src') || '';
                    if (src && !urls.includes(src)) urls.push(src);
                }
                return urls.join('; ');
            }""")
            row["prop_photos"] = photo_urls.strip() if photo_urls else ""

        return row

    def _extract_dashboard_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract Dashboard / Home metrics from the Booking.com extranet dashboard.
        Looks for metric cards / KPI boxes that contain a value and label.
        """
        rows = []
        row = {}

        # Try to find metric cards — common patterns in Booking.com extranet dashboard
        # Often these are structured as: <value> <label> pairs in cards
        try:
            # Pattern 1: cards with a number/big text and a label below
            metric_cards = page.query_selector_all(
                "[class*='metric'], [class*='kpi'], [class*='stat'], "
                "[class*='dashboard-card'], [class*='summary-card'], "
                "[class*='overview'] > div, [class*='widget']"
            )
            if metric_cards:
                for card in metric_cards:
                    card_text = card.inner_text().strip()
                    if not card_text:
                        continue
                    lines = [l.strip() for l in card_text.split("\n") if l.strip()]
                    # Look for known labels
                    label_map = {
                        "occupancy": "dash_occupancy", "revenue ytd": "dash_revenue_ytd",
                        "average daily rate": "dash_avg_daily_rate", "adr": "dash_avg_daily_rate",
                        "revpar": "dash_revpar", "bookings today": "dash_bookings_today",
                        "check-ins": "dash_check_ins_today", "check-ins today": "dash_check_ins_today",
                        "check-outs": "dash_check_outs_today", "check-outs today": "dash_check_outs_today",
                        "net revenue": "dash_net_revenue", "total commission": "dash_commission_total",
                        "commission": "dash_commission_total",
                    }
                    for line in lines:
                        line_lower = line.lower()
                        for keyword, field_key in label_map.items():
                            if keyword in line_lower and field_key in field_keys and not row.get(field_key):
                                # The value is likely the numeric part (first line or adjacent)
                                if lines.index(line) > 0:
                                    val = lines[lines.index(line) - 1]
                                else:
                                    val = line
                                row[field_key] = val
                                break
        except Exception:
            pass

        if row:
            rows.append(row)
        return rows

    def _extract_review_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract review data from review cards/lists.
        Booking.com extranet reviews show guest name, score, comment, etc.
        """
        rows = []

        # Try table-based extraction first (column mapping)
        table_rows = self._extract_table_fields(page, field_keys, {
            "guest": "rev_guest_name", "name": "rev_guest_name",
            "score": "rev_score", "rating": "rev_score",
            "comment": "rev_comment", "review": "rev_comment",
            "response": "rev_response", "reply": "rev_response",
            "date": "rev_date", "language": "rev_language", "lang": "rev_language",
        })
        if table_rows:
            return table_rows

        # Fallback: parse review cards
        try:
            cards = page.query_selector_all(
                "[class*='review'], [class*='feedback'], "
                ".review-card, .guest-review, [data-review-id]"
            )
            for card in cards:
                row = {}
                full_text = card.inner_text()
                if "rev_guest_name" in field_keys:
                    name_el = card.query_selector("[class*='name'], [class*='guest'], h3, h4")
                    row["rev_guest_name"] = name_el.inner_text().strip() if name_el else ""
                if "rev_score" in field_keys:
                    score_el = card.query_selector("[class*='score'], [class*='rating'], [class*='grade']")
                    row["rev_score"] = score_el.inner_text().strip() if score_el else ""
                if "rev_comment" in field_keys:
                    comment_el = card.query_selector("[class*='comment'], [class*='review-text'], [class*='text']")
                    row["rev_comment"] = comment_el.inner_text().strip() if comment_el else ""
                if "rev_date" in field_keys:
                    date_el = card.query_selector("[class*='date'], [class*='time'], time")
                    row["rev_date"] = date_el.inner_text().strip() if date_el else ""
                if "rev_response" in field_keys:
                    resp_el = card.query_selector("[class*='response'], [class*='reply']")
                    row["rev_response"] = resp_el.inner_text().strip() if resp_el else ""
                if row:
                    rows.append(row)
        except Exception:
            pass

        return rows

    def _extract_inbox_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract messages from the Inbox section."""
        rows = []

        # Try table first with column mapping
        table_rows = self._extract_table_fields(page, field_keys, {
            "guest": "inb_guest_name", "name": "inb_guest_name",
            "from": "inb_guest_name", "sender": "inb_guest_name",
            "reservation": "inb_reservation_id", "booking": "inb_reservation_id", "id": "inb_reservation_id",
            "subject": "inb_subject", "message": "inb_message",
            "date": "inb_date", "time": "inb_date",
            "status": "inb_status", "read": "inb_status", "unread": "inb_status",
        })
        if table_rows:
            return table_rows

        # Fallback: parse message list items
        try:
            items = page.query_selector_all(
                "[class*='message'], .inbox-item, [class*='conversation'], "
                "tr[class*='message'], [data-message-id]"
            )
            for item in items:
                row = {}
                if "inb_guest_name" in field_keys:
                    el = item.query_selector("[class*='sender'], [class*='from'], [class*='guest']")
                    row["inb_guest_name"] = el.inner_text().strip() if el else ""
                if "inb_subject" in field_keys:
                    el = item.query_selector("[class*='subject'], [class*='title']")
                    row["inb_subject"] = el.inner_text().strip() if el else ""
                if "inb_message" in field_keys:
                    el = item.query_selector("[class*='preview'], [class*='snippet'], [class*='body']")
                    row["inb_message"] = el.inner_text().strip() if el else ""
                if "inb_date" in field_keys:
                    el = item.query_selector("[class*='date'], [class*='time'], time")
                    row["inb_date"] = el.inner_text().strip() if el else ""
                if "inb_status" in field_keys:
                    # Check for read/unread indicators
                    unread = item.query_selector("[class*='unread'], [class*='new']")
                    row["inb_status"] = "Unread" if unread else "Read"
                if row:
                    rows.append(row)
        except Exception:
            pass

        return rows

    def _extract_boost_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract Boost / Performance section data."""
        rows = []
        row = {}

        try:
            # Try to find metric cards on the boost performance page
            cards = page.query_selector_all(
                "[class*='metric'], [class*='kpi'], [class*='stat'], "
                "[class*='boost'], [class*='performance'], "
                ".bui-card, [class*='card']"
            )
            label_map = {
                "visibility": "boost_visibility_score",
                "preferred": "boost_preferred_status", "partner": "boost_preferred_status",
                "genius": "boost_genius_tier",
                "conversion": "boost_conversion_rate",
                "competitor": "boost_competitor_rank", "rank": "boost_competitor_rank",
                "search": "boost_search_views",
                "property views": "boost_property_views", "page views": "boost_property_views",
                "bookings": "boost_bookings",
                "cancellations": "boost_cancellations",
            }
            for card in cards:
                text = card.inner_text().strip().lower()
                for keyword, fk in label_map.items():
                    if keyword in text and fk in field_keys and not row.get(fk):
                        # Value is often the numeric/status part preceding the label
                        lines = [l.strip() for l in card.inner_text().split("\n") if l.strip()]
                        for line in lines:
                            if keyword not in line.lower():
                                row[fk] = line
                                break
                        if not row.get(fk):
                            row[fk] = card.inner_text().strip()[:200]
        except Exception:
            pass

        if row:
            rows.append(row)
        return rows

    def _extract_analytics_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract Analytics section data."""
        rows = []
        row = {}

        try:
            cards = page.query_selector_all(
                "[class*='metric'], [class*='kpi'], [class*='stat'], "
                "[class*='analytics'], [class*='chart'], "
                ".bui-card, [class*='card'], [class*='widget']"
            )
            label_map = {
                "sales": "anl_sales_revenue", "revenue": "anl_sales_revenue",
                "room nights": "anl_room_nights", "nights sold": "anl_room_nights",
                "adr": "anl_adr", "daily rate": "anl_adr",
                "cancellations": "anl_cancellations", "canceled": "anl_cancellations",
                "page views": "anl_page_views", "views": "anl_page_views",
                "click-through": "anl_click_through", "ctr": "anl_click_through",
                "booking demand": "anl_booking_demand", "demand": "anl_booking_demand",
                "market share": "anl_market_share",
                "competitor pricing": "anl_competitor_pricing", "pricing": "anl_competitor_pricing",
                "booking window": "anl_booking_window", "window": "anl_booking_window",
                "country": "anl_country", "origin": "anl_country",
                "device": "anl_device", "mobile": "anl_device",
                "traveler": "anl_traveler", "guest type": "anl_traveler",
            }
            for card in cards:
                text = card.inner_text().strip().lower()
                for keyword, fk in label_map.items():
                    if keyword in text and fk in field_keys and not row.get(fk):
                        lines = [l.strip() for l in card.inner_text().split("\n") if l.strip()]
                        for line in lines:
                            if keyword not in line.lower():
                                row[fk] = line
                                break
                        if not row.get(fk):
                            row[fk] = card.inner_text().strip()[:200]
        except Exception:
            pass

        if row:
            rows.append(row)
        return rows

    def _extract_promotions_fields(self, page, field_keys: list[str]) -> list[dict]:
        """Extract Promotions / Offers data."""
        rows = []

        # Check if the page has an indicator of "no active promotions"
        # FIX #1: Make detection more specific to avoid false positives
        # - Only check within the main content area, not full body
        # - Require the text to appear in a specific context (empty state container)
        # - Check for positive indicators of promotions existing first
        try:
            # First check if there ARE promotions visible (positive indicator)
            main_content = page.query_selector("main, [class*='content'], [class*='container'], .bui-main-layout__content")
            check_area = main_content if main_content else page
            body_text = check_area.inner_text().lower() if check_area else ""

            # Positive indicators that promotions DO exist
            promo_exists_indicators = [
                "active promotion", "your promotion", "promotion name",
                "discount %", "bookable period", "stay dates", "room nights",
                "average daily rate", "cancelled room nights", "promo-",
                "data-promo", "offer-card", "promotion-card"
            ]
            has_promo_content = any(ind in body_text for ind in promo_exists_indicators)

            # Negative indicators (only trust if no positive content found)
            no_promo_indicators = [
                "don't have any active promotions",
                "no active promotions",
                "no promotions running",
                "you don't have any promotions",
                "create your first promotion",
                "get started with promotions",
            ]

            matched_negative = [ind for ind in no_promo_indicators if ind in body_text]
            if matched_negative and not has_promo_content:
                print(f"DEBUG [promotions]: No promotions indicator found (negative match, no positive content): {matched_negative}")
                return []
            elif matched_negative and has_promo_content:
                print(f"DEBUG [promotions]: Negative indicator found but positive content exists - continuing: {matched_negative}")
        except Exception as e:
            print(f"DEBUG [promotions]: Error in no-promo check: {e}")

        # FIX #2: Improved table extraction - try multiple table locations
        print(f"DEBUG [promotions]: Attempting table extraction...")
        table_rows = self._extract_table_fields(page, field_keys, {
            "name": "promo_Name", "promotion": "promo_Name", "offer": "promo_Name", "title": "promo_Name",
            "discount": "promo_Discount", "%": "promo_Discount", "discount %": "promo_Discount", "amount": "promo_Discount",
            "bookable period": "promo_Bookable_period", "booking period": "promo_Bookable_period", "valid from": "promo_Bookable_period", "from": "promo_Bookable_period",
            "stay dates": "promo_Stay_dates", "stay period": "promo_Stay_dates", "valid to": "promo_Stay_dates", "to": "promo_Stay_dates", "dates": "promo_Stay_dates",
            "bookings": "promo_Bookings", "reservations": "promo_Bookings",
            "room nights": "promo_Room_nights", "nights": "promo_Room_nights", "room night": "promo_Room_nights",
            "average daily rate": "promo_Average_daily_rate", "adr": "promo_Average_daily_rate", "avg daily rate": "promo_Average_daily_rate",
            "revenue": "promo_Revenue", "total revenue": "promo_Revenue",
            "cancelled room nights": "promo_Cancelled_room_nights", "cancelled nights": "promo_Cancelled_room_nights", "cancellations": "promo_Cancelled_room_nights",
            "status": "promo_Status", "state": "promo_Status",
        })
        if table_rows:
            print(f"DEBUG [promotions]: Table extraction SUCCESS - got {len(table_rows)} rows")
            return table_rows
        else:
            print(f"DEBUG [promotions]: Table extraction returned empty, trying card fallback...")

        # FIX #3: Updated CSS selectors for current Booking.com structure
        # Try multiple selector strategies in order of specificity
        selector_strategies = [
            # Strategy 1: Data attributes (most stable)
            "[data-testid*='promo'], [data-testid*='offer'], [data-promo-id], [data-offer-id]",
            # Strategy 2: BUI component classes (Booking.com's design system)
            "[class*='PromotionCard'], [class*='OfferCard'], [class*='promotion-card'], [class*='offer-card']",
            # Strategy 3: Table row patterns
            "tbody tr[class*='promo'], tbody tr[class*='offer'], tr[data-promo], tr[data-offer]",
            # Strategy 4: Generic card/row patterns
            "[class*='promo-row'], [class*='offer-row'], [class*='promo-item'], [class*='offer-item']",
            # Strategy 5: List/group containers
            "[class*='promotion-list'] > *, [class*='offer-list'] > *, [class*='promos-container'] > *",
            # Strategy 6: Original fallback (broad)
            ".promo-card, .offer-card, [data-promo-id], [class*='promo-card'], [class*='offer-card'], [class*='promotion-card'], div[class*='promo-row'], div[class*='offer-row']",
        ]

        items = []
        for i, selector in enumerate(selector_strategies):
            try:
                found = page.query_selector_all(selector)
                if found and len(found) > 0:
                    items = found
                    print(f"DEBUG [promotions]: Selector strategy {i+1} matched {len(items)} elements: {selector}")
                    break
                else:
                    print(f"DEBUG [promotions]: Selector strategy {i+1} returned 0 elements: {selector}")
            except Exception as e:
                print(f"DEBUG [promotions]: Selector strategy {i+1} error: {e}")

        if not items:
            print(f"DEBUG [promotions]: No promotion elements found with any selector strategy")
            return rows

        # Fallback: parse promotion cards with improved line classifier
        for idx, item in enumerate(items):
            try:
                row = {}
                full_text = item.inner_text().strip()
                if not full_text or len(full_text) < 10:
                    print(f"DEBUG [promotions]: Skipping item {idx} - text too short: '{full_text[:50]}'")
                    continue

                print(f"DEBUG [promotions]: Processing item {idx}: '{full_text[:100]}...'")
                lines = [l.strip() for l in full_text.split("\n") if l.strip()]

                # Improved classification with more patterns
                discount = ""
                name = ""
                dates = []
                status = "Active"
                conditions = ""
                bookings = ""
                room_nights = ""
                adr = ""
                revenue = ""
                cancelled_nights = ""

                for line_idx, line in enumerate(lines):
                    l_lower = line.lower()

                    # Discount detection
                    if any(curr in line for curr in ["%", "$", "€", "£", "₹", "rs", "inr"]) or \
                       any(kw in l_lower for kw in ["off", "discount", "save", "% off"]):
                        if not discount or len(line) < len(discount):  # Prefer shorter, cleaner match
                            discount = line

                    # Date detection - broader patterns
                    elif any(kw in l_lower for kw in ["valid", "from", "to", "between", "until", "through", "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "2024", "2025", "2026", "/", "-"]):
                        # Filter out lines that are clearly not date ranges
                        if not any(skip in l_lower for skip in ["status", "bookings", "nights", "revenue", "adr", "rate", "cancel"]):
                            dates.append(line)

                    # Status detection
                    elif any(kw in l_lower for kw in ["status", "state", "ended", "inactive", "expired", "active", "running", "paused", "draft", "scheduled"]):
                        status = line

                    # Conditions/terms
                    elif any(kw in l_lower for kw in ["terms", "condition", "min stay", "minimum stay", "min. stay", "lead time"]):
                        conditions = line

                    # Bookings / metrics
                    elif "booking" in l_lower and not any(skip in l_lower for skip in ["period", "date", "window", "engine"]):
                        bookings = line
                    elif "night" in l_lower and "room" in l_lower:
                        room_nights = line
                    elif "adr" in l_lower or "average daily rate" in l_lower or "avg rate" in l_lower:
                        adr = line
                    elif "revenue" in l_lower and "total" in l_lower:
                        revenue = line
                    elif "cancel" in l_lower and "night" in l_lower:
                        cancelled_nights = line

                    # Name - first meaningful line that isn't a metric
                    elif line_idx == 0 and not any(kw in l_lower for kw in ["%", "$", "€", "£", "status", "valid", "from", "to", "book", "night", "revenue", "adr", "cancel"]):
                        name = line
                    elif line_idx == 1 and not name and not any(kw in l_lower for kw in ["%", "$", "€", "£", "status", "valid", "from", "to"]):
                        name = line

                # Map extracted variables to the requested field keys
                row["promo_Name"] = name or (lines[0] if lines else "Promotion")
                row["promo_Discount"] = discount or "See details"

                if len(dates) >= 2:
                    row["promo_Bookable_period"] = dates[0]
                    row["promo_Stay_dates"] = dates[1]
                elif len(dates) == 1:
                    row["promo_Bookable_period"] = dates[0]
                    row["promo_Stay_dates"] = "Always active"
                else:
                    row["promo_Bookable_period"] = "N/A"
                    row["promo_Stay_dates"] = "Always active"

                row["promo_Bookings"] = bookings
                row["promo_Room_nights"] = room_nights
                row["promo_Average_daily_rate"] = adr
                row["promo_Revenue"] = revenue
                row["promo_Cancelled_room_nights"] = cancelled_nights
                row["promo_Status"] = status

                # Only add if we have at least a name or discount
                if row["promo_Name"] != "Promotion" or discount:
                    rows.append(row)
                    print(f"DEBUG [promotions]: Added row for '{row['promo_Name']}' with discount='{row['promo_Discount']}'")
                else:
                    print(f"DEBUG [promotions]: Skipped item {idx} - insufficient data")

            except Exception as e:
                print(f"DEBUG [promotions]: Error processing item {idx}: {e}")
                continue

        print(f"DEBUG [promotions]: Final result - {len(rows)} rows extracted")
        return rows

    def _fast_regex_extract(self, html: str, url: str, field_keys: list[str], section: str) -> list[dict]:
        """
        Subclasses can override this to implement pure Regex/JSON parsing on the raw HTML
        to completely bypass Playwright DOM rendering for maximum speed.
        Returns a list of data rows if successful, or an empty list to fall back to Playwright DOM parsing.
        """
        return []

    def _extract_single_property_data(self, page, field_keys: list[str], section: str) -> list[dict]:
        # Get labels for the keys
        job = getattr(self, "job", None)
        selected_fields_map = {}
        if job:
            for f in job.selected_fields:
                selected_fields_map[f["key"]] = f["label"]
        
        # Fallback maps
        if not selected_fields_map:
            # Reconstruct labels from our available_fields if job is not populated
            for group_info in self.available_fields:
                for f in group_info["fields"]:
                    selected_fields_map[f["key"]] = f["label"]

        all_rows = []
        worker = getattr(self, "worker", None)
        
        # Separate sub-tab keys from specific data/column keys
        sub_tab_keys = [k for k in field_keys if k in self.SUB_TAB_FIELDS]
        data_keys = [k for k in field_keys if k not in self.SUB_TAB_FIELDS]
        
        # 1. Scrape specific data fields in a single pass (if any)
        if data_keys:
            if worker:
                worker.log_msg.emit(f"  -> Extracting standard data fields: {', '.join(data_keys)}")
            rows = []
            if section == "property":
                row = self._extract_property_fields(page, data_keys)
                if row:
                    rows = [row]
            elif section == "dashboard":
                rows = self._extract_dashboard_fields(page, data_keys)
            elif section == "reviews":
                rows = self._extract_review_fields(page, data_keys)
            elif section == "inbox":
                rows = self._extract_inbox_fields(page, data_keys)
            elif section == "boost":
                rows = self._extract_boost_fields(page, data_keys)
            elif section == "analytics":
                rows = self._extract_analytics_fields(page, data_keys)
            elif section == "promotions":
                rows = self._extract_promotions_fields(page, data_keys)
            else:
                rows = self._generic_fallback(page)
                
            if rows:
                for r in rows:
                    r["sub_tab"] = "General Data"
                all_rows.extend(rows)
                
        # 2. Fetch and scrape sub-tabs sequentially
        for key in sub_tab_keys:
            label = selected_fields_map.get(key, key)
            
            path = self.SUB_TAB_URLS.get(key)
            clicked = False
            
            if path:
                # Use stored hotel_id and ses if available, otherwise get from page URL/DOM/cookies
                hotel_id = getattr(self, "current_hotel_id", None)
                if not hotel_id or hotel_id == "single":
                    hotel_id_match = re.search(r'hotel_id=(\d+)', page.url)
                    if hotel_id_match:
                        hotel_id = hotel_id_match.group(1)
                    else:
                        try:
                            el = page.locator("[data-hotel-id]").first
                            if el.count() > 0:
                                val = el.get_attribute("data-hotel-id")
                                if val and val.isdigit():
                                    hotel_id = val
                        except Exception:
                            pass
                        
                        if not hotel_id or hotel_id == "single":
                            try:
                                link = page.locator("a[href*='hotel_id=']").first
                                if link.count() > 0:
                                    href = link.get_attribute("href")
                                    m = re.search(r'hotel_id=(\d+)', href)
                                    if m:
                                        hotel_id = m.group(1)
                            except Exception:
                                pass
                        
                        if not hotel_id or hotel_id == "single":
                            try:
                                cookies = page.context.cookies()
                                for cookie in cookies:
                                    if cookie['name'] in ['last_hotel_id', 'hotel_id'] and cookie['value'].isdigit():
                                        hotel_id = cookie['value']
                                        break
                            except Exception:
                                pass

                        if not hotel_id or hotel_id == "single":
                            try:
                                html_content = page.content()
                                m = re.search(r'hotel_id=(\d+)', html_content)
                                if m:
                                    hotel_id = m.group(1)
                            except Exception:
                                pass
                if hotel_id and hotel_id != "single":
                    self.current_hotel_id = hotel_id

                ses = getattr(self, "current_ses", None)
                if not ses:
                    ses_match = re.search(r'ses=([a-f0-9]+)', page.url)
                    if ses_match:
                        ses = ses_match.group(1)
                    else:
                        try:
                            link = page.locator("a[href*='ses=']").first
                            if link.count() > 0:
                                href = link.get_attribute("href")
                                m = re.search(r'ses=([a-f0-9]+)', href)
                                if m:
                                    ses = m.group(1)
                        except Exception:
                            pass
                        
                        if not ses:
                            try:
                                html_content = page.content()
                                m = re.search(r'ses=([a-f0-9]+)', html_content)
                                if m:
                                    ses = m.group(1)
                            except Exception:
                                pass
                if ses:
                    self.current_ses = ses
                
                if hotel_id and hotel_id != "single" and ses:
                    sub_tab_url = f"https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/{path}?hotel_id={hotel_id}&ses={ses}&lang=en"
                    if worker:
                        worker.log_msg.emit(f"  -> Background fetching sub-tab: {label}...")
                    
                    html = self._fast_fetch_html(page, sub_tab_url)
                    is_login_redirect = "sign-in" in html or "op_token" in html or len(html) < 200
                    
                    if html and not html.startswith("FETCH_ERROR") and not is_login_redirect:
                        page.set_content(html, wait_until="commit")
                        clicked = True
                    else:
                        if worker:
                            worker.log_msg.emit(f"  -> Fast fetch failed for sub-tab {label}. Falling back to click.")
                        clicked = self._navigate_to_sub_tab(page, label)
                else:
                    clicked = self._navigate_to_sub_tab(page, label)
            else:
                clicked = self._navigate_to_sub_tab(page, label)
                
            if clicked:
                if worker:
                    worker.log_msg.emit(f"  -> Navigated/Loaded sub-tab: {label}")
            else:
                if worker:
                    worker.log_msg.emit(f"  -> (Sub-tab link '{label}' not loaded or already active)")
                
            # Now extract data from this sub-page
            rows = []
            
            if key == "prop_photos":
                photo_urls = page.evaluate("""() => {
                    const imgs = document.querySelectorAll('img[class*="photo"], img[class*="gallery"], [class*="photo"] img, [class*="gallery"] img');
                    const urls = [];
                    for (const img of imgs) {
                        const src = img.getAttribute('src') || img.getAttribute('data-src') || '';
                        if (src && !urls.includes(src)) urls.push(src);
                    }
                    return urls.join('; ');
                }""")
                if photo_urls:
                    rows = [{"photos": photo_urls}]
            elif key == "rev_guest_reviews":
                rows = self._extract_review_fields(page, ["rev_guest_name", "rev_score", "rev_comment", "rev_date", "rev_response"])
            elif key in ["inb_reservation_messages", "inb_booking_messages", "inb_guest_qa"]:
                rows = self._extract_inbox_fields(page, ["inb_guest_name", "inb_subject", "inb_message", "inb_date", "inb_status"])
            elif key == "promo_active":
                rows = self._extract_promotions_fields(page, ["promo_Name", "promo_Discount", "promo_Bookable_period", "promo_Stay_dates", "promo_Bookings", "promo_Room_nights", "promo_Average_daily_rate", "promo_Revenue", "promo_Cancelled_room_nights", "promo_Status"])
            elif key in ["boost_opportunity_center", "boost_genius_program", "boost_preferred_program", "boost_long_stays", "boost_visibility_booster", "boost_smart_flex", "boost_sponsored_listings"]:
                rows = self._extract_boost_fields(page, ["boost_visibility_score", "boost_preferred_status", "boost_genius_tier", "boost_conversion_rate", "boost_competitor_rank", "boost_search_views", "boost_property_views", "boost_bookings", "boost_cancellations"])
            elif key in ["anl_dashboard", "anl_demand", "anl_pace", "anl_sales_stats", "anl_booker_insights", "anl_book_window", "anl_cancellation_char", "anl_comparable", "anl_genius_report", "anl_ranking", "anl_performance"]:
                rows = self._extract_analytics_fields(page, ["anl_sales_revenue", "anl_room_nights", "anl_adr", "anl_cancellations", "anl_page_views", "anl_click_through", "anl_booking_demand", "anl_market_share", "anl_competitor_pricing", "anl_booking_window", "anl_country", "anl_device", "anl_traveler"])
            else:
                # Use generic fallback to parse tables, lists, or text
                rows = self._generic_fallback(page)
                
            # Tag rows with the sub-tab name
            if rows:
                for r in rows:
                    r["sub_tab"] = label
                all_rows.extend(rows)
                
        # If nothing was extracted at all, do a generic fallback on the current page
        if not all_rows:
            all_rows = self._generic_fallback(page)
            for r in all_rows:
                r["sub_tab"] = "General"
                
        return all_rows


    def _fast_fetch_html(self, page, url: str) -> str:
        js_code = """
        async (targetUrl) => {
            try {
                const r = await fetch(targetUrl);
                return await r.text();
            } catch (e) {
                return "FETCH_ERROR: " + e.message;
            }
        }
        """
        return page.evaluate(js_code, url)

    def _extract_property_name_from_page(self, page) -> str:
        """Extract the exact hotel/property name from the current extranet page."""
        try:
            # Settle DOM briefly
            page.wait_for_timeout(1000)
            
            # Words to exclude (since they represent section/page names rather than hotel names)
            exclude_words = {
                "promotions", "rates & availability", "rates", "availability", 
                "reservations", "property", "inbox", "reviews", "finance", 
                "analytics", "dashboard", "home", "group home", "extranet", 
                "calendar", "bulk edit", "deals", "offers", "opportunities"
            }
            
            # Selector list representing elements containing hotel/property name on Booking.com
            selectors = [
                "[class*='hotel-name']",
                "[class*='property-name']",
                "[data-name='hotel-name']",
                "[data-name='property-name']",
                "div[class*='property-name']",
                "[class*='property-selector']",
                "span[class*='property_name']",
                ".property_name",
                ".hotel_name",
                "span[class*='bui-header__title']",
                ".bui-header__title",
                "h1[class*='name']",
                "h1[class*='title']",
                "[class*='header-title']",
            ]
            for sel in selectors:
                try:
                    el = page.query_selector(sel)
                    if el:
                        text = el.inner_text().strip()
                        if text and len(text) > 2:
                            if text.lower() not in exclude_words and not any(w in text.lower() for w in exclude_words):
                                return text
                except Exception:
                    continue
            
            # Fallback: Extract from document.title by splitting on structural separators
            title = page.evaluate("document.title") or ""
            parts = re.split(r'[\-\|·•:—–\(\)]+', title)
            cleaned_parts = []
            for part in parts:
                p = part.strip()
                if not p:
                    continue
                # Clean up Booking keywords
                p_clean = p.replace("Booking.com Extranet", "").replace("Booking.com", "").replace("Extranet", "").strip()
                p_clean_lower = p_clean.lower()
                
                if not p_clean or len(p_clean) < 3:
                    continue
                if p_clean_lower in exclude_words:
                    continue
                if p_clean_lower in ["booking", "extranet", "com", "admin", "hotel", "property"]:
                    continue
                if re.match(r'^\d+$', p_clean):
                    continue
                cleaned_parts.append(p_clean)
                
            if cleaned_parts:
                return cleaned_parts[0]
        except Exception:
            pass
        return ""

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from the currently loaded Booking.com extranet page.
        Routes to section-specific extraction methods based on the field key prefixes.
        Falls back to generic table/body extraction if no fields are matched.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        # Check if we have multiple properties scanned from the Group Homepage
        properties = getattr(self, "_properties", [])
        if properties and len(properties) > 1:
            all_rows = []
            ses_match = re.search(r'ses=([a-f0-9]+)', page.url)
            ses = ses_match.group(1) if ses_match else ""
            
            # Keep a copy of properties and clear self._properties to prevent infinite recursion
            props_list = list(properties)
            self._properties = []
            
            # Update SQLite session total properties count
            session_id = getattr(self, "session_id", None)
            if session_id:
                ScrapeHistoryManager.update_session_counts(session_id, len(props_list))
                
            # Get already completed properties to skip them (resume functionality)
            completed_props = set()
            if session_id:
                completed_props = ScrapeHistoryManager.get_completed_properties(session_id)
            
            base = "https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage"
            section_paths = {
                "dashboard": "home.html",
                "reservations": "search_reservations.html",
                "rates": "rates_availability.html",
                "property": "content_score.html",
                "boost": "opportunities.html",
                "inbox": "messaging/inbox.html",
                "reviews": "reviews.html",
                "financial": "finance_overview.html",
                "analytics": "statistics/index.html",
                "promotions": "promotions/list.html",
            }
            path = section_paths.get(section, "promotions/list.html")
            
            worker = getattr(self, "worker", None)
            
            for idx, prop in enumerate(props_list):
                # Graceful cancellation check
                if worker and worker._stop:
                    if worker:
                        worker.log_msg.emit("Scrape job stop requested. Stopping multi-property loop...")
                    break
                    
                hotel_id = prop["id"]
                hotel_name = prop["name"]
                
                # Check for resume skip
                if hotel_id in completed_props:
                    if worker:
                        worker.log_msg.emit(f"Skipping already completed property {idx+1}/{len(props_list)}: {hotel_name} ({hotel_id})")
                        worker.progress.emit(idx, len(props_list), f"Skipping {hotel_name}...")
                    continue
                
                if worker:
                    worker.log_msg.emit(f"Scraping property {idx+1}/{len(props_list)}: {hotel_name} ({hotel_id})")
                    worker.progress.emit(idx, len(props_list), f"Scraping {hotel_name}...")
                
                url = f"{base}/{path}?hotel_id={hotel_id}&ses={ses}&lang=en"
                try:
                    self.current_hotel_id = hotel_id
                    self.current_ses = ses
                    
                    use_fast_fetch = getattr(getattr(self, 'job', None), 'fast_mode', True)
                    if use_fast_fetch:
                        if worker:
                            worker.log_msg.emit(f"  -> Background fetching main page for {hotel_name} ({hotel_id})...")
                        html = self._fast_fetch_html(page, url)
                        is_login_redirect = "sign-in" in html or "op_token" in html or len(html) < 200
                        if html and not html.startswith("FETCH_ERROR") and not is_login_redirect:
                            # 1) Attempt pure regex extraction on the raw HTML string
                            regex_rows = self._fast_regex_extract(html, url, field_keys, section)
                            if regex_rows:
                                if worker:
                                    worker.log_msg.emit(f"  -> ⚡ Fast regex extraction successful for {hotel_name}!")
                                for r in regex_rows:
                                    if not r.get("hotel_id"):
                                        r["hotel_id"] = hotel_id
                                    if not r.get("hotel_name"):
                                        r["hotel_name"] = hotel_name
                                    all_rows.append(r)
                                    if worker:
                                        worker.live_data.emit(r)
                                self._append_scraped_property_data(hotel_id, hotel_name, regex_rows, "Completed")
                                continue # Bypass all Playwright DOM parsing for this property/section!
                            
                            # 2) Fallback: Inject HTML into DOM for generic selectors
                            page.set_content(html, wait_until="commit")
                        else:
                            if worker:
                                worker.log_msg.emit(f"  -> Fast fetch failed or session expired. Falling back to page.goto.")
                            page.goto(url, timeout=30000, wait_until="domcontentloaded")
                    else:
                        page.goto(url, timeout=30000, wait_until="domcontentloaded")
                    
                    # Dynamic wait for page contents to render (promotions table, cards, or empty state text)
                    try:
                        wait_selectors = {
                            "dashboard": ".dashboard-card, [class*='metric'], [class*='kpi'], table",
                            "reservations": "table, tr, [class*='reservation']",
                            "rates": "table, [class*='calendar'], [class*='rate']",
                            "reviews": ".review-card, [class*='review'], table",
                            "promotions": ".promo-card, .offer-card, table, [class*='promo']",
                            "financial": "table, [class*='finance'], [class*='invoice']",
                            "inbox": "[class*='message'], [class*='inbox'], table",
                            "boost": "[class*='metric'], [class*='boost'], table",
                            "analytics": "[class*='chart'], [class*='metric'], table",
                            "property": "input, textarea, [class*='property'], table"
                        }
                        selector_to_wait = wait_selectors.get(section, "table, .bui-card, [class*='metric'], [class*='promo']")
                        page.wait_for_selector(selector_to_wait, timeout=3000)
                    except Exception:
                        page.wait_for_timeout(1000) # fallback sleep
                    
                    # Extract the exact property name from the actual property page!
                    exact_name = self._extract_property_name_from_page(page)
                    if exact_name and exact_name.lower() not in ["promotions", "rates", "availability", "reservations", "property", "inbox", "reviews", "finance", "analytics", "dashboard", "home", "extranet"]:
                        hotel_name = exact_name
                    
                    # Call single-property extraction
                    rows = self._extract_single_property_data(page, field_keys, section)
                    
                    # Tag with hotel info
                    for r in rows:
                        r["hotel_id"] = hotel_id
                        r["hotel_name"] = hotel_name
                        all_rows.append(r)
                        if worker:
                            worker.live_data.emit(r)
                        
                    # Save incremental data & update SQLite
                    self._append_scraped_property_data(hotel_id, hotel_name, rows, "Completed")
                except Exception as e:
                    err_row = {
                        "_error": f"Failed to scrape hotel {hotel_name} ({hotel_id}): {str(e)}",
                        "hotel_id": hotel_id,
                        "hotel_name": hotel_name
                    }
                    all_rows.append(err_row)
                    self._append_scraped_property_data(hotel_id, hotel_name, [err_row], "Failed")
            
            return all_rows

        # Single-property extraction logic
        # For single property, also try to tag with the exact hotel name and ID if we are on a valid extranet page
        exact_hotel_name = self._extract_property_name_from_page(page)
        
        hotel_id = getattr(self, "current_hotel_id", None)
        if not hotel_id or hotel_id == "single":
            hotel_id_match = re.search(r'hotel_id=(\d+)', page.url)
            if hotel_id_match:
                hotel_id = hotel_id_match.group(1)
            else:
                try:
                    el = page.locator("[data-hotel-id]").first
                    if el.count() > 0:
                        val = el.get_attribute("data-hotel-id")
                        if val and val.isdigit():
                            hotel_id = val
                except Exception:
                    pass
                
                if not hotel_id or hotel_id == "single":
                    try:
                        link = page.locator("a[href*='hotel_id=']").first
                        if link.count() > 0:
                            href = link.get_attribute("href")
                            m = re.search(r'hotel_id=(\d+)', href)
                            if m:
                                hotel_id = m.group(1)
                    except Exception:
                        pass
                
                if not hotel_id or hotel_id == "single":
                    try:
                        cookies = page.context.cookies()
                        for cookie in cookies:
                            if cookie['name'] in ['last_hotel_id', 'hotel_id'] and cookie['value'].isdigit():
                                hotel_id = cookie['value']
                                break
                    except Exception:
                        pass

                if not hotel_id or hotel_id == "single":
                    try:
                        html_content = page.content()
                        m = re.search(r'hotel_id=(\d+)', html_content)
                        if m:
                            hotel_id = m.group(1)
                    except Exception:
                        pass
        if hotel_id and hotel_id != "single":
            self.current_hotel_id = hotel_id
        else:
            hotel_id = "single"
        
        ses = getattr(self, "current_ses", None)
        if not ses:
            ses_match = re.search(r'ses=([a-f0-9]+)', page.url)
            if ses_match:
                ses = ses_match.group(1)
            else:
                try:
                    link = page.locator("a[href*='ses=']").first
                    if link.count() > 0:
                        href = link.get_attribute("href")
                        m = re.search(r'ses=([a-f0-9]+)', href)
                        if m:
                            ses = m.group(1)
                except Exception:
                    pass
                
                if not ses:
                    try:
                        html_content = page.content()
                        m = re.search(r'ses=([a-f0-9]+)', html_content)
                        if m:
                            ses = m.group(1)
                    except Exception:
                        pass
        if ses:
            self.current_ses = ses
        else:
            ses = ""

        # Persist updated parameters to JSON if we found new ones
        if ses and hotel_id and hotel_id != "single":
            try:
                params_path = COOKIES_DIR / f"{self.source_key}_params.json"
                with open(params_path, "w") as f:
                    json.dump({"hotel_id": hotel_id, "ses": ses}, f)
            except Exception:
                pass

        
        if not exact_hotel_name:
            exact_hotel_name = "Single Property"
            
        session_id = getattr(self, "session_id", None)
        if session_id:
            ScrapeHistoryManager.update_session_counts(session_id, 1)
        
        rows = self._extract_single_property_data(page, field_keys, section)
        if exact_hotel_name or hotel_id:
            for r in rows:
                if exact_hotel_name and not r.get("hotel_name"):
                    r["hotel_name"] = exact_hotel_name
                if hotel_id and not r.get("hotel_id"):
                    r["hotel_id"] = hotel_id
                    
        # Save incremental data & update SQLite
        self._append_scraped_property_data(hotel_id, exact_hotel_name, rows, "Completed")
        
        return rows


# ────────────────────────────────────────────────────────────
#  MMT Extranet Source
# ────────────────────────────────────────────────────────────

class MMTExtranetSource(ExtranetSource):
    source_name = "MMT (MakeMyTrip) Extranet"
    login_url = "https://in.goibibo.com/newextranet/dashboard"

    @property
    def cookies_path(self):
        return MMT_EXTRANET_COOKIES

    @property
    def available_fields(self):
        return [
            {
                "group": "Reservations / Bookings",
                "section": "reservations",
                "fields": [
                    {"key": "mmt_booking_id",     "label": "Booking ID"},
                    {"key": "mmt_guest_name",     "label": "Guest Name"},
                    {"key": "mmt_check_in",       "label": "Check-in Date"},
                    {"key": "mmt_check_out",      "label": "Check-out Date"},
                    {"key": "mmt_room_type",      "label": "Room Type"},
                    {"key": "mmt_status",         "label": "Booking Status"},
                    {"key": "mmt_total_amount",   "label": "Total Amount"},
                ]
            },
            {
                "group": "Property Details",
                "section": "property",
                "fields": [
                    {"key": "mmt_prop_name",      "label": "Property Name"},
                    {"key": "mmt_prop_desc",      "label": "Description"},
                    {"key": "mmt_amenities",      "label": "Amenities"},
                    {"key": "mmt_room_inventory", "label": "Room Inventory / Rates"},
                ]
            },
            {
                "group": "Reviews",
                "section": "reviews",
                "fields": [
                    {"key": "mmt_rev_guest",     "label": "Guest Name"},
                    {"key": "mmt_rev_score",     "label": "Rating / Score"},
                    {"key": "mmt_rev_comment",   "label": "Review Comment"},
                    {"key": "mmt_rev_date",      "label": "Review Date"},
                ]
            },
            {
                "group": "Financial / Settlement",
                "section": "financial",
                "fields": [
                    {"key": "mmt_settlement_amt","label": "Settlement Amount"},
                    {"key": "mmt_settlement_date","label": "Settlement Date"},
                    {"key": "mmt_commission",    "label": "Commission"},
                    {"key": "mmt_tds",           "label": "TDS"},
                    {"key": "mmt_invoice_no",    "label": "Invoice Number"},
                ]
            },
            {
                "group": "Promotions / Offers",
                "section": "promotions",
                "fields": [
                    {"key": "promo_mmt_name",       "label": "Offer Name"},
                    {"key": "promo_mmt_type",       "label": "Offer Type"},
                    {"key": "promo_mmt_discount",   "label": "Discount % / Amount"},
                    {"key": "promo_mmt_valid_from", "label": "Valid From"},
                    {"key": "promo_mmt_valid_to",   "label": "Valid To"},
                    {"key": "promo_mmt_conditions", "label": "Terms & Conditions"},
                    {"key": "promo_mmt_status",     "label": "Status"},
                ]
            },
        ]

    @property
    def multi_tab(self) -> bool:
        return True

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _fast_regex_extract(self, html: str, url: str, field_keys: list[str], section: str) -> list[dict]:
        """
        Fast Regex parsing for MMT Extranet.
        Matches exact HTML/JSON structure from the endpoints.
        """
        row = {}
        if section == "reviews":
            # Example: Try to parse JSON from Next.js __INITIAL_STATE__ if present
            m = re.search(r'__INITIAL_STATE__\s*=\s*({.*});', html)
            if m:
                try:
                    data = json.loads(m.group(1))
                    # Attempt to extract review score/count if structure is known
                    # For now, fallback to DOM if not confident
                except Exception:
                    pass
            # Or if it's a direct JSON API response:
            if html.strip().startswith("{"):
                try:
                    data = json.loads(html)
                    if "reviewScore" in data:
                        row["mmt_rev_score"] = str(data["reviewScore"])
                    if "totalReviews" in data:
                        row["mmt_rev_count"] = str(data["totalReviews"])
                    if row:
                        return [row]
                except Exception:
                    pass
        return []

    def navigate_to_section(self, page, section_key: str) -> None:
        base = "https://hotel.makemytrip.com"
        section_map = {
            "reservations": f"{base}/bookings",
            "property":     f"{base}/property",
            "reviews":      f"{base}/reviews",
            "financial":    f"{base}/settlement",
            "promotions":   f"{base}/promotions",
        }
        url = section_map.get(section_key, base)
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which MMT section we're on by field key prefix."""
        for key in field_keys:
            if key.startswith("promo_mmt"):
                return "promotions"
            if key.startswith("mmt_rev"):
                return "reviews"
            if key.startswith("mmt_settlement") or key.startswith("mmt_commission") or key.startswith("mmt_tds") or key.startswith("mmt_invoice"):
                return "financial"
            if key.startswith("mmt_prop") or key.startswith("mmt_amenities") or key.startswith("mmt_room"):
                return "property"
            if key.startswith("mmt"):
                return "reservations"
        return "general"

    def _extract_mmt_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract MMT Property Details fields."""
        row = {}
        if "mmt_prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1", "h2", "input[name*='name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[class*='page-title']", "[class*='headline']",
            ])
            if not val:
                try:
                    val = page.evaluate("document.title").strip()
                    for suf in [" - MakeMyTrip", " | MakeMyTrip", " - MMT"]:
                        if val.endswith(suf):
                            val = val[:-len(suf)].strip()
                            break
                except Exception:
                    pass
            row["mmt_prop_name"] = val or ""
        if "mmt_prop_desc" in field_keys:
            val = self._try_selectors(page, [
                "textarea[name*='description']", "textarea[id*='description']",
                "[class*='description']", "[class*='desc']",
                ".editor-content", "[contenteditable='true']",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Description")
            row["mmt_prop_desc"] = val or ""
        if "mmt_amenities" in field_keys:
            val = self._try_selectors(page, [
                "[class*='amenity']", "[class*='facilities']", ".amenities-list",
            ])
            if not val:
                checked = page.evaluate("""() => {
                    const cbs = document.querySelectorAll('input[type="checkbox"]:checked');
                    return Array.from(cbs).map(cb => {
                        const lbl = cb.closest('label');
                        return lbl ? lbl.textContent.trim() : (cb.parentElement ? cb.parentElement.textContent.trim() : '');
                    }).filter(t => t).join('; ');
                }""")
                val = checked
            row["mmt_amenities"] = val or ""
        if "mmt_room_inventory" in field_keys:
            val = self._try_selectors(page, [
                "[class*='room']", "[class*='inventory']", "[class*='rate-plan']",
                "table",
            ])
            row["mmt_room_inventory"] = val or ""
        return row

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from the currently loaded MMT extranet page.
        Routes to section-specific extraction based on field key prefixes.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        if section == "property":
            row = self._extract_mmt_property_fields(page, field_keys)
            if any(row.values()):
                return [row]

        elif section == "reservations":
            rows = self._extract_table_fields(page, field_keys, {
                "booking": "mmt_booking_id", "id": "mmt_booking_id", "confirmation": "mmt_booking_id",
                "guest": "mmt_guest_name", "name": "mmt_guest_name",
                "check-in": "mmt_check_in", "check in": "mmt_check_in", "arrival": "mmt_check_in",
                "check-out": "mmt_check_out", "check out": "mmt_check_out", "departure": "mmt_check_out",
                "room": "mmt_room_type", "room type": "mmt_room_type",
                "status": "mmt_status",
                "total": "mmt_total_amount", "amount": "mmt_total_amount", "price": "mmt_total_amount",
            })
            if rows:
                return rows

        elif section == "reviews":
            rows = self._extract_table_fields(page, field_keys, {
                "guest": "mmt_rev_guest", "name": "mmt_rev_guest",
                "score": "mmt_rev_score", "rating": "mmt_rev_score",
                "comment": "mmt_rev_comment", "review": "mmt_rev_comment",
                "date": "mmt_rev_date", "time": "mmt_rev_date",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='review'], [class*='feedback'], .review-card, [data-review-id]",
                {
                    "mmt_rev_guest": "[class*='name'], [class*='guest'], h3, h4",
                    "mmt_rev_score": "[class*='score'], [class*='rating'], [class*='grade']",
                    "mmt_rev_comment": "[class*='comment'], [class*='review-text'], [class*='text']",
                    "mmt_rev_date": "[class*='date'], [class*='time'], time",
                }
            )
            if rows:
                return rows

        elif section == "financial":
            rows = self._extract_table_fields(page, field_keys, {
                "amount": "mmt_settlement_amt", "settlement": "mmt_settlement_amt",
                "date": "mmt_settlement_date", "settlement date": "mmt_settlement_date",
                "commission": "mmt_commission",
                "tds": "mmt_tds",
                "invoice": "mmt_invoice_no", "number": "mmt_invoice_no",
            })
            if rows:
                return rows

        elif section == "promotions":
            rows = self._extract_table_fields(page, field_keys, {
                "name": "promo_mmt_name", "offer": "promo_mmt_name", "promotion": "promo_mmt_name",
                "type": "promo_mmt_type",
                "discount": "promo_mmt_discount", "%": "promo_mmt_discount",
                "from": "promo_mmt_valid_from", "valid from": "promo_mmt_valid_from",
                "to": "promo_mmt_valid_to", "valid to": "promo_mmt_valid_to",
                "terms": "promo_mmt_conditions", "conditions": "promo_mmt_conditions",
                "status": "promo_mmt_status",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='promotion'], [class*='offer'], .promo-card, .offer-card",
                {fk: "." for fk in field_keys}  # Full card text as fallback
            )
            if rows:
                for r in rows:
                    for fk in field_keys:
                        if fk not in r:
                            r[fk] = ""
                return rows

        # Generic fallback
        return self._generic_fallback(page)


# ────────────────────────────────────────────────────────────
#  Goibibo Extranet Source  (shares InGo-MMT platform with MMT)
# ────────────────────────────────────────────────────────────

class GoibiboExtranetSource(ExtranetSource):
    """Goibibo partner extranet via the shared Go-MMT / InGo-MMT platform."""
    source_name = "Goibibo Extranet"
    login_url = "https://partners.go-mmt.com/"

    @property
    def cookies_path(self):
        return GOIBIBO_EXTRANET_COOKIES

    @property
    def available_fields(self):
        return [
            {
                "group": "Reservations / Bookings",
                "section": "reservations",
                "fields": [
                    {"key": "goi_booking_id",     "label": "Booking ID"},
                    {"key": "goi_guest_name",     "label": "Guest Name"},
                    {"key": "goi_check_in",       "label": "Check-in Date"},
                    {"key": "goi_check_out",      "label": "Check-out Date"},
                    {"key": "goi_room_type",      "label": "Room Type"},
                    {"key": "goi_status",         "label": "Booking Status"},
                    {"key": "goi_total_amount",   "label": "Total Amount"},
                    {"key": "goi_payment_mode",   "label": "Payment Mode"},
                    {"key": "goi_contact",        "label": "Guest Contact"},
                ]
            },
            {
                "group": "Property Details",
                "section": "property",
                "fields": [
                    {"key": "goi_prop_name",      "label": "Property Name"},
                    {"key": "goi_prop_desc",      "label": "Description"},
                    {"key": "goi_amenities",      "label": "Amenities"},
                    {"key": "goi_room_inventory", "label": "Room Inventory / Rates"},
                    {"key": "goi_prop_images",    "label": "Property Images"},
                ]
            },
            {
                "group": "Reviews",
                "section": "reviews",
                "fields": [
                    {"key": "goi_rev_guest",      "label": "Guest Name"},
                    {"key": "goi_rev_score",      "label": "Rating / Score"},
                    {"key": "goi_rev_comment",    "label": "Review Comment"},
                    {"key": "goi_rev_date",       "label": "Review Date"},
                    {"key": "goi_rev_response",   "label": "Your Response"},
                ]
            },
            {
                "group": "Financial / Settlement",
                "section": "financial",
                "fields": [
                    {"key": "goi_settlement_amt", "label": "Settlement Amount"},
                    {"key": "goi_settlement_date","label": "Settlement Date"},
                    {"key": "goi_commission",     "label": "Commission"},
                    {"key": "goi_tds",            "label": "TDS"},
                    {"key": "goi_invoice_no",     "label": "Invoice Number"},
                    {"key": "goi_payout_status",   "label": "Payout Status"},
                ]
            },
            {
                "group": "Reports / Analytics",
                "section": "reports",
                "fields": [
                    {"key": "goi_rpt_occupancy",  "label": "Occupancy Rate"},
                    {"key": "goi_rpt_revenue",    "label": "Revenue Summary"},
                    {"key": "goi_rpt_avg_rate",   "label": "Average Daily Rate (ADR)"},
                    {"key": "goi_rpt_revpar",     "label": "RevPAR"},
                ]
            },
            {
                "group": "Promotions / Offers",
                "section": "promotions",
                "fields": [
                    {"key": "promo_goi_name",       "label": "Offer Name"},
                    {"key": "promo_goi_type",       "label": "Offer Type"},
                    {"key": "promo_goi_discount",   "label": "Discount % / Amount"},
                    {"key": "promo_goi_valid_from", "label": "Valid From"},
                    {"key": "promo_goi_valid_to",   "label": "Valid To"},
                    {"key": "promo_goi_conditions", "label": "Terms & Conditions"},
                    {"key": "promo_goi_status",     "label": "Status"},
                ]
            },
        ]

    @property
    def multi_tab(self) -> bool:
        return True

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _fast_regex_extract(self, html: str, url: str, field_keys: list[str], section: str) -> list[dict]:
        """
        Fast Regex parsing for Goibibo Extranet.
        """
        row = {}
        if section == "reviews":
            if html.strip().startswith("{"):
                try:
                    data = json.loads(html)
                    if "reviewScore" in data:
                        row["goi_rev_score"] = str(data["reviewScore"])
                    if "totalReviews" in data:
                        row["goi_rev_count"] = str(data["totalReviews"])
                    if row:
                        return [row]
                except Exception:
                    pass
        return []

    def navigate_to_section(self, page, section_key: str) -> None:
        base = "https://partners.go-mmt.com"
        section_map = {
            "reservations": f"{base}/bookings",
            "property":     f"{base}/property",
            "reviews":      f"{base}/reviews",
            "financial":    f"{base}/settlement",
            "reports":      f"{base}/reports/analytics",
            "promotions":   f"{base}/promotions",
        }
        url = section_map.get(section_key, base)
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which Goibibo section we're on by field key prefix."""
        for key in field_keys:
            if key.startswith("promo_goi"):
                return "promotions"
            if key.startswith("goi_rpt"):
                return "reports"
            if key.startswith("goi_rev"):
                return "reviews"
            if key.startswith("goi_settlement") or key.startswith("goi_commission") or key.startswith("goi_tds") or key.startswith("goi_invoice") or key.startswith("goi_payout"):
                return "financial"
            if key.startswith("goi_prop") or key.startswith("goi_amenities") or key.startswith("goi_room"):
                return "property"
            if key.startswith("goi"):
                return "reservations"
        return "general"

    def _extract_goi_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract Goibibo Property Details fields."""
        row = {}
        if "goi_prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1", "h2", "input[name*='name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[class*='page-title']", "[class*='headline']",
            ])
            if not val:
                try:
                    val = page.evaluate("document.title").strip()
                    for suf in [" - Goibibo", " | Goibibo", " - Go-MMT"]:
                        if val.endswith(suf):
                            val = val[:-len(suf)].strip()
                            break
                except Exception:
                    pass
            row["goi_prop_name"] = val or ""
        if "goi_prop_desc" in field_keys:
            val = self._try_selectors(page, [
                "textarea[name*='description']", "textarea[id*='description']",
                "[class*='description']", ".editor-content",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Description")
            row["goi_prop_desc"] = val or ""
        if "goi_amenities" in field_keys:
            val = self._try_selectors(page, [
                "[class*='amenity']", "[class*='facilities']", ".amenities-list",
            ])
            row["goi_amenities"] = val or ""
        if "goi_room_inventory" in field_keys:
            val = self._try_selectors(page, ["[class*='room']", "[class*='inventory']", "table"])
            row["goi_room_inventory"] = val or ""
        if "goi_prop_images" in field_keys:
            urls = page.evaluate("""() => {
                const imgs = document.querySelectorAll('img[src*="property"], img[class*="photo"], img[class*="gallery"]');
                return Array.from(new Set(Array.from(imgs).map(i => i.getAttribute('src') || i.getAttribute('data-src') || '').filter(s => s))).join('; ');
            }""")
            row["goi_prop_images"] = urls or ""
        return row

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from the currently loaded Go-MMT partner portal page.
        Routes to section-specific extraction based on field key prefixes.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        if section == "property":
            row = self._extract_goi_property_fields(page, field_keys)
            if any(row.values()):
                return [row]

        elif section == "reservations":
            rows = self._extract_table_fields(page, field_keys, {
                "booking": "goi_booking_id", "id": "goi_booking_id", "confirmation": "goi_booking_id",
                "guest": "goi_guest_name", "name": "goi_guest_name",
                "check-in": "goi_check_in", "check in": "goi_check_in", "arrival": "goi_check_in",
                "check-out": "goi_check_out", "check out": "goi_check_out", "departure": "goi_check_out",
                "room": "goi_room_type", "room type": "goi_room_type",
                "status": "goi_status",
                "total": "goi_total_amount", "amount": "goi_total_amount", "price": "goi_total_amount",
                "payment": "goi_payment_mode", "mode": "goi_payment_mode",
                "contact": "goi_contact", "phone": "goi_contact", "mobile": "goi_contact",
            })
            if rows:
                return rows

        elif section == "reviews":
            rows = self._extract_table_fields(page, field_keys, {
                "guest": "goi_rev_guest", "name": "goi_rev_guest",
                "score": "goi_rev_score", "rating": "goi_rev_score",
                "comment": "goi_rev_comment", "review": "goi_rev_comment",
                "date": "goi_rev_date", "time": "goi_rev_date",
                "response": "goi_rev_response", "reply": "goi_rev_response",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='review'], [class*='feedback'], .review-card",
                {
                    "goi_rev_guest": "[class*='name'], [class*='guest'], h3, h4",
                    "goi_rev_score": "[class*='score'], [class*='rating']",
                    "goi_rev_comment": "[class*='comment'], [class*='review-text']",
                    "goi_rev_date": "[class*='date'], [class*='time'], time",
                    "goi_rev_response": "[class*='response'], [class*='reply']",
                }
            )
            if rows:
                return rows

        elif section == "financial":
            rows = self._extract_table_fields(page, field_keys, {
                "amount": "goi_settlement_amt", "settlement": "goi_settlement_amt",
                "date": "goi_settlement_date", "settlement date": "goi_settlement_date",
                "commission": "goi_commission",
                "tds": "goi_tds",
                "invoice": "goi_invoice_no", "number": "goi_invoice_no",
                "status": "goi_payout_status", "payout status": "goi_payout_status",
            })
            if rows:
                return rows

        elif section == "reports":
            rows = self._extract_table_fields(page, field_keys, {
                "occupancy": "goi_rpt_occupancy",
                "revenue": "goi_rpt_revenue",
                "average daily rate": "goi_rpt_avg_rate", "adr": "goi_rpt_avg_rate", "avg rate": "goi_rpt_avg_rate",
                "revpar": "goi_rpt_revpar",
            })
            if rows:
                return rows
            rows = self._extract_metric_cards(page, field_keys, {
                "occupancy": "goi_rpt_occupancy",
                "revenue": "goi_rpt_revenue",
                "average daily rate": "goi_rpt_avg_rate", "adr": "goi_rpt_avg_rate",
                "revpar": "goi_rpt_revpar",
            })
            if rows:
                return rows

        elif section == "promotions":
            rows = self._extract_table_fields(page, field_keys, {
                "name": "promo_goi_name", "offer": "promo_goi_name", "promotion": "promo_goi_name",
                "type": "promo_goi_type",
                "discount": "promo_goi_discount", "%": "promo_goi_discount",
                "from": "promo_goi_valid_from", "valid from": "promo_goi_valid_from",
                "to": "promo_goi_valid_to", "valid to": "promo_goi_valid_to",
                "terms": "promo_goi_conditions", "conditions": "promo_goi_conditions",
                "status": "promo_goi_status",
            })
            if rows:
                return rows

        return self._generic_fallback(page)


# ────────────────────────────────────────────────────────────
#  Agoda (YCS) Extranet Source
# ────────────────────────────────────────────────────────────

class AgodaExtranetSource(ExtranetSource):
    """Agoda YCS (Yield Control System) — the extranet for Agoda hotel partners."""
    source_name = "Agoda (YCS) Extranet"
    login_url = "https://ycs.agoda.com/"

    @property
    def cookies_path(self):
        return AGODA_EXTRANET_COOKIES

    @property
    def multi_tab(self) -> bool:
        return True

    @property
    def available_fields(self):
        return [
            {
                "group": "Reservations / Bookings",
                "section": "reservations",
                "fields": [
                    {"key": "agd_booking_id",      "label": "Booking ID"},
                    {"key": "agd_guest_name",      "label": "Guest Name"},
                    {"key": "agd_check_in",        "label": "Check-in Date"},
                    {"key": "agd_check_out",       "label": "Check-out Date"},
                    {"key": "agd_room_type",       "label": "Room Type"},
                    {"key": "agd_status",          "label": "Booking Status"},
                    {"key": "agd_total_amount",    "label": "Total Amount"},
                    {"key": "agd_currency",        "label": "Currency"},
                    {"key": "agd_cancel_policy",   "label": "Cancellation Policy"},
                ]
            },
            {
                "group": "Property Details / Rates",
                "section": "property",
                "fields": [
                    {"key": "agd_prop_name",       "label": "Property Name"},
                    {"key": "agd_room_types",      "label": "Room Types"},
                    {"key": "agd_rate_plans",      "label": "Rate Plans"},
                    {"key": "agd_amenities",       "label": "Amenities"},
                    {"key": "agd_policies",        "label": "Property Policies"},
                    {"key": "agd_photos",          "label": "Photo URLs"},
                ]
            },
            {
                "group": "Reviews",
                "section": "reviews",
                "fields": [
                    {"key": "agd_rev_guest",       "label": "Guest Name"},
                    {"key": "agd_rev_score",       "label": "Review Score"},
                    {"key": "agd_rev_comment",     "label": "Review Comment"},
                    {"key": "agd_rev_response",    "label": "Your Response"},
                    {"key": "agd_rev_date",        "label": "Review Date"},
                    {"key": "agd_rev_language",    "label": "Language"},
                ]
            },
            {
                "group": "Financial / Payouts",
                "section": "financial",
                "fields": [
                    {"key": "agd_payout_amount",   "label": "Payout Amount"},
                    {"key": "agd_payout_date",     "label": "Payout Date"},
                    {"key": "agd_commission",      "label": "Commission"},
                    {"key": "agd_invoice_id",      "label": "Invoice ID"},
                    {"key": "agd_txn_id",          "label": "Transaction ID"},
                    {"key": "agd_payment_status",  "label": "Payment Status"},
                ]
            },
            {
                "group": "Promotions / Offers",
                "section": "promotions",
                "fields": [
                    {"key": "promo_agd_name",       "label": "Promotion Name"},
                    {"key": "promo_agd_type",       "label": "Promotion Type"},
                    {"key": "promo_agd_discount",   "label": "Discount % / Amount"},
                    {"key": "promo_agd_valid_from", "label": "Valid From"},
                    {"key": "promo_agd_valid_to",   "label": "Valid To"},
                    {"key": "promo_agd_conditions", "label": "Terms & Conditions"},
                    {"key": "promo_agd_status",     "label": "Status"},
                ]
            },
        ]

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _fast_regex_extract(self, html: str, url: str, field_keys: list[str], section: str) -> list[dict]:
        """
        Fast Regex parsing for Agoda Extranet (GraphQL / API response).
        """
        row = {}
        if section == "reviews":
            if html.strip().startswith("{"):
                try:
                    data = json.loads(html)
                    # Example generic extraction if GraphQL returns expected keys
                    if "data" in data:
                        d = str(data["data"])
                        ms = re.search(r"'ratingScore':\s*([\d.]+)", d)
                        mc = re.search(r"'totalReviews':\s*(\d+)", d)
                        if "agd_rev_score" in field_keys and ms:
                            row["agd_rev_score"] = ms.group(1)
                        if mc:
                            row["agd_rev_count"] = mc.group(1)
                    if row:
                        return [row]
                except Exception:
                    pass
        return []

    def navigate_to_section(self, page, section_key: str) -> None:
        base = "https://ycs.agoda.com"
        section_map = {
            "reservations": f"{base}/en-us/manage-reservations",
            "property":     f"{base}/en-us/property",
            "reviews":      f"{base}/en-us/reviews",
            "financial":    f"{base}/en-us/financial",
            "promotions":   f"{base}/en-us/promotions",
        }
        url = section_map.get(section_key, base)
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which Agoda YCS section we're on by field key prefix."""
        for key in field_keys:
            if key.startswith("promo_agd"):
                return "promotions"
            if key.startswith("agd_rev"):
                return "reviews"
            if key.startswith("agd_prop") or key.startswith("agd_room") or key.startswith("agd_rate") or key.startswith("agd_amenities") or key.startswith("agd_policies") or key.startswith("agd_photos"):
                return "property"
            if key.startswith("agd_payout") or key.startswith("agd_commission") or key.startswith("agd_invoice") or key.startswith("agd_txn") or key.startswith("agd_payment"):
                return "financial"
            if key.startswith("agd"):
                return "reservations"
        return "general"

    def _extract_agd_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract Agoda YCS Property Details fields."""
        row = {}
        if "agd_prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1", "h2", "input[name*='name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[class*='page-title']", "[class*='headline']",
                ".ycs-header-title", "[class*='ycs-'] h1",
            ])
            if not val:
                try:
                    val = page.evaluate("document.title").strip()
                    for suf in [" - Agoda", " | Agoda", " - YCS", " - Agoda YCS"]:
                        if val.endswith(suf):
                            val = val[:-len(suf)].strip()
                            break
                except Exception:
                    pass
            row["agd_prop_name"] = val or ""
        if "agd_room_types" in field_keys:
            val = self._try_selectors(page, [
                "[class*='room-type']", "[class*='roomtype']", "[class*='room'] h3", "[class*='room'] h4",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Room Types")
            row["agd_room_types"] = val or ""
        if "agd_rate_plans" in field_keys:
            val = self._try_selectors(page, [
                "[class*='rate-plan']", "[class*='rateplan']", "table",
            ])
            row["agd_rate_plans"] = val or ""
        if "agd_amenities" in field_keys:
            val = self._try_selectors(page, ["[class*='amenity']", "[class*='facilities']", ".amenities-list"])
            row["agd_amenities"] = val or ""
        if "agd_policies" in field_keys:
            val = self._try_selectors(page, ["[class*='policy']", "[class*='policy-section']"])
            if not val:
                val = self._extract_value_by_label(page, "Policies")
            row["agd_policies"] = val or ""
        if "agd_photos" in field_keys:
            urls = page.evaluate("""() => {
                const imgs = document.querySelectorAll('img[class*="photo"], img[class*="gallery"], [class*="photo"] img');
                return Array.from(new Set(Array.from(imgs).map(i => i.getAttribute('src') || i.getAttribute('data-src') || '').filter(s => s))).join('; ');
            }""")
            row["agd_photos"] = urls or ""
        return row

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from the currently loaded Agoda YCS page.
        Routes to section-specific extraction based on field key prefixes.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        if section == "property":
            row = self._extract_agd_property_fields(page, field_keys)
            if any(row.values()):
                return [row]

        elif section == "reservations":
            rows = self._extract_table_fields(page, field_keys, {
                "booking": "agd_booking_id", "id": "agd_booking_id", "confirmation": "agd_booking_id",
                "guest": "agd_guest_name", "name": "agd_guest_name",
                "check-in": "agd_check_in", "check in": "agd_check_in", "arrival": "agd_check_in",
                "check-out": "agd_check_out", "check out": "agd_check_out", "departure": "agd_check_out",
                "room": "agd_room_type", "room type": "agd_room_type",
                "status": "agd_status",
                "total": "agd_total_amount", "amount": "agd_total_amount", "price": "agd_total_amount",
                "currency": "agd_currency",
                "cancel": "agd_cancel_policy", "cancellation": "agd_cancel_policy",
            })
            if rows:
                return rows

        elif section == "reviews":
            rows = self._extract_table_fields(page, field_keys, {
                "guest": "agd_rev_guest", "name": "agd_rev_guest",
                "score": "agd_rev_score", "rating": "agd_rev_score",
                "comment": "agd_rev_comment", "review": "agd_rev_comment",
                "response": "agd_rev_response", "reply": "agd_rev_response",
                "date": "agd_rev_date", "time": "agd_rev_date",
                "language": "agd_rev_language", "lang": "agd_rev_language",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='review'], .review-card, .guest-review, [data-review-id]",
                {
                    "agd_rev_guest": "[class*='name'], [class*='guest'], h3, h4",
                    "agd_rev_score": "[class*='score'], [class*='rating']",
                    "agd_rev_comment": "[class*='comment'], [class*='review-text']",
                    "agd_rev_date": "[class*='date'], [class*='time'], time",
                    "agd_rev_response": "[class*='response'], [class*='reply']",
                    "agd_rev_language": "[class*='language'], [class*='lang']",
                }
            )
            if rows:
                return rows

        elif section == "financial":
            rows = self._extract_table_fields(page, field_keys, {
                "amount": "agd_payout_amount", "payout": "agd_payout_amount",
                "date": "agd_payout_date", "payout date": "agd_payout_date",
                "commission": "agd_commission",
                "invoice": "agd_invoice_id", "id": "agd_invoice_id",
                "transaction": "agd_txn_id", "txn": "agd_txn_id",
                "status": "agd_payment_status", "payment": "agd_payment_status",
            })
            if rows:
                return rows

        elif section == "promotions":
            rows = self._extract_table_fields(page, field_keys, {
                "name": "promo_agd_name", "promotion": "promo_agd_name", "offer": "promo_agd_name",
                "type": "promo_agd_type",
                "discount": "promo_agd_discount", "%": "promo_agd_discount",
                "from": "promo_agd_valid_from", "valid from": "promo_agd_valid_from",
                "to": "promo_agd_valid_to", "valid to": "promo_agd_valid_to",
                "terms": "promo_agd_conditions", "conditions": "promo_agd_conditions",
                "status": "promo_agd_status",
            })
            if rows:
                return rows

        return self._generic_fallback(page)


# ────────────────────────────────────────────────────────────
#  Expedia Partner Central Source
# ────────────────────────────────────────────────────────────

class ExpediaExtranetSource(ExtranetSource):
    """Expedia Partner Central (EPC) — the extranet for Expedia Group hotel partners.
    Manages listings across Expedia, Hotels.com, Vrbo, Orbitz, Travelocity, and more.
    """
    source_name = "Expedia Partner Central"
    login_url = "https://expediapartnercentral.com/"

    @property
    def cookies_path(self):
        return EXPEDIA_EXTRANET_COOKIES

    @property
    def multi_tab(self) -> bool:
        return True

    @property
    def available_fields(self):
        return [
            {
                "group": "Reservations / Bookings",
                "section": "reservations",
                "fields": [
                    {"key": "exp_booking_id",      "label": "Booking / Itinerary ID"},
                    {"key": "exp_guest_name",      "label": "Guest Name"},
                    {"key": "exp_check_in",        "label": "Check-in Date"},
                    {"key": "exp_check_out",       "label": "Check-out Date"},
                    {"key": "exp_room_type",       "label": "Room Type"},
                    {"key": "exp_status",          "label": "Booking Status"},
                    {"key": "exp_total_charged",   "label": "Total Charged"},
                    {"key": "exp_source",          "label": "Booking Source (Expedia / Hotels.com / etc.)"},
                    {"key": "exp_cancel_policy",    "label": "Cancellation Policy"},
                ]
            },
            {
                "group": "Property Details / Rates",
                "section": "property",
                "fields": [
                    {"key": "exp_prop_name",       "label": "Property Name"},
                    {"key": "exp_prop_desc",       "label": "Description"},
                    {"key": "exp_room_types",      "label": "Room Types"},
                    {"key": "exp_rate_plans",      "label": "Rate Plans"},
                    {"key": "exp_amenities",       "label": "Amenities"},
                    {"key": "exp_policies",        "label": "Property Policies"},
                    {"key": "exp_photos",          "label": "Photo URLs"},
                ]
            },
            {
                "group": "Reviews & Guest Feedback",
                "section": "reviews",
                "fields": [
                    {"key": "exp_rev_guest",       "label": "Guest Name"},
                    {"key": "exp_rev_score",       "label": "Overall Rating"},
                    {"key": "exp_rev_comment",     "label": "Review Comment"},
                    {"key": "exp_rev_response",    "label": "Your Response"},
                    {"key": "exp_rev_date",        "label": "Review Date"},
                    {"key": "exp_rev_cleanliness", "label": "Cleanliness Score"},
                    {"key": "exp_rev_service",     "label": "Service Score"},
                ]
            },
            {
                "group": "Financial / Payouts",
                "section": "financial",
                "fields": [
                    {"key": "exp_payout_amount",   "label": "Payout Amount"},
                    {"key": "exp_payout_date",     "label": "Payout Date"},
                    {"key": "exp_commission",      "label": "Commission"},
                    {"key": "exp_invoice_id",      "label": "Invoice / Statement ID"},
                    {"key": "exp_transaction_fee", "label": "Transaction Fee"},
                    {"key": "exp_payout_status",    "label": "Payout Status"},
                ]
            },
            {
                "group": "Promotions / Offers",
                "section": "promotions",
                "fields": [
                    {"key": "promo_exp_name",        "label": "Promotion Name"},
                    {"key": "promo_exp_type",        "label": "Promotion Type"},
                    {"key": "promo_exp_discount",    "label": "Discount % / Amount"},
                    {"key": "promo_exp_valid_from",  "label": "Valid From"},
                    {"key": "promo_exp_valid_to",    "label": "Valid To"},
                    {"key": "promo_exp_conditions",  "label": "Terms & Conditions"},
                    {"key": "promo_exp_status",      "label": "Status"},
                ]
            },
            {
                "group": "Competitive Insights",
                "section": "insights",
                "fields": [
                    {"key": "exp_ci_occupancy",      "label": "Occupancy Rate"},
                    {"key": "exp_ci_avg_rate",       "label": "Average Daily Rate (ADR)"},
                    {"key": "exp_ci_revpar",         "label": "RevPAR"},
                    {"key": "exp_ci_market_position", "label": "Market Position"},
                ]
            },
        ]

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def navigate_to_section(self, page, section_key: str) -> None:
        base = "https://expediapartnercentral.com"
        section_map = {
            "reservations": f"{base}/reservations",
            "property":     f"{base}/property",
            "reviews":      f"{base}/reviews",
            "financial":    f"{base}/financial",
            "promotions":   f"{base}/promotions",
            "insights":     f"{base}/insights",
        }
        url = section_map.get(section_key, base)
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which Expedia Partner Central section we're on."""
        for key in field_keys:
            if key.startswith("exp_ci"):
                return "insights"
            if key.startswith("promo_exp"):
                return "promotions"
            if key.startswith("exp_rev"):
                return "reviews"
            if key.startswith("exp_prop"):
                return "property"
            if key.startswith("exp"):
                return "reservations"
        return "general"

    def _extract_exp_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract Expedia Property Details fields."""
        row = {}
        if "exp_prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1", "h2", "input[name*='name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[class*='page-title']", "[class*='headline']",
                ".epc-header-title", "[class*='epc-'] h1",
            ])
            if not val:
                try:
                    val = page.evaluate("document.title").strip()
                    for suf in [" - Expedia Partner Central", " | Expedia", " - Expedia", " - EPC"]:
                        if val.endswith(suf):
                            val = val[:-len(suf)].strip()
                            break
                except Exception:
                    pass
            row["exp_prop_name"] = val or ""
        if "exp_prop_desc" in field_keys:
            val = self._try_selectors(page, [
                "textarea[name*='description']", "textarea[id*='description']",
                "[class*='description']", ".editor-content",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Description")
            row["exp_prop_desc"] = val or ""
        if "exp_room_types" in field_keys:
            val = self._try_selectors(page, [
                "[class*='room-type']", "[class*='roomtype']", "[class*='room'] h3",
            ])
            row["exp_room_types"] = val or ""
        if "exp_rate_plans" in field_keys:
            val = self._try_selectors(page, ["[class*='rate-plan']", "[class*='rateplan']", "table"])
            row["exp_rate_plans"] = val or ""
        if "exp_amenities" in field_keys:
            val = self._try_selectors(page, ["[class*='amenity']", "[class*='facilities']", ".amenities-list"])
            row["exp_amenities"] = val or ""
        if "exp_policies" in field_keys:
            val = self._try_selectors(page, ["[class*='policy']", "[class*='policy-section']"])
            if not val:
                val = self._extract_value_by_label(page, "Policies")
            row["exp_policies"] = val or ""
        if "exp_photos" in field_keys:
            urls = page.evaluate("""() => {
                const imgs = document.querySelectorAll('img[class*="photo"], img[class*="gallery"], [class*="photo"] img');
                return Array.from(new Set(Array.from(imgs).map(i => i.getAttribute('src') || i.getAttribute('data-src') || '').filter(s => s))).join('; ');
            }""")
            row["exp_photos"] = urls or ""
        return row

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from the currently loaded Expedia Partner Central page.
        Routes to section-specific extraction based on field key prefixes.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        if section == "property":
            row = self._extract_exp_property_fields(page, field_keys)
            if any(row.values()):
                return [row]

        elif section == "reservations":
            rows = self._extract_table_fields(page, field_keys, {
                "booking": "exp_booking_id", "id": "exp_booking_id", "itinerary": "exp_booking_id",
                "guest": "exp_guest_name", "name": "exp_guest_name",
                "check-in": "exp_check_in", "check in": "exp_check_in", "arrival": "exp_check_in",
                "check-out": "exp_check_out", "check out": "exp_check_out", "departure": "exp_check_out",
                "room": "exp_room_type", "room type": "exp_room_type",
                "status": "exp_status",
                "charged": "exp_total_charged", "total": "exp_total_charged",
                "source": "exp_source", "brand": "exp_source",
                "cancel": "exp_cancel_policy", "cancellation": "exp_cancel_policy",
            })
            if rows:
                return rows

        elif section == "reviews":
            rows = self._extract_table_fields(page, field_keys, {
                "guest": "exp_rev_guest", "name": "exp_rev_guest",
                "score": "exp_rev_score", "rating": "exp_rev_score", "overall": "exp_rev_score",
                "comment": "exp_rev_comment", "review": "exp_rev_comment",
                "response": "exp_rev_response", "reply": "exp_rev_response",
                "date": "exp_rev_date", "time": "exp_rev_date",
                "cleanliness": "exp_rev_cleanliness", "clean": "exp_rev_cleanliness",
                "service": "exp_rev_service", "staff": "exp_rev_service",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='review'], .review-card, .guest-review, [data-review-id]",
                {
                    "exp_rev_guest": "[class*='name'], [class*='guest'], h3, h4",
                    "exp_rev_score": "[class*='score'], [class*='rating']",
                    "exp_rev_comment": "[class*='comment'], [class*='review-text']",
                    "exp_rev_date": "[class*='date'], [class*='time'], time",
                    "exp_rev_response": "[class*='response'], [class*='reply']",
                    "exp_rev_cleanliness": "[class*='cleanliness'], [class*='clean']",
                    "exp_rev_service": "[class*='service']",
                }
            )
            if rows:
                return rows

        elif section == "financial":
            rows = self._extract_table_fields(page, field_keys, {
                "amount": "exp_payout_amount", "payout": "exp_payout_amount",
                "date": "exp_payout_date", "payout date": "exp_payout_date",
                "commission": "exp_commission",
                "invoice": "exp_invoice_id", "statement": "exp_invoice_id", "id": "exp_invoice_id",
                "fee": "exp_transaction_fee", "transaction fee": "exp_transaction_fee",
                "status": "exp_payout_status", "payout status": "exp_payout_status",
            })
            if rows:
                return rows

        elif section == "promotions":
            rows = self._extract_table_fields(page, field_keys, {
                "name": "promo_exp_name", "promotion": "promo_exp_name", "offer": "promo_exp_name",
                "type": "promo_exp_type",
                "discount": "promo_exp_discount", "%": "promo_exp_discount",
                "from": "promo_exp_valid_from", "valid from": "promo_exp_valid_from",
                "to": "promo_exp_valid_to", "valid to": "promo_exp_valid_to",
                "terms": "promo_exp_conditions", "conditions": "promo_exp_conditions",
                "status": "promo_exp_status",
            })
            if rows:
                return rows

        elif section == "insights":
            rows = self._extract_table_fields(page, field_keys, {
                "occupancy": "exp_ci_occupancy",
                "average daily rate": "exp_ci_avg_rate", "adr": "exp_ci_avg_rate", "avg rate": "exp_ci_avg_rate",
                "revpar": "exp_ci_revpar",
                "market": "exp_ci_market_position", "position": "exp_ci_market_position",
            })
            if rows:
                return rows
            rows = self._extract_metric_cards(page, field_keys, {
                "occupancy": "exp_ci_occupancy",
                "average daily rate": "exp_ci_avg_rate", "adr": "exp_ci_avg_rate",
                "revpar": "exp_ci_revpar",
                "market position": "exp_ci_market_position", "market share": "exp_ci_market_position",
            })
            if rows:
                return rows

        return self._generic_fallback(page)


# ────────────────────────────────────────────────────────────
#  Hotels.com Partner Central Source  (same EPC platform as Expedia)
# ────────────────────────────────────────────────────────────

class HotelsExtranetSource(ExtranetSource):
    """Hotels.com Partner Central — shares the Expedia Group Partner Central platform (EPC).
    Hotels.com is part of Expedia Group, so hotel partners manage all Expedia Group
    brands (Expedia, Hotels.com, Vrbo, Orbitz, Travelocity) through the same portal.
    """
    source_name = "Hotels.com Partner Central"
    login_url = "https://expediapartnercentral.com/"

    @property
    def cookies_path(self):
        return EXPEDIA_EXTRANET_COOKIES  # Same platform as Expedia — shared cookies

    @property
    def multi_tab(self) -> bool:
        return True

    @property
    def available_fields(self):
        return [
            {
                "group": "Reservations / Bookings",
                "section": "reservations",
                "fields": [
                    {"key": "htl_booking_id",     "label": "Booking ID"},
                    {"key": "htl_guest_name",     "label": "Guest Name"},
                    {"key": "htl_check_in",       "label": "Check-in Date"},
                    {"key": "htl_check_out",      "label": "Check-out Date"},
                    {"key": "htl_room_type",      "label": "Room Type"},
                    {"key": "htl_status",         "label": "Booking Status"},
                    {"key": "htl_total_charged",  "label": "Total Charged"},
                    {"key": "htl_cancel_policy",  "label": "Cancellation Policy"},
                ]
            },
            {
                "group": "Property Details / Rates",
                "section": "property",
                "fields": [
                    {"key": "htl_prop_name",      "label": "Property Name"},
                    {"key": "htl_prop_desc",      "label": "Description"},
                    {"key": "htl_room_types",     "label": "Room Types"},
                    {"key": "htl_rate_plans",     "label": "Rate Plans"},
                    {"key": "htl_amenities",      "label": "Amenities"},
                    {"key": "htl_policies",       "label": "Policies"},
                    {"key": "htl_photos",         "label": "Photo URLs"},
                ]
            },
            {
                "group": "Reviews & Guest Feedback",
                "section": "reviews",
                "fields": [
                    {"key": "htl_rev_guest",      "label": "Guest Name"},
                    {"key": "htl_rev_score",      "label": "Overall Rating"},
                    {"key": "htl_rev_comment",    "label": "Review Comment"},
                    {"key": "htl_rev_response",   "label": "Your Response"},
                    {"key": "htl_rev_date",       "label": "Review Date"},
                    {"key": "htl_rev_cleanliness","label": "Cleanliness Score"},
                    {"key": "htl_rev_service",    "label": "Service Score"},
                ]
            },
            {
                "group": "Financial / Payouts",
                "section": "financial",
                "fields": [
                    {"key": "htl_payout_amount",  "label": "Payout Amount"},
                    {"key": "htl_payout_date",    "label": "Payout Date"},
                    {"key": "htl_commission",     "label": "Commission"},
                    {"key": "htl_invoice_id",     "label": "Invoice ID"},
                    {"key": "htl_transaction_fee","label": "Transaction Fee"},
                    {"key": "htl_payout_status",  "label": "Payout Status"},
                ]
            },
            {
                "group": "Promotions / Offers",
                "section": "promotions",
                "fields": [
                    {"key": "promo_htl_name",       "label": "Promotion Name"},
                    {"key": "promo_htl_type",       "label": "Promotion Type"},
                    {"key": "promo_htl_discount",   "label": "Discount % / Amount"},
                    {"key": "promo_htl_valid_from", "label": "Valid From"},
                    {"key": "promo_htl_valid_to",   "label": "Valid To"},
                    {"key": "promo_htl_conditions", "label": "Terms & Conditions"},
                    {"key": "promo_htl_status",     "label": "Status"},
                ]
            },
            {
                "group": "Competitive Insights",
                "section": "insights",
                "fields": [
                    {"key": "htl_ci_occupancy",     "label": "Occupancy Rate"},
                    {"key": "htl_ci_avg_rate",      "label": "Average Daily Rate (ADR)"},
                    {"key": "htl_ci_revpar",        "label": "RevPAR"},
                    {"key": "htl_ci_market_position","label": "Market Position"},
                ]
            },
        ]

    def login(self, page):
        page.goto(self.login_url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def navigate_to_section(self, page, section_key: str) -> None:
        base = "https://expediapartnercentral.com"
        section_map = {
            "reservations": f"{base}/reservations",
            "property":     f"{base}/property",
            "reviews":      f"{base}/reviews",
            "financial":    f"{base}/financial",
            "promotions":   f"{base}/promotions",
            "insights":     f"{base}/insights",
        }
        url = section_map.get(section_key, base)
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(1000)

    def _detect_section(self, field_keys: list[str]) -> str:
        """Determine which Hotels.com Partner Central section we're on."""
        for key in field_keys:
            if key.startswith("htl_ci"):
                return "insights"
            if key.startswith("promo_htl"):
                return "promotions"
            if key.startswith("htl_rev"):
                return "reviews"
            if key.startswith("htl_prop"):
                return "property"
            if key.startswith("htl"):
                return "reservations"
        return "general"

    def _extract_htl_property_fields(self, page, field_keys: list[str]) -> dict:
        """Extract Hotels.com Property Details fields."""
        row = {}
        if "htl_prop_name" in field_keys:
            val = self._try_selectors(page, [
                "h1", "h2", "input[name*='name']", "input[id*='name']",
                "[class*='property-name']", "[class*='hotel-name']",
                "[class*='page-title']", "[class*='headline']",
                ".epc-header-title", "[class*='epc-'] h1",
            ])
            if not val:
                try:
                    val = page.evaluate("document.title").strip()
                    for suf in [" - Hotels.com", " | Hotels.com", " - Expedia", " - EPC"]:
                        if val.endswith(suf):
                            val = val[:-len(suf)].strip()
                            break
                except Exception:
                    pass
            row["htl_prop_name"] = val or ""
        if "htl_prop_desc" in field_keys:
            val = self._try_selectors(page, [
                "textarea[name*='description']", "textarea[id*='description']",
                "[class*='description']", ".editor-content",
            ])
            if not val:
                val = self._extract_value_by_label(page, "Description")
            row["htl_prop_desc"] = val or ""
        if "htl_room_types" in field_keys:
            val = self._try_selectors(page, [
                "[class*='room-type']", "[class*='roomtype']", "[class*='room'] h3",
            ])
            row["htl_room_types"] = val or ""
        if "htl_rate_plans" in field_keys:
            val = self._try_selectors(page, ["[class*='rate-plan']", "[class*='rateplan']", "table"])
            row["htl_rate_plans"] = val or ""
        if "htl_amenities" in field_keys:
            val = self._try_selectors(page, ["[class*='amenity']", "[class*='facilities']", ".amenities-list"])
            row["htl_amenities"] = val or ""
        if "htl_policies" in field_keys:
            val = self._try_selectors(page, ["[class*='policy']", "[class*='policy-section']"])
            if not val:
                val = self._extract_value_by_label(page, "Policies")
            row["htl_policies"] = val or ""
        if "htl_photos" in field_keys:
            urls = page.evaluate("""() => {
                const imgs = document.querySelectorAll('img[class*="photo"], img[class*="gallery"], [class*="photo"] img');
                return Array.from(new Set(Array.from(imgs).map(i => i.getAttribute('src') || i.getAttribute('data-src') || '').filter(s => s))).join('; ');
            }""")
            row["htl_photos"] = urls or ""
        return row

    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """Extract rows from Hotels.com Partner Central (same EPC platform as Expedia).
        Routes to section-specific extraction based on field key prefixes.
        """
        field_keys = [f["key"] for f in selected_fields]
        section = self._detect_section(field_keys)

        if section == "property":
            row = self._extract_htl_property_fields(page, field_keys)
            if any(row.values()):
                return [row]

        elif section == "reservations":
            rows = self._extract_table_fields(page, field_keys, {
                "booking": "htl_booking_id", "id": "htl_booking_id", "itinerary": "htl_booking_id",
                "guest": "htl_guest_name", "name": "htl_guest_name",
                "check-in": "htl_check_in", "check in": "htl_check_in", "arrival": "htl_check_in",
                "check-out": "htl_check_out", "check out": "htl_check_out", "departure": "htl_check_out",
                "room": "htl_room_type", "room type": "htl_room_type",
                "status": "htl_status",
                "charged": "htl_total_charged", "total": "htl_total_charged",
                "cancel": "htl_cancel_policy", "cancellation": "htl_cancel_policy",
            })
            if rows:
                return rows

        elif section == "reviews":
            rows = self._extract_table_fields(page, field_keys, {
                "guest": "htl_rev_guest", "name": "htl_rev_guest",
                "score": "htl_rev_score", "rating": "htl_rev_score", "overall": "htl_rev_score",
                "comment": "htl_rev_comment", "review": "htl_rev_comment",
                "response": "htl_rev_response", "reply": "htl_rev_response",
                "date": "htl_rev_date", "time": "htl_rev_date",
                "cleanliness": "htl_rev_cleanliness", "clean": "htl_rev_cleanliness",
                "service": "htl_rev_service", "staff": "htl_rev_service",
            })
            if rows:
                return rows
            rows = self._extract_card_fields(page, field_keys,
                "[class*='review'], .review-card, .guest-review, [data-review-id]",
                {
                    "htl_rev_guest": "[class*='name'], [class*='guest'], h3, h4",
                    "htl_rev_score": "[class*='score'], [class*='rating']",
                    "htl_rev_comment": "[class*='comment'], [class*='review-text']",
                    "htl_rev_date": "[class*='date'], [class*='time'], time",
                    "htl_rev_response": "[class*='response'], [class*='reply']",
                    "htl_rev_cleanliness": "[class*='cleanliness'], [class*='clean']",
                    "htl_rev_service": "[class*='service']",
                }
            )
            if rows:
                return rows

        elif section == "financial":
            rows = self._extract_table_fields(page, field_keys, {
                "amount": "htl_payout_amount", "payout": "htl_payout_amount",
                "date": "htl_payout_date", "payout date": "htl_payout_date",
                "commission": "htl_commission",
                "invoice": "htl_invoice_id", "id": "htl_invoice_id",
                "fee": "htl_transaction_fee", "transaction fee": "htl_transaction_fee",
                "status": "htl_payout_status", "payout status": "htl_payout_status",
            })
            if rows:
                return rows

        elif section == "promotions":
            rows = self._extract_table_fields(page, field_keys, {
                "name": "promo_htl_name", "promotion": "promo_htl_name", "offer": "promo_htl_name",
                "type": "promo_htl_type",
                "discount": "promo_htl_discount", "%": "promo_htl_discount",
                "from": "promo_htl_valid_from", "valid from": "promo_htl_valid_from",
                "to": "promo_htl_valid_to", "valid to": "promo_htl_valid_to",
                "terms": "promo_htl_conditions", "conditions": "promo_htl_conditions",
                "status": "promo_htl_status",
            })
            if rows:
                return rows

        elif section == "insights":
            rows = self._extract_table_fields(page, field_keys, {
                "occupancy": "htl_ci_occupancy",
                "average daily rate": "htl_ci_avg_rate", "adr": "htl_ci_avg_rate", "avg rate": "htl_ci_avg_rate",
                "revpar": "htl_ci_revpar",
                "market": "htl_ci_market_position", "position": "htl_ci_market_position",
            })
            if rows:
                return rows
            rows = self._extract_metric_cards(page, field_keys, {
                "occupancy": "htl_ci_occupancy",
                "average daily rate": "htl_ci_avg_rate", "adr": "htl_ci_avg_rate",
                "revpar": "htl_ci_revpar",
                "market position": "htl_ci_market_position", "market share": "htl_ci_market_position",
            })
            if rows:
                return rows

        return self._generic_fallback(page)


# ───────────────────────────────────────────────────────────
#  Async HTTP Scraper Source — Integration with Universal Scraper
# ───────────────────────────────────────────────────────────

class AsyncHTTPSource(ExtranetSource):
    """
    High-performance async HTTP scraper for any URL-based data source.
    Integrates with the universal scraper framework.
    
    Use case: Scrape any website or API endpoint with:
      - Concurrency control (anti-blocking)
      - Browser-like headers
      - Structured data extraction
      - CSV export
    """
    
    source_key = "async_http"
    
    @property
    def source_name(self) -> str:
        return "Async HTTP Scraper"
    
    @property
    def login_url(self) -> str:
        return "about:blank"
    
    @property
    def available_fields(self) -> list[dict]:
        """Fields for async HTTP scraper."""
        return [
            {
                "group": "URL & Content",
                "section": "content",
                "fields": [
                    {"key": "async_url", "label": "Source URL"},
                    {"key": "async_title", "label": "Page Title"},
                    {"key": "async_status", "label": "Scrape Status"},
                    {"key": "async_timestamp", "label": "Timestamp"},
                ]
            },
            {
                "group": "Structured Data",
                "section": "structured",
                "fields": [
                    {"key": "async_schema_type", "label": "Schema Type"},
                    {"key": "async_name", "label": "Name (JSON-LD)"},
                    {"key": "async_rating", "label": "Rating (JSON-LD)"},
                    {"key": "async_description", "label": "Description"},
                ]
            },
            {
                "group": "API Discovery",
                "section": "api",
                "fields": [
                    {"key": "async_api_endpoints", "label": "Discovered API Endpoints"},
                ]
            },
        ]
    
    @property
    def cookies_path(self) -> Path:
        """No cookies needed for async HTTP scraper."""
        return COOKIES_DIR / "async_http_session.pkl"
    
    def login(self, page) -> None:
        """No login needed for async HTTP scraper."""
        pass
    
    def navigate_to_section(self, page, section_key: str) -> None:
        """No navigation needed — URLs are provided directly."""
        pass
    
    def extract_data(self, page, selected_fields: list[dict]) -> list[dict]:
        """
        Extract data using the async HTTP scraper.
        Note: page argument is ignored (async scraper doesn't use Playwright).
        
        For actual usage, the universal scraper UI should prompt for URLs,
        then invoke async_scraper_core.RatingScraper directly.
        """
        return []


# ───────────────────────────────────────────────────────────
#  Source registry — add new sources here
# ───────────────────────────────────────────────────────────

EXTRANET_SOURCES: dict[str, ExtranetSource] = {
    "booking_extranet": BookingExtranetSource(),
    "mmt_extranet":     MMTExtranetSource(),
    "goibibo_extranet": GoibiboExtranetSource(),
    "agoda_extranet":   AgodaExtranetSource(),
    "expedia_extranet": ExpediaExtranetSource(),
    "hotels_extranet":  HotelsExtranetSource(),
    "async_http":       AsyncHTTPSource(),
}

for key, src in EXTRANET_SOURCES.items():
    src.source_key = key



# ────────────────────────────────────────────────────────────
#  Config system
# ────────────────────────────────────────────────────────────

class ScrapeJob:
    """A single scrape job configuration — serializable to/from JSON."""

    def __init__(self, source_key: str = "", selected_fields: list[dict] = None,
                 label: str = "", output_path: str = "", headless: bool = False,
                 fast_mode: bool = False, concurrency: int = 1):
        self.source_key = source_key
        self.selected_fields = selected_fields or []
        self.label = label or f"Scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.output_path = output_path or ""
        self.headless = headless
        self.fast_mode = fast_mode
        self.concurrency = concurrency

    def to_dict(self) -> dict:
        return {
            "source": self.source_key,
            "label": self.label,
            "output_path": self.output_path,
            "headless": self.headless,
            "fast_mode": self.fast_mode,
            "concurrency": self.concurrency,
            "fields": [
                {"key": f["key"], "label": f.get("label", f["key"])}
                for f in self.selected_fields
            ],
        }

    @classmethod
    def from_dict(cls, data: dict):
        return cls(
            source_key=data.get("source", ""),
            selected_fields=data.get("fields", []),
            label=data.get("label", ""),
            output_path=data.get("output_path", ""),
            headless=data.get("headless", False),
            fast_mode=data.get("fast_mode", False),
            concurrency=data.get("concurrency", 1),
        )

    @classmethod
    def from_file(cls, path: str):
        with open(path, encoding="utf-8") as f:
            return cls.from_dict(json.load(f))

    def to_file(self, path: str):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.to_dict(), f, indent=2)

    def display_summary(self) -> str:
        source = EXTRANET_SOURCES.get(self.source_key)
        name = source.source_name if source else self.source_key
        field_labels = [f.get("label", f["key"]) for f in self.selected_fields]
        return f"[{name}]  {len(field_labels)} fields: {', '.join(field_labels[:5])}{'...' if len(field_labels) > 5 else ''}"


# ────────────────────────────────────────────────────────────
#  SQLite-based Scrape History and Progress Auto-Save Manager
# ────────────────────────────────────────────────────────────

class ScrapeHistoryManager:
    DB_PATH = COOKIES_DIR / "scrape_history.db"

    @classmethod
    def init_db(cls):
        """Initialize the SQLite database schema."""
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS scrape_sessions (
                id TEXT PRIMARY KEY,
                timestamp TEXT,
                platform TEXT,
                source_key TEXT,
                fields TEXT,
                output_path TEXT,
                status TEXT,
                total_properties INTEGER,
                processed_properties INTEGER,
                total_rows INTEGER
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS scraped_properties (
                session_id TEXT,
                hotel_id TEXT,
                hotel_name TEXT,
                status TEXT,
                rows_count INTEGER,
                timestamp TEXT,
                PRIMARY KEY (session_id, hotel_id)
            )
        """)
        conn.commit()
        conn.close()

    @classmethod
    def create_session(cls, session_id: str, platform: str, source_key: str, fields: list[dict], output_path: str):
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        fields_json = json.dumps(fields)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT OR REPLACE INTO scrape_sessions 
            (id, timestamp, platform, source_key, fields, output_path, status, total_properties, processed_properties, total_rows)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (session_id, timestamp, platform, source_key, fields_json, output_path, "Running", 0, 0, 0))
        conn.commit()
        conn.close()

    @classmethod
    def update_session_counts(cls, session_id: str, total_properties: int):
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE scrape_sessions 
            SET total_properties = ?
            WHERE id = ?
        """, (total_properties, session_id))
        conn.commit()
        conn.close()

    @classmethod
    def add_scraped_property(cls, session_id: str, hotel_id: str, hotel_name: str, status: str, rows_count: int):
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT OR REPLACE INTO scraped_properties 
            (session_id, hotel_id, hotel_name, status, rows_count, timestamp)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (session_id, hotel_id, hotel_name, status, rows_count, timestamp))
        
        # Update processed count and total rows in session
        cursor.execute("""
            SELECT COUNT(*), SUM(rows_count) FROM scraped_properties
            WHERE session_id = ? AND (status = 'Completed' OR rows_count > 0)
        """, (session_id,))
        processed, total_rows = cursor.fetchone()
        
        cursor.execute("""
            UPDATE scrape_sessions 
            SET processed_properties = ?, total_rows = ?
            WHERE id = ?
        """, (processed or 0, total_rows or 0, session_id))
        
        conn.commit()
        conn.close()

    @classmethod
    def complete_session(cls, session_id: str, status: str = "Completed"):
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE scrape_sessions 
            SET status = ?
            WHERE id = ?
        """, (status, session_id))
        conn.commit()
        conn.close()

    @classmethod
    def get_completed_properties(cls, session_id: str) -> set[str]:
        """Get set of hotel_ids already successfully scraped in this session."""
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT hotel_id FROM scraped_properties
            WHERE session_id = ? AND status = 'Completed'
        """, (session_id,))
        rows = cursor.fetchall()
        conn.close()
        return {r[0] for r in rows}

    @classmethod
    def get_session(cls, session_id: str) -> dict:
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM scrape_sessions WHERE id = ?", (session_id,))
        row = cursor.fetchone()
        result = dict(row) if row else {}
        conn.close()
        return result

    @classmethod
    def get_history(cls) -> list[dict]:
        """Fetch all sessions ordered by timestamp descending."""
        cls.init_db()
        conn = sqlite3.connect(str(cls.DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM scrape_sessions ORDER BY timestamp DESC")
        rows = cursor.fetchall()
        result = [dict(r) for r in rows]
        conn.close()
        return result


# ────────────────────────────────────────────────────────────
#  Scrape engine (runs a job in a background thread)
# ────────────────────────────────────────────────────────────

class ExtranetScrapeWorker(QThread):
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(str, int)
    log_msg = pyqtSignal(str)
    login_required = pyqtSignal(str)
    live_data = pyqtSignal(dict)

    def __init__(self, job: ScrapeJob, session_id: str = None):
        super().__init__()
        self.job = job
        self._stop = False
        self.login_event = threading.Event()
        self.session_id = session_id or f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.is_resume = session_id is not None

    def stop(self):
        self._stop = True

    def run(self):
        source = EXTRANET_SOURCES.get(self.job.source_key)
        if not source:
            self.log_msg.emit(f"Unknown source: {self.job.source_key}")
            self.finished.emit("", 0)
            return

        source_name = source.source_name
        
        today_str = datetime.now().strftime("%Y-%m-%d")
        last_year_str = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
        default_label = f"ACTIVE_promotions_report__from_{last_year_str}_to_{today_str}"
        
        output_path = self.job.output_path or str(
            Path.home() / "Downloads" / f"{self.job.label or default_label}.xlsx"
        )
        
        # Set attributes on source singleton immediately (early initialization)
        source.output_path = output_path
        source.session_id = self.session_id
        source.job = self.job
        source.worker = self

        # Initialize or update session in SQLite history
        selected_field_keys = self.job.selected_fields
        if not self.is_resume:
            ScrapeHistoryManager.create_session(
                self.session_id, source_name, self.job.source_key, selected_field_keys, output_path
            )
        else:
            ScrapeHistoryManager.complete_session(self.session_id, "Running")

        # Prepare Excel file and headers in real-time mode
        field_keys = [clean_row_keys({f["key"]: ""}).popitem()[0] for f in self.job.selected_fields]
        ordered_keys = []
        for k in field_keys:
            if k not in ordered_keys:
                ordered_keys.append(k)
        for k in ["hotel_id", "hotel_name", "sub_tab", "_source", "_error"]:
            if k not in ordered_keys:
                ordered_keys.append(k)
        
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        file_exists = os.path.exists(output_path) and os.path.getsize(output_path) > 0
        if not file_exists:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(ordered_keys)
                wb.save(output_path)
            except Exception as e:
                self.log_msg.emit(f"Warning: could not initialize Excel file: {e}")

        self.log_msg.emit(f"Starting scrape from {source_name}")
        self.progress.emit(0, 1, "Launching browser...")

        try:
            pw = sync_playwright().start()
            # Use launch_persistent_context to safely maintain logins/profile persistent state
            args = [
                f"--remote-debugging-port={EXTRANET_DEBUG_PORT}",
                "--no-first-run",
                "--window-size=1280,900",
            ]
            if self.job.headless:
                args.append("--headless=new")
                
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(COOKIES_DIR / 'chrome_extranet'),
                headless=False,
                channel="chrome",
                args=args
            )

            if getattr(self.job, 'fast_mode', False):
                def block_resources(route):
                    if route.request.resource_type in ["image", "stylesheet", "font", "media"]:
                        route.abort()
                    else:
                        route.continue_()
                context.route("**/*", block_resources)

            page = context.pages[0] if context.pages else context.new_page()
            
            # Singleton attributes already early-initialized at start of run()
            pass

            # ── Login (if needed) ──────────────────────────
            cookies_path = source.cookies_path
            if cookies_path and cookies_path.exists():
                with open(cookies_path, "rb") as f:
                    cookies = pickle.load(f)
                context.add_cookies(cookies)
                self.log_msg.emit("Session cookies loaded.")
                # Navigate to the source domain to establish session context.
                # This does two things:
                #   1. Verifies the cookies are still valid (will redirect to dashboard if so)
                #   2. Populates page.url with session params (ses, hotel_id, etc.)
                #   3. Essential for Booking.com — navigate_to_section needs ses params from URL
                page.goto(source.login_url, timeout=30000, wait_until="domcontentloaded")
                
                # Let's wait up to 8 seconds for redirects to settle and evaluate if session is active
                is_session_active = False
                for i in range(8):
                    page.wait_for_timeout(1000)
                    current_url = page.url.lower()
                    
                    body_text = ""
                    try:
                        body_text = page.inner_text("body")[:2000].lower()
                    except Exception:
                        pass
                    
                    has_password_input = False
                    try:
                        has_password_input = page.query_selector("input[type='password']") is not None
                    except Exception:
                        pass
                        
                    login_indicators = [
                        "sign in to manage",      # Booking.com
                        "sign in with",           # Generic
                        "log in to your",         # Generic
                        "enter your password",    # Login form
                        "forgot password",        # Login form
                        "create your account",    # Booking.com signup
                        "login-form",             # Form class
                        "username",               # Input name
                        "password",               # Input name
                    ]
                    has_login_text = any(ind in body_text for ind in login_indicators)
                    
                    # Detect if we landed on an error page (e.g. "sorry, this page does not exist")
                    is_err, _ = source._is_error_page(page)
                    
                    if "admin.booking.com" in current_url:
                        has_session_params = "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url
                        has_dashboard_path = any(x in current_url for x in ("/hotel/", "/extranet/", "/dashboard/"))
                        
                        if has_session_params and has_dashboard_path and not has_password_input and not has_login_text and not is_err:
                            is_session_active = True
                            break
                        elif has_password_input or has_login_text or "login" in current_url or is_err:
                            is_session_active = False
                            break
                    else:
                        # Non-booking sources logic
                        if not has_password_input and not has_login_text and not is_err and any(x in current_url for x in ("dashboard", "home", "extranet")):
                            is_session_active = True
                            break
                        elif has_password_input or has_login_text or "login" in current_url or is_err:
                            is_session_active = False
                            break
                
                if is_session_active:
                    self.log_msg.emit("Session active.")
                else:
                    self.log_msg.emit("Session expired — re-login required.")
                    self.login_required.emit(f"{source.source_name} session expired, please log in again.")
                    self.log_msg.emit("Chrome opened — log in and confirm in the app.")
                    self.login_event.wait()
                    
                    # Wait for dashboard/home URL with session parameters to appear to ensure login succeeded
                    self.log_msg.emit("Waiting for dashboard to load...")
                    has_params = False
                    for _ in range(15):
                        current_url = page.url.lower()
                        if "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url:
                            has_params = True
                            break
                        page.wait_for_timeout(1000)
                    
                    if not has_params:
                        self.log_msg.emit("Warning: Dashboard URL with session parameters not detected yet.")
                        
                    cookies = context.cookies()
                    with open(cookies_path, "wb") as f:
                        pickle.dump(cookies, f)
                    
                    # Extract and save session params
                    try:
                        url = page.url
                        hotel_id_match = re.search(r'hotel_id=(\d+)', url)
                        ses_match = re.search(r'ses=([a-f0-9]+)', url)
                        params = {}
                        if hotel_id_match:
                            params["hotel_id"] = hotel_id_match.group(1)
                        if ses_match:
                            params["ses"] = ses_match.group(1)
                        if params:
                            params_path = COOKIES_DIR / f"{source.source_key}_params.json"
                            with open(params_path, "w") as f:
                                json.dump(params, f)
                    except Exception:
                        pass
                    
                    self.log_msg.emit("Session refreshed.")
            else:
                self.log_msg.emit(f"Navigating to {source.source_name} login page...")
                page.goto(source.login_url, timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)
                self.login_required.emit(f"{source.source_name} login required.")
                self.log_msg.emit("Chrome opened — log in and confirm in the app.")
                self.login_event.wait()
                
                # Wait for dashboard/home URL with session parameters to appear to ensure login succeeded
                self.log_msg.emit("Waiting for dashboard to load...")
                has_params = False
                for _ in range(15):
                    current_url = page.url.lower()
                    if "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url:
                        has_params = True
                        break
                    page.wait_for_timeout(1000)
                
                cookies = context.cookies()
                if cookies_path:
                    with open(cookies_path, "wb") as f:
                        pickle.dump(cookies, f)
                    self.log_msg.emit("Session saved.")

            # ── Group fields by section ────────────────────
            sections = {}
            for field in self.job.selected_fields:
                key = field["key"]
                parts = key.split("_")
                # Try compound prefix first (e.g. "exp_ci" -> insights),
                # then fall back to single prefix (e.g. "exp" -> reservations)
                compound_key = "_".join(parts[:2]) if len(parts) >= 3 else None
                simple_key = parts[0] if len(parts) >= 2 else "general"
                # Map short prefixes to actual section keys
                section_map = {
                    "res": "reservations", "prop": "property",
                    "rev": "reviews", "fin": "financial",
                    "promo": "promotions",
                    # Source-specific section prefixes (compound 2-part keys)
                    "exp_ci": "insights",
                    "htl_ci": "insights",
                    "goi_rpt": "reports",
                    "mmt_rev": "reviews",          # MMT Reviews
                    "mmt_settlement": "financial",  # MMT Financial
                    "mmt_prop": "property",         # MMT Property
                    "mmt_amenities": "property",    # MMT Amenities
                    "mmt_room_inventory": "property", # MMT Room Inventory
                    "mmt_content": "property",      # MMT Content Score
                    "mmt_missing": "property",      # MMT Missing Checklist
                    "goi_rev": "reviews",           # Goibibo Reviews
                    "goi_settlement": "financial",  # Goibibo Financial
                    "goi_prop": "property",         # Goibibo Property
                    "goi_amenities": "property",    # Goibibo Amenities
                    "goi_room_inventory": "property", # Goibibo Room Inventory
                    "goi_content": "property",      # Goibibo Content Score
                    "goi_missing": "property",      # Goibibo Missing Checklist
                    "agd_rev": "reviews",           # Agoda Reviews
                    "agd_prop": "property",         # Agoda Property
                    "exp_rev": "reviews",           # Expedia Reviews
                    "exp_prop": "property",         # Expedia Property
                    "htl_rev": "reviews",           # Hotels.com Reviews
                    "htl_prop": "property",         # Hotels.com Property
                    # Booking.com section prefixes (single-word keys)
                    "dash": "dashboard",
                    "rate": "rates",
                    "boost": "boost",
                    "inb": "inbox",
                    "anl": "analytics",
                    # Fallback simple prefixes -> "reservations"
                    "mmt": "reservations",
                    "goi": "reservations",
                    "agd": "reservations",
                    "exp": "reservations",
                    "htl": "reservations",
                }
                # Check compound prefix first
                if compound_key and compound_key in section_map:
                    section_key = section_map[compound_key]
                else:
                    section_key = section_map.get(simple_key, simple_key)
                if section_key not in sections:
                    sections[section_key] = []
                sections[section_key].append(field)

            if not sections:
                sections["general"] = self.job.selected_fields

            # ── Scrape each section ────────────────────────
            all_rows = []
            section_count = len(sections)
            for idx, (section_key, fields) in enumerate(sections.items()):
                if self._stop:
                    break
                self.log_msg.emit(f"Navigating to section: {section_key} ({idx+1}/{section_count})")
                self.progress.emit(idx, section_count, f"Scraping {section_key}...")

                if source.multi_tab:
                    # Open a new tab for each section (MMT/Goibibo SPAs don't
                    # handle cross-section navigation well on a single page)
                    tab = context.new_page()
                    try:
                        source.navigate_to_section(tab, section_key)
                        page.wait_for_timeout(1000)
                        rows = source.extract_data(tab, fields)
                    finally:
                        tab.close()
                else:
                    source.navigate_to_section(page, section_key)
                    page.wait_for_timeout(1000)  # let the page render
                    rows = source.extract_data(page, fields)

                self.log_msg.emit(f"  -> Got {len(rows)} rows from '{section_key}'")
                all_rows.extend(rows)

            # Extract property name before closing the browser context to prevent Playwright target closed exceptions.
            hotel_name = "Active Property"
            try:
                if page and not page.is_closed():
                    hotel_name = source._extract_property_name_from_page(page) or "Active Property"
            except Exception:
                pass

            context.close()
            pw.stop()

            # ── Double-Safety Excel Write & Consolidation ───────────────────────────
            if self._stop:
                ScrapeHistoryManager.complete_session(self.session_id, "Interrupted")
                self.log_msg.emit("\nScrape job stopped/interrupted by user.")
                self.finished.emit(output_path, 0)
            elif all_rows:
                # Read any existing rows from Excel first (for resume safety)
                existing_rows = []
                if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(output_path)
                        ws = wb.active
                        headers = [cell.value for cell in ws[1]]
                        for row_cells in ws.iter_rows(min_row=2, values_only=True):
                            existing_rows.append(dict(zip(headers, row_cells)))
                    except Exception as e:
                        self.log_msg.emit(f"Warning: could not read existing Excel for consolidation: {e}")
                
                # Clean row keys of the newly scraped rows
                cleaned_all_rows = [clean_row_keys(r) for r in all_rows]
                
                # Combine: remove old rows for hotels we just scraped to prevent duplicates
                scraped_hotel_ids = {r.get("hotel_id") for r in cleaned_all_rows if r.get("hotel_id")}
                consolidated = [r for r in existing_rows if r.get("hotel_id") not in scraped_hotel_ids]
                consolidated.extend(cleaned_all_rows)
                
                # Rewrite consolidated data
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(ordered_keys)
                    for r in consolidated:
                        row_data = [r.get(k, "") for k in ordered_keys]
                        ws.append(row_data)
                    wb.save(output_path)
                except Exception as e:
                    self.log_msg.emit(f"Error writing final Excel file: {e}")

                # Log single property scrape status to SQLite
                try:
                    hotel_id = "single"
                    ScrapeHistoryManager.add_scraped_property(self.session_id, hotel_id, hotel_name, "Completed", len(cleaned_all_rows))
                except Exception as e:
                    self.log_msg.emit(f"Warning: could not write scrape history: {e}")

                ScrapeHistoryManager.complete_session(self.session_id, "Completed")
                
                # Get the actual total row count for the session from SQLite for accuracy
                session_stats = ScrapeHistoryManager.get_session(self.session_id)
                final_row_count = session_stats.get("total_rows", len(consolidated)) if session_stats else len(consolidated)
                
                self.log_msg.emit(f"\nDone! Scrape finished successfully.")
                self.log_msg.emit(f"  Output saved to: {output_path}")
                self.progress.emit(section_count, section_count, "Complete!")
                self.finished.emit(output_path, final_row_count)
            else:
                # If no rows extracted in this run, check if there's data in the database
                session_stats = ScrapeHistoryManager.get_session(self.session_id)
                total_rows = session_stats.get("total_rows", 0) if session_stats else 0
                if total_rows > 0:
                    ScrapeHistoryManager.complete_session(self.session_id, "Completed")
                    self.log_msg.emit(f"\nDone! Scrape finished successfully with {total_rows} total records.")
                    self.log_msg.emit(f"  Output saved to: {output_path}")
                    self.progress.emit(section_count, section_count, "Complete!")
                    self.finished.emit(output_path, total_rows)
                else:
                    ScrapeHistoryManager.complete_session(self.session_id, "Completed")
                    self.log_msg.emit("Scrape finished. No new active promotions or data extracted.")
                    self.finished.emit(output_path, 0)

        except Exception as e:
            self.log_msg.emit(f"ERROR: {e}")
            import traceback
            self.log_msg.emit(traceback.format_exc())
            ScrapeHistoryManager.complete_session(self.session_id, "Failed")
            self.finished.emit("", 0)


# ────────────────────────────────────────────────────────────
#  UI Widget — embeddable in the main window's tab
# ────────────────────────────────────────────────────────────

class UniversalScraperTab(QWidget):
    """A full widget that can be added as a tab in the main window."""

    log_signal = pyqtSignal(str)
    ui_signal = pyqtSignal(object)

    def __init__(self):
        super().__init__()
        self.current_job = None
        self.worker = None
        self.log_signal.connect(self._append_log)
        self.ui_signal.connect(lambda fn: fn())
        self._build_ui()
        self._refresh_source_fields()
        self._load_history_into_table()

    def _append_log(self, msg: str):
        self.log.append(msg)
        scrollbar = self.log.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        self.sub_tabs = QTabWidget()
        self.sub_tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #333; background: #1a1a2e; }
            QTabBar::tab { background: #0f3460; color: #aaa; padding: 10px 24px;
                          margin-right: 2px; border-top-left-radius: 4px;
                          border-top-right-radius: 4px; font-weight: bold; }
            QTabBar::tab:selected { background: #16213e; color: #e94560; border-bottom: 2px solid #e94560; }
            QTabBar::tab:hover { background: #1a3a6a; color: white; }
        """)
        main_layout.addWidget(self.sub_tabs)
        
        # ── Tab 1: Config ────────────────────────────────────
        config_widget = QWidget()
        self._build_config_ui(config_widget)
        self.sub_tabs.addTab(config_widget, "Scraper Config")
        
        # ── Tab 2: History ───────────────────────────────────
        history_widget = QWidget()
        self._build_history_ui(history_widget)
        self.sub_tabs.addTab(history_widget, "Scrape History & Resume")

    def _build_config_ui(self, widget):
        layout = QVBoxLayout(widget)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # ── Header ────────────────────────────────────────
        title = QLabel("Universal Data Scraper")
        title.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        subtitle = QLabel("Configure and run targeted extranet scrapes")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(subtitle)

        # ── Source selector ────────────────────────────────
        src_row = QHBoxLayout()
        src_row.addWidget(QLabel("Data Source:"))
        self.source_combo = QComboBox()
        for key, src in EXTRANET_SOURCES.items():
            self.source_combo.addItem(src.source_name, userData=key)
        self.source_combo.currentIndexChanged.connect(self._refresh_source_fields)
        self.source_combo.setStyleSheet(
            "background: #16213e; color: white; border: 1px solid #444; "
            "border-radius: 4px; padding: 6px; font-size: 13px;"
        )
        src_row.addWidget(self.source_combo, 1)

        self.login_btn = QPushButton("Login")
        self.login_btn.setStyleSheet("background-color: #0a7; padding: 8px 16px;")
        self._login_mode = "login"  # "login" | "confirm"
        self.login_btn.clicked.connect(self._on_login_btn_clicked)
        src_row.addWidget(self.login_btn)

        self.clear_cache_btn = QPushButton("Clear Cache")
        self.clear_cache_btn.setStyleSheet("background-color: #c0392b; padding: 8px 16px;")
        self.clear_cache_btn.clicked.connect(self._clear_cache)
        src_row.addWidget(self.clear_cache_btn)

        self.session_label = QLabel("Not logged in")
        self.session_label.setStyleSheet("color: #888; font-size: 11px;")
        src_row.addWidget(self.session_label)
        layout.addLayout(src_row)

        # ── Field selection (scrollable) ──────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        scroll_container = QWidget()
        self.fields_layout = QVBoxLayout(scroll_container)
        scroll.setWidget(scroll_container)
        scroll.setMaximumHeight(260)
        scroll.setMinimumHeight(150)
        layout.addWidget(QLabel("Select fields to scrape:"))
        layout.addWidget(scroll)

        # ── Job controls ──────────────────────────────────
        ctrl_row = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(self._select_all_fields)
        self.select_all_btn.setStyleSheet("background-color: #34495e;")
        ctrl_row.addWidget(self.select_all_btn)

        self.deselect_all_btn = QPushButton("Deselect All")
        self.deselect_all_btn.clicked.connect(self._deselect_all_fields)
        self.deselect_all_btn.setStyleSheet("background-color: #34495e;")
        ctrl_row.addWidget(self.deselect_all_btn)

        self.scrape_all_btn = QPushButton("Scrape All Data")
        self.scrape_all_btn.clicked.connect(self._scrape_all_data)
        self.scrape_all_btn.setStyleSheet("background-color: #8e44ad; font-weight: bold;")
        ctrl_row.addWidget(self.scrape_all_btn)

        ctrl_row.addStretch()

        self.export_config_btn = QPushButton("Export Config")
        self.export_config_btn.clicked.connect(self._export_config)
        self.export_config_btn.setStyleSheet("background-color: #555;")
        ctrl_row.addWidget(self.export_config_btn)

        self.import_config_btn = QPushButton("Import Config")
        self.import_config_btn.clicked.connect(self._import_config)
        self.import_config_btn.setStyleSheet("background-color: #555;")
        ctrl_row.addWidget(self.import_config_btn)
        layout.addLayout(ctrl_row)

        # ── Settings & Config Panel ───────────────────────
        settings_group = QGroupBox("Advanced Settings")
        settings_group.setStyleSheet("QGroupBox { border: 1px solid #444; margin-top: 10px; font-weight: bold; }")
        settings_layout = QHBoxLayout(settings_group)
        
        self.fast_mode_chk = QCheckBox("Enable Fast Mode (Blocks images/CSS)")
        self.fast_mode_chk.setChecked(True)
        self.fast_mode_chk.setToolTip("Aborts loading images, fonts, and stylesheets to massively speed up scraping.")
        settings_layout.addWidget(self.fast_mode_chk)

        self.headless_chk = QCheckBox("Headless Mode (Hide browser)")
        self.headless_chk.setChecked(False)
        settings_layout.addWidget(self.headless_chk)

        from PyQt6.QtWidgets import QSpinBox
        settings_layout.addWidget(QLabel("Concurrency limit:"))
        self.concurrency_spin = QSpinBox()
        self.concurrency_spin.setRange(1, 10)
        self.concurrency_spin.setValue(1)
        self.concurrency_spin.setToolTip("Number of parallel browser tabs to open when scraping multiple properties")
        settings_layout.addWidget(self.concurrency_spin)
        settings_layout.addStretch()

        layout.addWidget(settings_group)

        # ── Run controls ──────────────────────────────────
        run_row = QHBoxLayout()
        self.output_label = QLabel("Output label:")
        self.output_label.setStyleSheet("color: #ccc;")
        run_row.addWidget(self.output_label)

        self.label_input = QTextEdit()
        self.label_input.setMaximumHeight(32)
        self.label_input.setPlaceholderText("e.g. Oct2024_Bookings")
        self.label_input.setStyleSheet(
            "background: #16213e; color: white; border: 1px solid #444; "
            "border-radius: 4px; font-size: 12px;"
        )
        run_row.addWidget(self.label_input, 1)

        self.run_btn = QPushButton("▶  Start Scrape")
        self.run_btn.setStyleSheet("background-color: #27ae60; font-weight: bold; padding: 10px 24px;")
        self.run_btn.clicked.connect(lambda: self._run_job())
        run_row.addWidget(self.run_btn)

        self.stop_btn = QPushButton("■ Stop")
        self.stop_btn.setStyleSheet("background-color: #c0392b; font-weight: bold;")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop_job)
        run_row.addWidget(self.stop_btn)
        layout.addLayout(run_row)

        # ── Progress ──────────────────────────────────────
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # ── Log ───────────────────────────────────────────
        log_header_layout = QHBoxLayout()
        log_header_layout.addWidget(QLabel("Log:"))
        log_header_layout.addStretch()
        
        clear_log_btn = QPushButton("Clear Logs")
        clear_log_btn.setStyleSheet("background-color: #7f8c8d; font-size: 10px; padding: 2px 8px; max-height: 20px;")
        clear_log_btn.clicked.connect(lambda: self.log.clear())
        log_header_layout.addWidget(clear_log_btn)
        
        layout.addLayout(log_header_layout)
        
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumHeight(80)
        self.log.setStyleSheet(
            "background-color: #16213e; color: #a0e0a0; border: 1px solid #333; "
            "border-radius: 4px; font-family: Consolas; font-size: 11px;"
        )
        layout.addWidget(self.log)

        # ── Live Preview ──────────────────────────────────
        self.live_preview = QTableWidget()
        self.live_preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.live_preview.setColumnCount(4)
        self.live_preview.setHorizontalHeaderLabels(["Hotel ID", "Name", "Key", "Value"])
        self.live_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.live_preview.setStyleSheet("""
            QTableWidget { background-color: #1a1a2e; color: #e0e0e0; gridline-color: #333; border: 1px solid #333; font-size: 11px; }
            QHeaderView::section { background-color: #0f3460; color: white; padding: 4px; border: 1px solid #333; font-weight: bold; }
        """)
        layout.addWidget(QLabel("Live Preview Data:"))
        layout.addWidget(self.live_preview)

    def _build_history_ui(self, widget):
        layout = QVBoxLayout(widget)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # Title
        hist_title = QLabel("Scrape Session History & Resume")
        hist_title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        hist_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(hist_title)

        # Filters Row
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Filter by Month:"))
        self.month_filter = QComboBox()
        self.month_filter.addItem("All Months")
        self.month_filter.currentIndexChanged.connect(self._load_history_into_table)
        filter_row.addWidget(self.month_filter)

        filter_row.addWidget(QLabel("Filter by Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All Statuses", "Completed", "Running", "Interrupted", "Failed"])
        self.status_filter.currentIndexChanged.connect(self._load_history_into_table)
        filter_row.addWidget(self.status_filter)

        filter_row.addStretch()

        self.refresh_hist_btn = QPushButton("Refresh")
        self.refresh_hist_btn.setStyleSheet("background-color: #2980b9; padding: 6px 14px;")
        self.refresh_hist_btn.clicked.connect(self._load_history_into_table)
        filter_row.addWidget(self.refresh_hist_btn)
        layout.addLayout(filter_row)

        # History Table
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels([
            "Date & Time", "Platform", "Progress", "Records", "Status", "Actions"
        ])
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.setStyleSheet("""
            QTableWidget { background-color: #16213e; color: #e0e0e0; gridline-color: #333; border: 1px solid #333; }
            QHeaderView::section { background-color: #0f3460; color: white; padding: 6px; border: 1px solid #333; font-weight: bold; }
            QTableWidget::item { padding: 6px; }
        """)
        
        # Set column widths/strech
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        self.history_table.setColumnWidth(5, 180)
        self.history_table.setColumnWidth(1, 150)
        self.history_table.setColumnWidth(2, 80)
        self.history_table.setColumnWidth(3, 80)
        self.history_table.setColumnWidth(4, 90)
        
        layout.addWidget(self.history_table)

    def _load_history_into_table(self):
        """Load, filter, and render history items from the SQLite database."""
        history = ScrapeHistoryManager.get_history()
        
        # Populate month filter combo dynamically if not already filled
        current_month_sel = self.month_filter.currentText()
        self.month_filter.blockSignals(True)
        self.month_filter.clear()
        self.month_filter.addItem("All Months")
        
        months = set()
        for h in history:
            ts = h.get("timestamp", "")
            if len(ts) >= 7:
                year_month = ts[:7]
                try:
                    dt = datetime.strptime(year_month, "%Y-%m")
                    months.add(dt.strftime("%B %Y"))
                except Exception:
                    months.add(year_month)
                    
        for m in sorted(list(months), reverse=True):
            self.month_filter.addItem(m)
            
        idx = self.month_filter.findText(current_month_sel)
        if idx >= 0:
            self.month_filter.setCurrentIndex(idx)
        self.month_filter.blockSignals(False)

        # Apply filters
        month_sel = self.month_filter.currentText()
        status_sel = self.status_filter.currentText()
        
        filtered_history = []
        for h in history:
            if month_sel != "All Months":
                ts = h.get("timestamp", "")
                if len(ts) >= 7:
                    try:
                        dt = datetime.strptime(ts[:7], "%Y-%m")
                        m_str = dt.strftime("%B %Y")
                        if m_str != month_sel:
                            continue
                    except Exception:
                        if ts[:7] != month_sel:
                            continue
                            
            if status_sel != "All Statuses":
                if h.get("status", "") != status_sel:
                    continue
                    
            filtered_history.append(h)

        self.history_table.setRowCount(0)
        self.history_table.setRowCount(len(filtered_history))
        
        for row_idx, h in enumerate(filtered_history):
            self.history_table.setItem(row_idx, 0, QTableWidgetItem(h.get("timestamp", "")))
            self.history_table.setItem(row_idx, 1, QTableWidgetItem(h.get("platform", "")))
            
            total_p = h.get("total_properties", 0) or 0
            proc_p = h.get("processed_properties", 0) or 0
            progress_str = f"{proc_p} / {total_p}" if total_p > 0 else "1 / 1"
            self.history_table.setItem(row_idx, 2, QTableWidgetItem(progress_str))
            self.history_table.setItem(row_idx, 3, QTableWidgetItem(str(h.get("total_rows", 0))))
            
            status = h.get("status", "")
            status_item = QTableWidgetItem(status)
            if status == "Completed":
                status_item.setForeground(Qt.GlobalColor.green)
            elif status == "Running":
                status_item.setForeground(Qt.GlobalColor.cyan)
            elif status == "Interrupted":
                status_item.setForeground(Qt.GlobalColor.yellow)
            elif status == "Failed":
                status_item.setForeground(Qt.GlobalColor.red)
            self.history_table.setItem(row_idx, 4, status_item)
            
            self.history_table.setRowHeight(row_idx, 36)
            
            # Actions cell
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(4, 2, 4, 2)
            actions_layout.setSpacing(4)
            
            open_btn = QPushButton("Open Excel")
            open_btn.setStyleSheet("background-color: #27ae60; color: white; border: none; border-radius: 4px; font-size: 11px; font-weight: bold; min-height: 22px; padding: 2px 6px;")
            out_path = h.get("output_path", "")
            open_btn.clicked.connect(lambda checked, p=out_path: self._open_excel_file(p))
            actions_layout.addWidget(open_btn)
            
            resume_btn = QPushButton("Resume")
            resume_btn.setStyleSheet("background-color: #d35400; color: white; border: none; border-radius: 4px; font-size: 11px; font-weight: bold; min-height: 22px; padding: 2px 6px;")
            session_id = h.get("id", "")
            
            # Can resume if running/interrupted/failed and there are properties remaining to scrape
            can_resume = status in ("Interrupted", "Failed", "Running") and total_p > proc_p
            resume_btn.setEnabled(can_resume)
            if not can_resume:
                resume_btn.setStyleSheet("background-color: #333; color: #666; border: none; border-radius: 4px; font-size: 11px; font-weight: bold; min-height: 22px; padding: 2px 6px;")
                
            resume_btn.clicked.connect(lambda checked, s_id=session_id: self._resume_session(s_id))
            actions_layout.addWidget(resume_btn)
            
            cell_widget = QWidget()
            cell_widget.setLayout(actions_layout)
            self.history_table.setCellWidget(row_idx, 5, cell_widget)

    def _open_excel_file(self, output_path: str):
        if not output_path or not os.path.exists(output_path):
            self.log_msg(f"Error: Excel file not found at {output_path}")
            return
        try:
            os.startfile(output_path)
            self.log_msg(f"Opened file: {output_path}")
        except Exception as e:
            self.log_msg(f"Failed to open Excel file: {e}")

    def _resume_session(self, session_id: str):
        session = ScrapeHistoryManager.get_session(session_id)
        if not session:
            self.log_msg("Error: Session not found in database.")
            return
            
        output_path = session.get("output_path", "")
        source_key = session.get("source_key", "")
        fields_str = session.get("fields", "[]")
        
        try:
            fields = json.loads(fields_str)
        except Exception:
            self.log_msg("Error: Failed to parse fields configuration for this session.")
            return
            
        label = session_id.replace("session_", "")
        job = ScrapeJob(
            source_key=source_key,
            selected_fields=fields,
            label=label,
            output_path=output_path
        )
        
        self.sub_tabs.setCurrentIndex(0)
        self._run_job(job, session_id=session_id)

    def _refresh_source_fields(self):
        """Rebuild the field checklist for the currently selected source."""
        while self.fields_layout.count():
            item = self.fields_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        source_key = self.source_combo.currentData()
        source = EXTRANET_SOURCES.get(source_key)
        if not source:
            return

        self._update_session_status()

        for group_def in source.available_fields:
            group_box = QGroupBox(group_def["group"])
            group_box.setStyleSheet("""
                QGroupBox { color: #e0e0e0; font-weight: bold; border: 1px solid #444;
                            border-radius: 6px; margin-top: 10px; padding-top: 16px; }
                QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
            """)
            group_layout = QVBoxLayout()
            for field in group_def["fields"]:
                cb = QCheckBox(field["label"])
                cb.setProperty("field_key", field["key"])
                cb.setStyleSheet("color: #ccc; spacing: 6px;")
                group_layout.addWidget(cb)
            group_box.setLayout(group_layout)
            self.fields_layout.addWidget(group_box)

        self.fields_layout.addStretch()

    def _update_session_status(self):
        source_key = self.source_combo.currentData()
        source = EXTRANET_SOURCES.get(source_key)
        if source and hasattr(source, 'cookies_path') and source.cookies_path.exists():
            self.session_label.setText("✓ Session active")
            self.session_label.setStyleSheet("color: #0a7; font-size: 11px;")
        else:
            self.session_label.setText("Not logged in")
            self.session_label.setStyleSheet("color: #888; font-size: 11px;")

    def _clear_cache(self):
        import shutil
        self.log_msg("\nClearing cache and login sessions...")
        cookie_files = [
            BOOKING_EXTRANET_COOKIES,
            MMT_EXTRANET_COOKIES,
            GOIBIBO_EXTRANET_COOKIES,
            AGODA_EXTRANET_COOKIES,
            EXPEDIA_EXTRANET_COOKIES
        ]
        cleared_count = 0
        for f in cookie_files:
            try:
                if f.exists():
                    f.unlink()
                    cleared_count += 1
            except Exception as e:
                self.log_msg(f"Could not delete cookie file {f.name}: {e}")

        # Clear saved session params JSON files
        for key in EXTRANET_SOURCES.keys():
            params_file = COOKIES_DIR / f"{key}_params.json"
            if params_file.exists():
                try:
                    params_file.unlink()
                except Exception:
                    pass

        chrome_profile_dir = COOKIES_DIR / 'chrome_extranet'
        if chrome_profile_dir.exists():
            try:
                shutil.rmtree(chrome_profile_dir, ignore_errors=False)
                self.log_msg("Chrome persistent profile directory deleted.")
            except Exception as e:
                self.log_msg(f"Warning: Could not completely delete Chrome profile folder: {e}")
                self.log_msg("Some browser files may be locked. Please make sure all scraper browser windows are closed.")
        
        self.log_msg("Session cache and local cookie files cleared successfully.")
        self._update_session_status()

    def _on_login_btn_clicked(self):
        if self._login_mode == "login":
            self._do_login()
        elif self._login_mode == "confirm":
            self._confirm_login()
        elif self._login_mode == "worker_confirm":
            self._confirm_worker_login()

    def _do_login(self):
        source_key = self.source_combo.currentData()
        source = EXTRANET_SOURCES.get(source_key)
        if not source:
            return

        self.log_signal.emit(f"\nOpening {source.source_name} login...")
        self.login_btn.setEnabled(False)
        self._login_event = threading.Event()

        def login_thread():
            try:
                # Terminate only Chrome instances running on our debug port or using our profile to prevent affecting user's personal Chrome
                try:
                    import psutil
                    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                        try:
                            if proc.info['name'] == 'chrome.exe':
                                cmdline = proc.info['cmdline'] or []
                                cmd_str = ' '.join(cmdline).lower()
                                if 'remote-debugging-port=9223' in cmd_str or 'chrome_extranet' in cmd_str:
                                    p = psutil.Process(proc.info['pid'])
                                    p.terminate()
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            pass
                    lock_file = COOKIES_DIR / 'chrome_extranet' / 'SingletonLock'
                    if lock_file.exists():
                        lock_file.unlink()
                except Exception:
                    pass

                pw = sync_playwright().start()
                context = pw.chromium.launch_persistent_context(
                    user_data_dir=str(COOKIES_DIR / 'chrome_extranet'),
                    headless=False,
                    channel="chrome",
                    args=[
                        f"--remote-debugging-port={EXTRANET_DEBUG_PORT}",
                        "--no-first-run",
                        "--window-size=1280,900",
                    ]
                )
                page = context.pages[0] if context.pages else context.new_page()
                page.goto(source.login_url, timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)
                self.log_signal.emit(f"Chrome opened at {source.login_url}")
                self.log_signal.emit("Log in to the browser window, then click 'Confirm Login'")
                self.ui_signal.emit(self._show_confirm_login_button)
                self._login_event.wait()
                cookies = context.cookies()
                if hasattr(source, 'cookies_path'):
                    with open(source.cookies_path, "wb") as f:
                        pickle.dump(cookies, f)
                
                # Extract and save session params
                try:
                    url = page.url
                    hotel_id_match = re.search(r'hotel_id=(\d+)', url)
                    ses_match = re.search(r'ses=([a-f0-9]+)', url)
                    params = {}
                    if hotel_id_match:
                        params["hotel_id"] = hotel_id_match.group(1)
                    if ses_match:
                        params["ses"] = ses_match.group(1)
                    if params:
                        params_path = COOKIES_DIR / f"{source.source_key}_params.json"
                        with open(params_path, "w") as f:
                            json.dump(params, f)
                except Exception:
                    pass
                
                context.close()
                pw.stop()
                self.log_signal.emit(f"{source.source_name} session saved!")
                self.ui_signal.emit(self._update_session_status)
            except Exception as e:
                self.log_signal.emit(f"Login error: {e}")
            self.ui_signal.emit(self._reset_login_button)

        threading.Thread(target=login_thread, daemon=True).start()

    def _show_confirm_login_button(self):
        self._login_mode = "confirm"
        self.login_btn.setText("Click here after logging in -> Confirm")
        self.login_btn.setStyleSheet(
            "background-color: #f39c12; font-weight: bold; padding: 8px 16px;"
        )
        self.login_btn.setEnabled(True)

    def _confirm_login(self):
        self._login_mode = "login"
        if hasattr(self, '_login_event') and self._login_event:
            self._login_event.set()
        self.login_btn.setEnabled(False)
        self.login_btn.setText("Saving session...")

    def _reset_login_button(self):
        self._login_mode = "login"
        self.login_btn.setText("Login")
        self.login_btn.setStyleSheet("background-color: #0a7; padding: 8px 16px;")
        self.login_btn.setEnabled(True)

    def log_msg(self, msg: str):
        self.log.append(msg)
        scrollbar = self.log.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def _select_all_fields(self):
        self._set_all_checkboxes(True)

    def _deselect_all_fields(self):
        self._set_all_checkboxes(False)

    def _scrape_all_data(self):
        self._set_all_checkboxes(True)
        self.log_msg("All fields selected — starting scrape with all available data...")
        self._run_job()

    def _set_all_checkboxes(self, checked: bool):
        for i in range(self.fields_layout.count()):
            item = self.fields_layout.itemAt(i)
            if item and item.widget() and isinstance(item.widget(), QGroupBox):
                for cb in item.widget().findChildren(QCheckBox):
                    cb.setChecked(checked)

    def _get_selected_fields(self) -> list[dict]:
        fields = []
        for i in range(self.fields_layout.count()):
            item = self.fields_layout.itemAt(i)
            if item and item.widget() and isinstance(item.widget(), QGroupBox):
                for cb in item.widget().findChildren(QCheckBox):
                    if cb.isChecked():
                        key = cb.property("field_key")
                        fields.append({"key": key, "label": cb.text()})
        return fields

    def _export_config(self):
        source_key = self.source_combo.currentData()
        fields = self._get_selected_fields()
        if not fields:
            self.log_msg("No fields selected to export.")
            return
        label = self.label_input.toPlainText().strip() or f"scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        job = ScrapeJob(source_key=source_key, selected_fields=fields, label=label)

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Config", str(Path.home() / "Downloads" / f"{label}.json"),
            "JSON Files (*.json)"
        )
        if path:
            job.to_file(path)
            self.log_msg(f"Config saved: {path}")

    def _import_config(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Load Config", str(Path.home() / "Downloads"), "JSON Files (*.json)"
        )
        if not path:
            return
        try:
            job = ScrapeJob.from_file(path)
            idx = self.source_combo.findData(job.source_key)
            if idx >= 0:
                self.source_combo.setCurrentIndex(idx)
            for field in job.selected_fields:
                for i in range(self.fields_layout.count()):
                    item = self.fields_layout.itemAt(i)
                    if item and item.widget() and isinstance(item.widget(), QGroupBox):
                        for cb in item.widget().findChildren(QCheckBox):
                            if cb.property("field_key") == field["key"]:
                                cb.setChecked(True)
            self.log_msg(f"Config loaded: {Path(path).name} — {len(job.selected_fields)} fields selected")
        except Exception as e:
            self.log_msg(f"Failed to load config: {e}")

    def _run_job(self, job=None, session_id=None):
        if job is None:
            source_key = self.source_combo.currentData()
            fields = self._get_selected_fields()
            if not fields:
                self.log_msg("Select at least one field to scrape.")
                return

            today_str = datetime.now().strftime("%Y-%m-%d")
            last_year_str = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
            default_label = f"ACTIVE_promotions_report__from_{last_year_str}_to_{today_str}"

            label = self.label_input.toPlainText().strip() or default_label
            output_path = str(Path.home() / "Downloads" / f"{label}.xlsx")
            
            headless_mode = self.headless_chk.isChecked()
            fast_mode = self.fast_mode_chk.isChecked()
            concurrency = self.concurrency_spin.value()

            job = ScrapeJob(
                source_key=source_key,
                selected_fields=fields,
                label=label,
                output_path=output_path,
                headless=headless_mode,
                fast_mode=fast_mode,
                concurrency=concurrency
            )
        self.current_job = job

        self.log_msg(f"\n{'='*50}")
        if session_id:
            self.log_msg(f"Resuming Job: {job.display_summary()}")
        else:
            self.log_msg(f"Job: {job.display_summary()}")
        self.log_msg(f"Output: {job.output_path}")
        self.log_msg(f"{'='*50}")

        self.run_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setVisible(True)
        self.progress.setMaximum(1)
        self.progress.setValue(0)
        self.live_preview.setRowCount(0)

        self.worker = ExtranetScrapeWorker(job, session_id=session_id)
        self.worker.progress.connect(self._on_worker_progress)
        self.worker.finished.connect(self._on_worker_finished)
        self.worker.log_msg.connect(self.log_msg)
        self.worker.login_required.connect(self._on_worker_login_required)
        self.worker.live_data.connect(self._on_live_data)
        self.worker.start()

    def _on_live_data(self, row_data: dict):
        # Insert a few key/value pairs from the row into the live preview table
        hotel_id = str(row_data.get("hotel_id", ""))
        hotel_name = str(row_data.get("hotel_name", ""))
        
        for k, v in row_data.items():
            if k in ["hotel_id", "hotel_name", "_source", "_error"]:
                continue
            if not v:
                continue
            row_idx = self.live_preview.rowCount()
            self.live_preview.insertRow(row_idx)
            self.live_preview.setItem(row_idx, 0, QTableWidgetItem(hotel_id))
            self.live_preview.setItem(row_idx, 1, QTableWidgetItem(hotel_name))
            self.live_preview.setItem(row_idx, 2, QTableWidgetItem(str(k)))
            self.live_preview.setItem(row_idx, 3, QTableWidgetItem(str(v)))
            self.live_preview.scrollToBottom()

    def _stop_job(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log_msg("Stopping...")

    def _on_worker_progress(self, current, total, status):
        self.progress.setMaximum(total)
        self.progress.setValue(current)
        self.progress.setFormat(f"{status}  ({current}/{total})")

    def _on_worker_finished(self, output_path: str, rows: int):
        self.run_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self._reset_login_button()
        if output_path and rows >= 0:
            self._load_history_into_table()
            self.log_msg(f"\n✓ Complete! {rows} rows -> {output_path}")
            self.progress.setFormat(f"Done! {rows} rows")
        else:
            self.progress.setFormat("No data extracted")
        import winsound
        try:
            winsound.MessageBeep(winsound.MB_ICONASTERISK)
        except Exception:
            pass

    def _on_worker_login_required(self, message: str):
        self.log_msg(message)
        self.log_msg("Chrome opened — log in to the browser window, then click Confirm Login below.")
        self.login_btn.setText("Click here after logging in -> Confirm (Worker)")
        self.login_btn.setStyleSheet(
            "background-color: #e67e22; font-weight: bold; padding: 8px 16px;"
        )
        self.login_btn.setEnabled(True)
        self._login_mode = "worker_confirm"

    def _confirm_worker_login(self):
        if self.worker and hasattr(self.worker, 'login_event'):
            self.worker.login_event.set()
        self._login_mode = "login"
        self.login_btn.setText("Scraping in progress...")
        self.login_btn.setStyleSheet(
            "background-color: #555; padding: 8px 16px;"
        )
        self.login_btn.setEnabled(False)
        self.log_msg("Session confirmed — continuing scrape...")
