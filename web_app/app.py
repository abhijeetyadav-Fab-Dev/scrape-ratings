import sys
import os
import threading
import json
import pickle
import re
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, render_template, jsonify, request, send_file
from playwright.sync_api import sync_playwright

# Add parent directory to path so we can import from universal_scraper
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from universal_scraper import (
    EXTRANET_SOURCES, ScrapeJob, ScrapeHistoryManager, COOKIES_DIR,
    EXTRANET_DEBUG_PORT, clean_row_keys
)

# Initialize database
ScrapeHistoryManager.init_db()

app = Flask(__name__)
active_jobs = {} # session_id -> WebScrapeWorker

class WebScrapeWorker(threading.Thread):
    def __init__(self, job, session_id):
        super().__init__()
        self.job = job
        self.session_id = session_id
        self._stop = False
        self.login_event = threading.Event()
        self.logs = []
        self.progress_val = 0
        self.progress_max = 1
        self.progress_status = "Waiting..."
        self.login_required_msg = None
        self.finished_state = False
        self.output_path = ""
        self.total_rows = 0

    def log(self, msg):
        self.logs.append(msg)
        try:
            print(f"[{self.session_id}] {msg}")
        except UnicodeEncodeError:
            try:
                encoding = sys.stdout.encoding or 'utf-8'
                safe_msg = str(msg).encode(encoding, errors='replace').decode(encoding)
                print(f"[{self.session_id}] {safe_msg}")
            except Exception:
                safe_msg = str(msg).encode('ascii', errors='replace').decode('ascii')
                print(f"[{self.session_id}] {safe_msg}")

    def stop(self):
        self._stop = True

    def run(self):
        self.log(f"Starting background scrape job for session {self.session_id}...")
        source = EXTRANET_SOURCES.get(self.job.source_key)
        if not source:
            self.log(f"Error: Unknown source: {self.job.source_key}")
            self.progress_status = "Error: Unknown source"
            self.finished_state = True
            return

        source_name = source.source_name
        today_str = datetime.now().strftime("%Y-%m-%d")
        last_year_str = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
        default_label = f"ACTIVE_promotions_report__from_{last_year_str}_to_{today_str}"
        
        output_path = self.job.output_path or str(
            Path.home() / "Downloads" / f"{self.job.label or default_label}.xlsx"
        )
        self.output_path = output_path
        
        # Set attributes on source singleton
        source.output_path = output_path
        source.session_id = self.session_id
        source.job = self.job
        
        # Mock signal callbacks to interface with the ExtranetSource plugin
        class MockWorkerSignals:
            def __init__(self, worker_ref):
                self.worker_ref = worker_ref
            def emit(self, *args):
                pass
                
        class MockSignalLog:
            def __init__(self, worker_ref):
                self.worker_ref = worker_ref
            def emit(self, msg):
                self.worker_ref.log(msg)
                
        class MockSignalProgress:
            def __init__(self, worker_ref):
                self.worker_ref = worker_ref
            def emit(self, current, total, status):
                self.worker_ref.progress_val = current
                self.worker_ref.progress_max = total
                self.worker_ref.progress_status = status
                
        class MockSignalLoginRequired:
            def __init__(self, worker_ref):
                self.worker_ref = worker_ref
            def emit(self, msg):
                self.worker_ref.login_required_msg = msg
                self.worker_ref.log(f"[LOGIN ACTION REQUIRED] {msg}")

        class MockWorker:
            def __init__(self, worker_ref):
                self.worker_ref = worker_ref
                self.log_msg = MockSignalLog(worker_ref)
                self.progress = MockSignalProgress(worker_ref)
                self.login_required = MockSignalLoginRequired(worker_ref)
                self.live_data = MockWorkerSignals(worker_ref)
                self.finished = MockWorkerSignals(worker_ref)
                self.login_event = worker_ref.login_event

            @property
            def _stop(self):
                return self.worker_ref._stop
            
            @_stop.setter
            def _stop(self, val):
                self.worker_ref._stop = val

        mock_worker = MockWorker(self)
        source.worker = mock_worker

        # Initialize session in SQLite database
        selected_field_keys = self.job.selected_fields
        ScrapeHistoryManager.create_session(
            self.session_id, source_name, self.job.source_key, selected_field_keys, output_path
        )

        # Prepare Excel headers
        field_keys = [clean_row_keys({f["key"]: ""}).popitem()[0] for f in self.job.selected_fields]
        ordered_keys = []
        for k in field_keys:
            if k not in ordered_keys:
                ordered_keys.append(k)
        for k in ["hotel_id", "hotel_name", "sub_tab", "_source", "_error"]:
            if k not in ordered_keys:
                ordered_keys.append(k)
        
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        file_exists = os.path.exists(output_path) and os.path.getsize(output_path) > 0
        if not file_exists:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(ordered_keys)
                wb.save(output_path)
            except Exception as e:
                self.log(f"Warning: could not initialize Excel file: {e}")

        self.log(f"Starting scrape from {source_name}")
        self.progress_status = "Launching browser..."
        self.progress_val = 0
        self.progress_max = 1

        try:
            pw = sync_playwright().start()
            # Launch persistent context using debug port 9224 to not conflict with desktop app's 9223
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(COOKIES_DIR / 'chrome_extranet_web'),
                headless=self.job.headless,
                channel="chrome",
                args=[
                    f"--remote-debugging-port={EXTRANET_DEBUG_PORT + 1}",
                    "--no-first-run",
                    "--window-size=1280,900",
                ]
            )

            if getattr(self.job, 'fast_mode', False):
                def block_resources(route):
                    if route.request.resource_type in ["image", "stylesheet", "font", "media"]:
                        route.abort()
                    else:
                        route.continue_()
                context.route("**/*", block_resources)

            page = context.pages[0] if context.pages else context.new_page()

            # ── Login (if needed) ──────────────────────────
            cookies_path = source.cookies_path
            if cookies_path and cookies_path.exists():
                with open(cookies_path, "rb") as f:
                    cookies = pickle.load(f)
                context.add_cookies(cookies)
                self.log("Session cookies loaded.")
                
                # Load saved session params if present
                params_path = COOKIES_DIR / f"{source.source_key}_params.json"
                if params_path.exists():
                    try:
                        with open(params_path, "r") as f:
                            params = json.load(f)
                        source.current_hotel_id = params.get("hotel_id")
                        source.current_ses = params.get("ses")
                        self.log(f"Loaded saved session params: hotel_id={source.current_hotel_id}")
                    except Exception as e:
                        self.log(f"Warning: could not load session params: {e}")
                
                if getattr(source, "current_hotel_id", None) and getattr(source, "current_ses", None):
                    target_url = f"https://admin.booking.com/hotel/hoteladmin/extranet_ng/manage/home.html?hotel_id={source.current_hotel_id}&ses={source.current_ses}&lang=en"
                    self.log(f"Navigating directly to property dashboard: {source.current_hotel_id}")
                    page.goto(target_url, timeout=30000, wait_until="domcontentloaded")
                else:
                    page.goto(source.login_url, timeout=30000, wait_until="domcontentloaded")
                
                is_session_active = False
                for i in range(8):
                    page.wait_for_timeout(1000)
                    current_url = page.url.lower()
                    
                    body_text = ""
                    try:
                        body_text = page.inner_text("body")[:2000].lower()
                    except Exception:
                        pass
                    
                    has_password_input = False
                    try:
                        has_password_input = page.query_selector("input[type='password']") is not None
                    except Exception:
                        pass
                        
                    login_indicators = [
                        "sign in to manage", "sign in with", "log in to your",
                        "enter your password", "forgot password", "create your account",
                        "login-form", "username", "password",
                    ]
                    has_login_text = any(ind in body_text for ind in login_indicators)
                    is_err, _ = source._is_error_page(page)
                    
                    if "admin.booking.com" in current_url:
                        has_session_params = "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url
                        has_dashboard_path = any(x in current_url for x in ("/hotel/", "/extranet/", "/dashboard/"))
                        
                        if has_session_params and has_dashboard_path and not has_password_input and not has_login_text and not is_err:
                            is_session_active = True
                            break
                        elif has_password_input or has_login_text or "login" in current_url or is_err:
                            is_session_active = False
                            break
                    else:
                        if not has_password_input and not has_login_text and not is_err and any(x in current_url for x in ("dashboard", "home", "extranet")):
                            is_session_active = True
                            break
                        elif has_password_input or has_login_text or "login" in current_url or is_err:
                            is_session_active = False
                            break
                
                if is_session_active:
                    self.log("Session active.")
                else:
                    self.log("Session expired — re-login required.")
                    self.login_required_msg = f"{source.source_name} session expired. Please log in using the opened desktop browser."
                    if self.job.headless:
                        self.log("WARNING: Job is running headless, but login is required. Restart with Headless option turned OFF to log in.")
                    
                    self.login_event.wait()
                    self.login_required_msg = None
                    self.log("Waiting for dashboard to load...")
                    
                    has_params = False
                    for _ in range(15):
                        current_url = page.url.lower()
                        if "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url:
                            has_params = True
                            break
                        page.wait_for_timeout(1000)
                    
                    cookies = context.cookies()
                    with open(cookies_path, "wb") as f:
                        pickle.dump(cookies, f)
                    
                    # Extract and save session params
                    try:
                        url = page.url
                        h_m = re.search(r'hotel_id=(\d+)', url)
                        s_m = re.search(r'ses=([a-f0-9]+)', url)
                        params = {}
                        if h_m:
                            params["hotel_id"] = h_m.group(1)
                        if s_m:
                            params["ses"] = s_m.group(1)
                        if params:
                            params_path = COOKIES_DIR / f"{source.source_key}_params.json"
                            with open(params_path, "w") as f:
                                json.dump(params, f)
                    except Exception:
                        pass
                    
                    self.log("Session refreshed.")
            else:
                self.log(f"Navigating to {source.source_name} login page...")
                page.goto(source.login_url, timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)
                self.login_required_msg = f"{source.source_name} login required. Please log in using the opened desktop browser."
                if self.job.headless:
                    self.log("WARNING: Job is running headless, but login is required. Restart with Headless option turned OFF to log in.")
                
                self.login_event.wait()
                self.login_required_msg = None
                self.log("Waiting for dashboard to load...")
                
                has_params = False
                for _ in range(15):
                    current_url = page.url.lower()
                    if "ses=" in current_url or "hotel_id=" in current_url or "hotel_account_id=" in current_url:
                        has_params = True
                        break
                    page.wait_for_timeout(1000)
                
                cookies = context.cookies()
                if cookies_path:
                    with open(cookies_path, "wb") as f:
                        pickle.dump(cookies, f)
                    
                    # Extract and save session params
                    try:
                        url = page.url
                        h_m = re.search(r'hotel_id=(\d+)', url)
                        s_m = re.search(r'ses=([a-f0-9]+)', url)
                        params = {}
                        if h_m:
                            params["hotel_id"] = h_m.group(1)
                        if s_m:
                            params["ses"] = s_m.group(1)
                        if params:
                            params_path = COOKIES_DIR / f"{source.source_key}_params.json"
                            with open(params_path, "w") as f:
                                json.dump(params, f)
                    except Exception:
                        pass
                    
                    self.log("Session saved.")

            # ── Group fields by section ────────────────────
            sections = {}
            for field in self.job.selected_fields:
                key = field["key"]
                parts = key.split("_")
                compound_key = "_".join(parts[:2]) if len(parts) >= 3 else None
                simple_key = parts[0] if len(parts) >= 2 else "general"
                
                section_map = {
                    "res": "reservations", "prop": "property",
                    "rev": "reviews", "fin": "financial",
                    "promo": "promotions",
                    "exp_ci": "insights", "htl_ci": "insights", "goi_rpt": "reports",
                    "mmt_rev": "reviews", "mmt_settlement": "financial",
                    "mmt_prop": "property", "mmt_amenities": "property",
                    "mmt_room_inventory": "property", "mmt_content": "property",
                    "mmt_missing": "property", "goi_rev": "reviews",
                    "goi_settlement": "financial", "goi_prop": "property",
                    "goi_amenities": "property", "goi_room_inventory": "property",
                    "goi_content": "property", "goi_missing": "property",
                    "agd_rev": "reviews", "agd_prop": "property",
                    "exp_rev": "reviews", "exp_prop": "property",
                    "htl_rev": "reviews", "htl_prop": "property",
                    "dash": "dashboard", "rate": "rates", "boost": "boost",
                    "inb": "inbox", "anl": "analytics",
                    "mmt": "reservations", "goi": "reservations",
                    "agd": "reservations", "exp": "reservations", "htl": "reservations",
                }
                
                if compound_key and compound_key in section_map:
                    section_key = section_map[compound_key]
                else:
                    section_key = section_map.get(simple_key, simple_key)
                
                sections.setdefault(section_key, []).append(field)

            if not sections:
                sections["general"] = self.job.selected_fields

            # ── Scrape each section ────────────────────────
            all_rows = []
            section_count = len(sections)
            for idx, (section_key, fields) in enumerate(sections.items()):
                if self._stop:
                    break
                self.log(f"Navigating to section: {section_key} ({idx+1}/{section_count})")
                self.progress_val = idx
                self.progress_max = section_count
                self.progress_status = f"Scraping {section_key}..."

                if source.multi_tab:
                    tab = context.new_page()
                    try:
                        source.navigate_to_section(tab, section_key)
                        page.wait_for_timeout(1000)
                        rows = source.extract_data(tab, fields)
                    finally:
                        tab.close()
                else:
                    source.navigate_to_section(page, section_key)
                    page.wait_for_timeout(1000)
                    rows = source.extract_data(page, fields)

                self.log(f"  -> Got {len(rows)} rows from '{section_key}'")
                all_rows.extend(rows)

            hotel_name = "Active Property"
            try:
                if page and not page.is_closed():
                    hotel_name = source._extract_property_name_from_page(page) or "Active Property"
            except Exception:
                pass

            # Cleanup happens in the finally block

            # ── Double-Safety Excel Write & Consolidation ───────────────────────────
            if self._stop:
                ScrapeHistoryManager.complete_session(self.session_id, "Interrupted")
                self.log("\nScrape job stopped/interrupted by user.")
                self.progress_status = "Interrupted"
            elif all_rows:
                existing_rows = []
                if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(output_path)
                        ws = wb.active
                        headers = [cell.value for cell in ws[1]]
                        for row_cells in ws.iter_rows(min_row=2, values_only=True):
                            existing_rows.append(dict(zip(headers, row_cells)))
                    except Exception as e:
                        self.log(f"Warning: could not read existing Excel: {e}")
                
                cleaned_all_rows = [clean_row_keys(r) for r in all_rows]
                scraped_hotel_ids = {r.get("hotel_id") for r in cleaned_all_rows if r.get("hotel_id")}
                consolidated = [r for r in existing_rows if r.get("hotel_id") not in scraped_hotel_ids]
                consolidated.extend(cleaned_all_rows)
                
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(ordered_keys)
                    for r in consolidated:
                        row_data = [r.get(k, "") for k in ordered_keys]
                        ws.append(row_data)
                    wb.save(output_path)
                except Exception as e:
                    self.log(f"Error writing final Excel file: {e}")

                try:
                    hotel_id = "single"
                    ScrapeHistoryManager.add_scraped_property(self.session_id, hotel_id, hotel_name, "Completed", len(cleaned_all_rows))
                except Exception as e:
                    self.log(f"Warning: could not write scrape history: {e}")

                ScrapeHistoryManager.complete_session(self.session_id, "Completed")
                
                session_stats = ScrapeHistoryManager.get_session(self.session_id)
                self.total_rows = session_stats.get("total_rows", len(consolidated)) if session_stats else len(consolidated)
                
                self.log(f"\nDone! Scrape finished successfully.")
                self.log(f"  Output saved to: {output_path}")
                self.progress_val = section_count
                self.progress_max = section_count
                self.progress_status = "Complete!"
            else:
                session_stats = ScrapeHistoryManager.get_session(self.session_id)
                self.total_rows = session_stats.get("total_rows", 0) if session_stats else 0
                if self.total_rows > 0:
                    ScrapeHistoryManager.complete_session(self.session_id, "Completed")
                    self.log(f"\nDone! Scrape finished successfully with {self.total_rows} total records.")
                    self.progress_val = section_count
                    self.progress_max = section_count
                    self.progress_status = "Complete!"
                else:
                    ScrapeHistoryManager.complete_session(self.session_id, "No Data")
                    self.log("No data scraped.")
                    self.progress_status = "Finished (No Data)"
        except Exception as e:
            self.log(f"Fatal error during scrape: {e}")
            ScrapeHistoryManager.complete_session(self.session_id, f"Error: {e}")
            self.progress_status = f"Error: {e}"
        finally:
            try:
                if 'context' in locals() and context:
                    context.close()
            except Exception:
                pass
            try:
                if 'pw' in locals() and pw:
                    pw.stop()
            except Exception:
                pass
            
        self.finished_state = True

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/sources")
def get_sources():
    sources_data = {}
    for key, src in EXTRANET_SOURCES.items():
        sources_data[key] = {
            "name": src.source_name,
            "available_fields": src.available_fields
        }
    return jsonify(sources_data)

@app.route("/api/scrape/start", methods=["POST"])
def start_scrape():
    data = request.json or {}
    source_key = data.get("source")
    field_keys = data.get("fields", [])
    headless = data.get("headless", False)
    fast_mode = data.get("fast_mode", False)
    
    if not source_key or not field_keys:
        return jsonify({"error": "Missing source or fields"}), 400
        
    src = EXTRANET_SOURCES.get(source_key)
    if not src:
        return jsonify({"error": "Invalid source"}), 400
        
    selected_fields = []
    for group in src.available_fields:
        for f in group["fields"]:
            if f["key"] in field_keys:
                selected_fields.append(f)
                
    session_id = f"web_session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    job = ScrapeJob(
        source_key=source_key,
        selected_fields=selected_fields,
        headless=headless,
        fast_mode=fast_mode
    )
    
    worker = WebScrapeWorker(job, session_id)
    active_jobs[session_id] = worker
    worker.start()
    
    return jsonify({
        "session_id": session_id,
        "status": "started",
        "message": f"Scrape job started for {src.source_name}"
    })

@app.route("/api/scrape/status/<session_id>")
def scrape_status(session_id):
    worker = active_jobs.get(session_id)
    if not worker:
        session_info = ScrapeHistoryManager.get_session(session_id)
        if session_info:
            return jsonify({
                "session_id": session_id,
                "finished": True,
                "status_text": session_info.get("status"),
                "progress_val": session_info.get("processed_properties", 0),
                "progress_max": session_info.get("total_properties", 0) or 1,
                "total_rows": session_info.get("total_rows", 0),
                "logs": [f"Session loaded from history: {session_info.get('status')}"]
            })
        return jsonify({"error": "Session not found"}), 404
        
    return jsonify({
        "session_id": session_id,
        "finished": worker.finished_state,
        "status_text": worker.progress_status,
        "progress_val": worker.progress_val,
        "progress_max": worker.progress_max,
        "login_required": worker.login_required_msg is not None,
        "login_msg": worker.login_required_msg,
        "total_rows": worker.total_rows,
        "logs": worker.logs
    })

@app.route("/api/scrape/stop/<session_id>", methods=["POST"])
def stop_scrape(session_id):
    worker = active_jobs.get(session_id)
    if worker:
        worker.stop()
        return jsonify({"message": "Stop request sent to worker"})
    return jsonify({"error": "Active session not found"}), 404

@app.route("/api/scrape/confirm_login/<session_id>", methods=["POST"])
def confirm_login(session_id):
    worker = active_jobs.get(session_id)
    if worker:
        worker.login_event.set()
        return jsonify({"message": "Login confirmed, resuming..."})
    return jsonify({"error": "Active session not found"}), 404

@app.route("/api/history")
def scrape_history():
    history = ScrapeHistoryManager.get_history()
    return jsonify(history)

@app.route("/api/download/<session_id>")
def download_file(session_id):
    worker = active_jobs.get(session_id)
    if worker and worker.output_path and os.path.exists(worker.output_path):
        return send_file(worker.output_path, as_attachment=True)
        
    session_info = ScrapeHistoryManager.get_session(session_id)
    if session_info and session_info.get("output_path"):
        path = session_info.get("output_path")
        if os.path.exists(path):
            return send_file(path, as_attachment=True)
            
    return jsonify({"error": "Excel file not found"}), 404

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
