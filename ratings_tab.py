"""
Ratings Scraper Tab — Platform Sub-Tabs
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Refactored Ratings Scraper with platform-specific sub-tabs:
  - Quick Search (auto-detect)
  - Booking.com
  - MMT (MakeMyTrip)
  - Agoda
  - Expedia

Each sub-tab drives which platform's scraper is used.
All common controls (bulk input, buttons, progress, log) are shared.
"""

import sys, os, io, csv, time, re, threading, pickle, json
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from playwright.sync_api import sync_playwright

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QProgressBar, QTextEdit, QSpinBox, QLineEdit,
    QFrame, QTabWidget, QComboBox, QGroupBox, QGridLayout, QCheckBox,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from ratings_platforms import (
    AVAILABLE_PLATFORMS, detect_input_type, get_platform,
    extract_rating_review_count, _get_headless_browser,
    clean_booking_url, search_booking_hotel,
    scrape_hotel, search_and_scrape, mmt_login, mmt_has_session,
    scrape_mmt_hotel, _start_mmt_chrome, scrape_goibibo_hotel,
    get_shared_browser, close_shared_browser, notify_complete,
    save_checkpoint, load_checkpoint, clear_checkpoint,
    COOKIES_DIR, MMT_COOKIES,
)
from agent_overlay import DeepResearchWorker

# ── Platform Descriptions ─────────────────────────────────

PLATFORM_DESCRIPTIONS = {
    'quick': {
        'title': 'Quick Search',
        'desc': 'Auto-detect: paste any hotel link (Booking, MMT, Agoda, Expedia) or just a hotel name.',
        'input_placeholder': 'Paste hotel link, name, or ID (auto-detects platform)...',
        'bulk_placeholder': "Paste links or hotel names here (one per line)...\nAccepts: Booking.com / MMT / Agoda / Expedia URLs, FH IDs, or hotel names",
    },
    'booking': {
        'title': 'Booking.com',
        'desc': 'Scrapes Booking.com hotel ratings. Runs in headless mode — no login needed. Accepts URLs or hotel names.',
        'input_placeholder': 'Paste Booking.com URL or hotel name...',
        'bulk_placeholder': "Paste Booking.com links or hotel names (one per line)...\nAlso accepts hotel name, city format.",
    },
    'mmt': {
        'title': 'MMT (MakeMyTrip)',
        'desc': 'Scrapes MMT hotel ratings. Uses visible browser with saved login session. Accepts FH IDs or MMT URLs.',
        'input_placeholder': 'Paste MMT hotel URL or FH ID (e.g. 32775)...',
        'bulk_placeholder': "Paste MMT links or FH IDs (one per line)...\nRequires an active MMT login session. Click 'Login to MMT' first.",
    },
    'goibibo': {
        'title': 'Goibibo',
        'desc': 'Scrapes Goibibo hotel ratings. Uses Chrome remote debugging — no login needed. Accepts Goibibo URLs or hotel names.',
        'input_placeholder': 'Paste Goibibo URL or hotel name...',
        'bulk_placeholder': "Paste Goibibo links or hotel names (one per line)...\nAlso accepts hotel name, city format.",
    },
    'agoda': {
        'title': 'Agoda',
        'desc': 'Scrapes Agoda hotel ratings. Runs in headless mode — no login needed. Accepts Agoda URLs or hotel names.',
        'input_placeholder': 'Paste Agoda URL or hotel name...',
        'bulk_placeholder': "Paste Agoda links or hotel names (one per line)...\nAlso accepts hotel name, city format.",
    },
    'expedia': {
        'title': 'Expedia',
        'desc': 'Scrapes Expedia hotel ratings. Runs in headless mode — no login needed. Accepts Expedia URLs or hotel names.',
        'input_placeholder': 'Paste Expedia URL or hotel name...',
        'bulk_placeholder': "Paste Expedia links or hotel names (one per line)...\nAlso accepts hotel name, city format.",
    },
}


# ── Platform Info Widget (shown inside each sub-tab) ─────

class PlatformInfoWidget(QWidget):
    """Shows platform-specific info inside a sub-tab."""

    def __init__(self, platform_key):
        super().__init__()
        info = PLATFORM_DESCRIPTIONS.get(platform_key, PLATFORM_DESCRIPTIONS['quick'])
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        title = QLabel(info['title'])
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setStyleSheet("color: #e94560;")
        layout.addWidget(title)

        desc = QLabel(info['desc'])
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #aaa; font-size: 12px; padding: 4px 0 8px 0;")
        layout.addWidget(desc)

        # Platform-specific details
        if platform_key == 'booking':
            details = QGroupBox("How it works")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("✓ Headless — no browser window opens"))
            d_layout.addWidget(QLabel("✓ No login required"))
            d_layout.addWidget(QLabel("✓ Accepts: booking.com/hotel/... URLs or hotel names"))
            d_layout.addWidget(QLabel("✓ Rating scale: /10"))
            d_layout.addWidget(QLabel("✓ Can search by hotel name + city if URL redirects"))
            layout.addWidget(details)

        elif platform_key == 'mmt':
            details = QGroupBox("How it works")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("⚠ Visible browser — Chrome opens to load MMT pages"))
            d_layout.addWidget(QLabel("⚠ Requires an active MMT login session"))
            login_status = "✅ Session active" if mmt_has_session() else "❌ No session — click 'Login to MMT'"
            d_layout.addWidget(QLabel(login_status))
            d_layout.addWidget(QLabel("✓ Accepts: MMT FH IDs (e.g. 32775) or full MMT URLs"))
            d_layout.addWidget(QLabel("✓ Rating scale: /5"))
            layout.addWidget(details)

        elif platform_key == 'goibibo':
            details = QGroupBox("How it works")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("✓ Runs in Chrome remote debugging mode to avoid protocol blocking"))
            d_layout.addWidget(QLabel("✓ No login required"))
            d_layout.addWidget(QLabel("✓ Accepts: goibibo.com/hotels/... URLs or hotel names"))
            d_layout.addWidget(QLabel("✓ Rating scale: /5"))
            layout.addWidget(details)

        elif platform_key == 'agoda':
            details = QGroupBox("How it works")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("✓ Headless — no browser window opens"))
            d_layout.addWidget(QLabel("✓ No login required"))
            d_layout.addWidget(QLabel("✓ Accepts: Agoda URLs or hotel names"))
            d_layout.addWidget(QLabel("✓ Rating scale: /10"))
            d_layout.addWidget(QLabel("ℹ Basic implementation — may need adjustments for Agoda's page structure"))
            layout.addWidget(details)

        elif platform_key == 'expedia':
            details = QGroupBox("How it works")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("✓ Headless — no browser window opens"))
            d_layout.addWidget(QLabel("✓ No login required"))
            d_layout.addWidget(QLabel("✓ Accepts: Expedia URLs or hotel names"))
            d_layout.addWidget(QLabel("✓ Rating scale: /10"))
            d_layout.addWidget(QLabel("ℹ Basic implementation — may need adjustments for Expedia's page structure"))
            layout.addWidget(details)

        elif platform_key == 'quick':
            details = QGroupBox("Auto-detection rules")
            d_layout = QVBoxLayout(details)
            d_layout.addWidget(QLabel("✓ booking.com/hotel/... → Booking.com"))
            d_layout.addWidget(QLabel("✓ makemytrip.com/hotelid=... → MMT"))
            d_layout.addWidget(QLabel("✓ goibibo.com/... → Goibibo"))
            d_layout.addWidget(QLabel("✓ agoda.com/... → Agoda"))
            d_layout.addWidget(QLabel("✓ expedia.com/... → Expedia"))
            d_layout.addWidget(QLabel("✓ Numeric FH ID → MMT"))
            d_layout.addWidget(QLabel("✓ Plain text → search by name"))
            layout.addWidget(details)

        layout.addStretch()


# ── Ratings Tab (Parent Widget) ──────────────────────────

class RatingsTab(QWidget):
    """The full Ratings Scraper tab with platform sub-tabs and shared controls."""

    log_signal = pyqtSignal(str)
    ui_signal = pyqtSignal(object)

    def __init__(self, csv_path=None, workers=10):
        super().__init__()
        self.csv_path = csv_path
        self.items = []
        self.original_rows = []
        self.original_headers = []
        self.output_path = None
        self.worker = None
        self._active_platform = 'quick'  # Currently selected sub-tab

        self._build_ui(workers)

        self.log_signal.connect(self.log.append)
        self.ui_signal.connect(lambda fn: fn())

        if csv_path:
            self.load_csv(csv_path)

    def _build_ui(self, workers):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(15, 10, 15, 10)

        # ── Title ───────────────────────────────────────────
        title = QLabel("Hotel Ratings & Reviews Scraper")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # ── Platform Sub-Tabs ───────────────────────────────
        self.platform_tabs = QTabWidget()
        self.platform_tabs.setStyleSheet("""
            QTabWidget::pane { border: none; background: #1a1a2e; }
            QTabBar::tab { background: #0f3460; color: #888; padding: 6px 16px;
                          margin-right: 2px; border-top-left-radius: 4px;
                          border-top-right-radius: 4px; font-size: 11px; }
            QTabBar::tab:selected { background: #16213e; color: white; }
            QTabBar::tab:hover { background: #1a3a6a; }
        """)

        for key in ['quick', 'booking', 'mmt', 'goibibo', 'agoda', 'expedia']:
            self.platform_tabs.addTab(PlatformInfoWidget(key), PLATFORM_DESCRIPTIONS[key]['title'])

        layout.addWidget(self.platform_tabs)

        # ── Quick Search Bar (always visible) ───────────────
        quick_row = QHBoxLayout()
        self.quick_input = QLineEdit()
        self.quick_input.setPlaceholderText(
            "Type hotel name or paste a link (Booking.com, MMT, etc.)..."
        )
        self.quick_input.setStyleSheet(
            "background: #16213e; color: white; border: 1px solid #444; "
            "border-radius: 6px; padding: 10px; font-size: 13px;"
        )
        self.quick_input.returnPressed.connect(self.quick_search)
        quick_row.addWidget(self.quick_input)

        self.quick_btn = QPushButton("Search")
        self.quick_btn.clicked.connect(self.quick_search)
        self.quick_btn.setStyleSheet(
            "background-color: #e94560; font-weight: bold; padding: 10px 20px;"
        )
        quick_row.addWidget(self.quick_btn)
        layout.addLayout(quick_row)

        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setStyleSheet("color: #333;")
        layout.addWidget(separator)

        bulk_label = QLabel("— Bulk scrape: paste links/names OR load CSV —")
        bulk_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        bulk_label.setStyleSheet("color: #555; font-size: 11px;")
        layout.addWidget(bulk_label)

        # ── Bulk Input ──────────────────────────────────────
        self.bulk_input = QTextEdit()
        self.bulk_input.setPlaceholderText(PLATFORM_DESCRIPTIONS['quick']['bulk_placeholder'])
        self.bulk_input.setMaximumHeight(70)
        self.bulk_input.setStyleSheet(
            "background: #16213e; color: white; border: 1px solid #444; "
            "border-radius: 6px; font-size: 12px;"
        )
        layout.addWidget(self.bulk_input)

        # ── Button Row 1 ────────────────────────────────────
        btn_row = QHBoxLayout()
        self.browse_btn = QPushButton("Browse CSV")
        self.browse_btn.clicked.connect(self.browse_file)
        btn_row.addWidget(self.browse_btn)

        self.sample_btn = QPushButton("Download Sample")
        self.sample_btn.clicked.connect(self.download_sample_csv)
        self.sample_btn.setStyleSheet("background-color: #3498db; font-weight: bold;")
        btn_row.addWidget(self.sample_btn)
        
        self.find_links_btn = QPushButton("Find Frontend Links")
        self.find_links_btn.clicked.connect(self.find_frontend_links)
        self.find_links_btn.setStyleSheet("background-color: #8e44ad; font-weight: bold;")
        self.find_links_btn.setToolTip("Resolve bulk hotel names into URLs")
        btn_row.addWidget(self.find_links_btn)

        self.deep_extract_cb = QCheckBox("Deep Extract MMT ID (Slower)")
        self.deep_extract_cb.setStyleSheet("color: #ccc; font-size: 11px;")
        btn_row.addWidget(self.deep_extract_cb)

        self.find_parallel_cb = QCheckBox("Find Parallel Listings")
        self.find_parallel_cb.setStyleSheet("color: #ccc; font-size: 11px;")
        self.find_parallel_cb.setToolTip("Find duplicate/parallel listings using Lat & Long coordinates")
        btn_row.addWidget(self.find_parallel_cb)

        btn_row.addWidget(QLabel("Workers:"))
        self.worker_spin = QSpinBox()
        self.worker_spin.setRange(1, 100)
        self.worker_spin.setValue(workers)
        self.worker_spin.setStyleSheet(
            "background: #16213e; color: white; border: 1px solid #333; padding: 4px;"
        )
        btn_row.addWidget(self.worker_spin)

        self.start_btn = QPushButton("Start")
        self.start_btn.clicked.connect(self.start_scraping)
        self.start_btn.setStyleSheet("background-color: #e94560; font-weight: bold;")
        btn_row.addWidget(self.start_btn)

        self.pause_btn = QPushButton("Pause")
        self.pause_btn.clicked.connect(self.pause_scraping)
        self.pause_btn.setStyleSheet("background-color: #f5a623; font-weight: bold;")
        self.pause_btn.setEnabled(False)
        btn_row.addWidget(self.pause_btn)

        self.stop_btn = QPushButton("Stop")
        self.stop_btn.clicked.connect(self.stop_scraping)
        self.stop_btn.setStyleSheet("background-color: #c0392b; font-weight: bold;")
        self.stop_btn.setEnabled(False)
        btn_row.addWidget(self.stop_btn)

        self.settings_btn = QPushButton("⚙ Settings")
        self.settings_btn.clicked.connect(self.show_settings)
        self.settings_btn.setStyleSheet("background-color: #4a5568; font-weight: bold;")
        btn_row.addWidget(self.settings_btn)

        layout.addLayout(btn_row)

        # ── Button Row 2 ────────────────────────────────────
        btn_row2 = QHBoxLayout()
        self.download_btn = QPushButton("Download CSV")
        self.download_btn.clicked.connect(self.download_csv)
        self.download_btn.setStyleSheet("background-color: #27ae60; font-weight: bold;")
        self.download_btn.setEnabled(False)
        btn_row2.addWidget(self.download_btn)

        self.clear_btn = QPushButton("Clear")
        self.clear_btn.clicked.connect(self.clear_all)
        self.clear_btn.setStyleSheet("background-color: #555; font-weight: bold;")
        btn_row2.addWidget(self.clear_btn)

        layout.addLayout(btn_row2)

        # ── MMT Session Row ─────────────────────────────────
        mmt_row = QHBoxLayout()
        mmt_status = "MMT session active" if mmt_has_session() else "MMT: No session"
        self.mmt_label = QLabel(mmt_status)
        self.mmt_label.setStyleSheet("color: #888; font-size: 11px;")
        mmt_row.addWidget(self.mmt_label)
        self.mmt_login_btn = QPushButton("Login to MMT")
        self.mmt_login_btn.setStyleSheet(
            "background-color: #0a7; padding: 8px 16px; font-size: 12px;"
        )
        self.mmt_login_btn.clicked.connect(self.do_mmt_login)
        mmt_row.addWidget(self.mmt_login_btn)
        layout.addLayout(mmt_row)

        # ── Progress Bar ────────────────────────────────────
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # ── Log Area ────────────────────────────────────────
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumHeight(140)
        layout.addWidget(self.log)

        # Connect sub-tab change signal now that all widgets (bulk_input, quick_input) are initialized
        self.platform_tabs.currentChanged.connect(self._on_platform_changed)

    def _on_platform_changed(self, index):
        """Update bulk input placeholder and active platform when sub-tab changes."""
        keys = ['quick', 'booking', 'mmt', 'goibibo', 'agoda', 'expedia']
        key = keys[index] if index < len(keys) else 'quick'
        self._active_platform = key
        info = PLATFORM_DESCRIPTIONS.get(key, PLATFORM_DESCRIPTIONS['quick'])
        if hasattr(self, 'bulk_input'):
            self.bulk_input.setPlaceholderText(info['bulk_placeholder'])
        if hasattr(self, 'quick_input'):
            self.quick_input.setPlaceholderText(info['input_placeholder'])

    # ── Input Parsing ──────────────────────────────────────

    def _detect_and_parse(self, line):
        """Parse a single line of input based on the active platform."""
        line = line.strip()
        if not line:
            return None

        # First try auto-detection
        detected = detect_input_type(line)

        # If a specific platform sub-tab is selected (not 'quick'), override
        platform = self._active_platform
        if platform != 'quick':
            # Force platform
            detected['platform'] = platform
            if platform == 'mmt' and detected['type'] == 'unknown':
                detected['type'] = 'id'
                detected['hotel_id'] = line.replace('#', '').strip()
            elif platform == 'booking' and detected['type'] == 'unknown':
                detected['type'] = 'name'
                detected['name'] = line

        return detected

    def quick_search(self):
        query = self.quick_input.text().strip()
        if not query:
            return
        self.quick_btn.setEnabled(False)
        self.quick_input.setEnabled(False)
        self.log.append(f"\nSearching: {query}...")

        detected = self._detect_and_parse(query)
        if not detected:
            self.log.append("Could not parse input")
            self.quick_btn.setEnabled(True)
            self.quick_input.setEnabled(True)
            return

        def do_search():
            try:
                if (detected['platform'] == 'mmt' or 'makemytrip.com' in query) and (detected.get('hotel_id') or 'hotelId=' in query):
                    hid = detected.get('hotel_id')
                    if not hid:
                        m_hid = re.search(r'hotelId=(\d+)', query)
                        if m_hid:
                            hid = m_hid.group(1)
                    if mmt_has_session():
                        rating, count = scrape_mmt_hotel(hid)
                        self.ui_signal.emit(lambda: self._show_result(
                            query, rating, count, 'mmt', detected.get('url') or query
                        ))
                    else:
                        self.log_signal.emit("MMT requires login first. Click 'Login to MMT' button.")
                elif detected['platform'] == 'goibibo' and detected.get('url'):
                    rating, count = scrape_goibibo_hotel(detected['url'])
                    self.ui_signal.emit(lambda: self._show_result(
                        query, rating, count, 'goibibo', detected.get('url')
                    ))
                elif detected['platform'] in ('booking', 'agoda', 'expedia') or 'booking.com' in (detected.get('url') or ''):
                    url = detected.get('url', '')
                    name = detected.get('name', query)
                    city = detected.get('city', '')
                    if url and 'booking.com' in url:
                        rating, count, _ = scrape_hotel(url, name, city)
                        self.ui_signal.emit(lambda: self._show_result(
                            query, rating, count, 'booking', url
                        ))
                    elif url and ('agoda.com' in url or 'expedia.com' in url):
                        page = None
                        try:
                            browser = _get_headless_browser()
                            page = browser.new_page()
                            page.goto(url, timeout=20000, wait_until="domcontentloaded")
                            page.wait_for_timeout(3000)
                            content = page.content()
                            rating, count = extract_rating_review_count(content, scale_10=True)
                            self.ui_signal.emit(lambda: self._show_result(
                                query, rating, count,
                                'agoda' if 'agoda.com' in url else 'expedia',
                                url
                            ))
                        except:
                            self.log_signal.emit(f"Could not scrape: {url}")
                        finally:
                            if page:
                                page.close()
                    else:
                        rating, count, found_url, _ = search_and_scrape(name, city)
                        self.ui_signal.emit(lambda: self._show_result(
                            name, rating, count, 'booking', found_url
                        ))
                elif detected['type'] == 'name' or (detected['platform'] in ('booking', 'mmt', 'goibibo', 'agoda', 'expedia') and not detected.get('hotel_id') and not detected.get('url')):
                    name = detected.get('name', query)
                    city = detected.get('city', '')
                    target_platform = detected.get('platform') or self._active_platform
                    
                    if target_platform == 'mmt':
                        if mmt_has_session():
                            # Resolve hotel name query using MMT platform scraper which supports query searching
                            rating, count = scrape_mmt_hotel(name) # Will trigger MMT search internally
                            self.ui_signal.emit(lambda: self._show_result(
                                name, rating, count, 'mmt', f"MMT Search: {name}"
                            ))
                        else:
                            self.log_signal.emit("MMT requires login first. Click 'Login to MMT' button.")
                    elif target_platform == 'goibibo':
                        rating, count = scrape_goibibo_hotel('', name, city)
                        self.ui_signal.emit(lambda: self._show_result(
                            name, rating, count, 'goibibo', f"Goibibo Search: {name}"
                        ))
                    elif target_platform == 'agoda':
                        # Headless Agoda search
                        page = None
                        try:
                            browser = _get_headless_browser()
                            page = browser.new_page()
                            url = f"https://www.agoda.com/search?text={name.replace(' ', '+')}"
                            page.goto(url, timeout=25000, wait_until="domcontentloaded")
                            page.wait_for_timeout(3000)
                            content = page.content()
                            rating, count = extract_rating_review_count(content, scale_10=True)
                            self.ui_signal.emit(lambda: self._show_result(name, rating, count, 'agoda', url))
                        except Exception as e:
                            self.log_signal.emit(f"Agoda search failed: {e}")
                        finally:
                            if page: page.close()
                    elif target_platform == 'expedia':
                        # Headless Expedia search
                        page = None
                        try:
                            browser = _get_headless_browser()
                            page = browser.new_page()
                            url = f"https://www.expedia.com/hotels/search?text={name.replace(' ', '+')}"
                            page.goto(url, timeout=25000, wait_until="domcontentloaded")
                            page.wait_for_timeout(3000)
                            content = page.content()
                            rating, count = extract_rating_review_count(content, scale_10=True)
                            self.ui_signal.emit(lambda: self._show_result(name, rating, count, 'expedia', url))
                        except Exception as e:
                            self.log_signal.emit(f"Expedia search failed: {e}")
                        finally:
                            if page: page.close()
                    else:
                        # Default Booking.com search
                        rating, count, found_url, _ = search_and_scrape(name, city)
                        self.ui_signal.emit(lambda: self._show_result(
                            name, rating, count, 'booking', found_url
                        ))
                else:
                    self.log_signal.emit(f"Could not determine how to scrape: {query}")
            except Exception as e:
                self.log_signal.emit(f"Error: {e}")
            self.ui_signal.emit(lambda: (
                self.quick_btn.setEnabled(True),
                self.quick_input.setEnabled(True)
            ))

        threading.Thread(target=do_search, daemon=True).start()

    def _show_result(self, query, rating, count, source, url=None):
        """Display search result in the log."""
        self.quick_btn.setEnabled(True)
        self.quick_input.setEnabled(True)
        if rating:
            scale = "/5" if source == "mmt" else "/10"
            self.log.append(f"  {query}")
            self.log.append(f"  Rating: {rating}{scale} | Reviews: {count or 'N/A'}")
            self.log.append(f"  Source: {'MakeMyTrip' if source == 'mmt' else source.title()}")
            if url and url != query:
                self.log.append(f"  Found at: {url[:80]}")
        else:
            self.log.append(f"  Could not find ratings for: {query}")
        self.log.append("")

    # ── Frontend Link Finder ──────────────────────────────────
    
    def find_frontend_links(self):
        text = ""
        if self.items:
            # If a CSV was loaded, use the names from self.items
            # Fallback to whatever URL or hotel_id if name is missing
            query_list = []
            for item in self.items:
                target = item.get('name') or item.get('url') or item.get('hotel_id') or ""
                query_list.append(target)
            text = "\n".join(query_list)
        else:
            # Otherwise use the pasted text
            text = self.bulk_input.toPlainText().strip()
            
        if not text:
            self.log.append("❌ Please paste hotel names or load a CSV first!")
            return
            
        self.find_links_btn.setEnabled(False)
        self.browse_btn.setEnabled(False)
        self.start_btn.setEnabled(False)
        self.bulk_input.clear()
        
        self.resolved_results = [None] * len(self.items) if self.items else []
        
        # Check for Gemini API key
        import os
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            from PyQt6.QtWidgets import QInputDialog
            key, ok = QInputDialog.getText(self, "AI Verification Setup", 
                                           "To enable AI Verification for better accuracy, please enter a Google Gemini API Key (or leave blank to skip):")
            if ok and key.strip():
                os.environ["GEMINI_API_KEY"] = key.strip()

        plat = self._active_platform
        if plat == 'quick':
            plat = 'any'
            
        deep_extract = self.deep_extract_cb.isChecked()
        find_parallel = self.find_parallel_cb.isChecked()
            
        self.research_worker = DeepResearchWorker(text, plat, deep_extract=deep_extract, items_context=self.items, find_parallel=find_parallel)
        self.research_worker.signals.log.connect(self.log.append)
        self.research_worker.signals.finished.connect(self._on_frontend_link_found)
        self.research_worker.start()

    def _on_frontend_link_found(self, result):
        if 'error' in result:
            if result.get('error') == 'No valid queries':
                self.find_links_btn.setEnabled(True)
                self.browse_btn.setEnabled(True)
                self.start_btn.setEnabled(True)
                self.research_worker = None
            return
            
        if result.get('batch_finished'):
            self.find_links_btn.setEnabled(True)
            self.browse_btn.setEnabled(True)
            self.start_btn.setEnabled(True)
            self.research_worker = None
            
            # Prompt user to download CSV
            current_text = self.bulk_input.toPlainText().strip()
            if current_text or self.resolved_results:
                from PyQt6.QtWidgets import QMessageBox
                reply = QMessageBox.question(
                    self, "Download Links",
                    "Frontend links have been extracted!\n\nWould you like to download the updated CSV (preserving all input columns)?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.Yes:
                    path, _ = QFileDialog.getSaveFileName(self, "Save Resolved CSV", "resolved_properties.csv", "CSV Files (*.csv)")
                    if path:
                        try:
                            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                                import csv
                                writer = csv.writer(f)
                                headers = list(self.original_headers)
                                url_header = "Resolved URL"
                                id_header = "Resolved Hotel ID"
                                
                                if url_header not in headers:
                                    headers.append(url_header)
                                if id_header not in headers:
                                    headers.append(id_header)
                                
                                writer.writerow(headers)
                                
                                for idx, row in enumerate(self.original_rows):
                                    new_row = list(row)
                                    resolved = self.resolved_results[idx] if idx < len(self.resolved_results) else None
                                    resolved_url = ""
                                    resolved_hid = ""
                                    if resolved:
                                        url_val = resolved.get('url', '')
                                        if '\n' in url_val:
                                            resolved_url = url_val.replace('\n', ', ')
                                        else:
                                            parts = url_val.split('|')
                                            resolved_url = parts[0].strip()
                                        resolved_hid = resolved.get('hotel_id', '')
                                    new_row.append(resolved_url)
                                    new_row.append(resolved_hid)
                                    writer.writerow(new_row)
                                    
                            QMessageBox.information(self, "Success", f"CSV successfully saved to:\n{path}")
                        except Exception as e:
                            QMessageBox.warning(self, "Error", f"Failed to save CSV:\n{e}")
            return
            
        # Store resolved result at query index
        idx = result.get('query_index')
        if idx is not None and idx < len(self.resolved_results):
            self.resolved_results[idx] = result

        # Append the resolved link to the bulk_input window
        current = self.bulk_input.toPlainText().strip()
        new_link = result['url']
        if current:
            self.bulk_input.setPlainText(f"{current}\n{new_link}")
        else:
            self.bulk_input.setPlainText(new_link)

    # ── CSV Loading ─────────────────────────────────────────

    def download_sample_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Sample CSV", "sample_ratings_scraper.csv", "CSV Files (*.csv)")
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Hotel Name", "City", "FHID", "URL"])
                writer.writerows([
                    ["FabHotel Raj Villa", "Indore", "1234", "http://booking.com/..."],
                    ["FabHotel The Corporate", "Mumbai", "", ""]
                ])
            QMessageBox.information(self, "Success", f"Sample CSV saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save CSV:\n{e}")

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 
            "Select Input File", 
            "", 
            "Supported Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.load_csv(path)

    def load_csv(self, path):
        self.csv_path = path
        self.items = []
        self.original_rows = []
        self.original_headers = []
        try:
            if path.lower().endswith(('.xlsx', '.xls')):
                rows = []
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, data_only=True)
                    ws = wb.active
                    for r in ws.iter_rows(values_only=True):
                        # Convert None to empty string, and convert other values to string to mimic csv.reader
                        row_vals = [str(cell) if cell is not None else '' for cell in r]
                        rows.append(row_vals)
                    wb.close()
                except Exception as ex:
                    try:
                        import pandas as pd
                        df = pd.read_excel(path)
                        headers = [str(c) for c in df.columns]
                        rows = [headers]
                        for r in df.values:
                            rows.append([str(c) if pd.notna(c) else '' for c in r])
                    except Exception as ex2:
                        raise Exception(f"Failed to read Excel file: {ex} (pandas fallback error: {ex2})")
            else:
                with open(path, newline='', encoding='utf-8') as f:
                    rows = list(csv.reader(f))

            header_idx = 0
            headers = []
            for i, row in enumerate(rows[:5]):
                lower_row = [c.lower().strip() for c in row]
                if 'name' in lower_row or 'hotel name' in lower_row:
                    header_idx = i
                    headers = [c.strip() for c in row]
                    break

            if not headers:
                headers = [c.strip() for c in rows[0]]

            lower_headers = [h.lower() for h in headers]

            def find_col(*names):
                for n in names:
                    if len(n) <= 3:
                        for i, h in enumerate(lower_headers):
                            if re.search(r'(?<![a-z])' + re.escape(n) + r'(?![a-z])', h):
                                return i
                    else:
                        for i, h in enumerate(lower_headers):
                            if n in h:
                                return i
                return None

            # Strict name column match (exclude ID, Code, Link, URL, and FH columns)
            name_idx = None
            for i, h in enumerate(lower_headers):
                if ('name' in h or 'hotel' in h) and not any(x in h for x in ('code', 'id', 'link', 'url', 'fh')):
                    name_idx = i
                    break
            
            if name_idx is None:
                name_idx = find_col('name', 'hotel')

            city_idx = find_col('city', 'location')
            link_idx = find_col('booking', 'mmt', 'goibibo', 'gi', 'agoda', 'expedia', 'link', 'url')
            id_idx = find_col('fhid', 'fh id', 'fh', 'front-end id', 'hotel code', 'hotel id', 'hotel_id', 'code', 'id')
            mmt_id_idx = find_col('mmt id', 'mmt_id', 'makemytrip id')
            bcom_id_idx = find_col('b.com id', 'booking id', 'booking.com id', 'bcom id')
            address_idx = find_col('address', 'addr')
            lat_idx = find_col('latitude', 'lat')
            lon_idx = find_col('longitude', 'lon', 'lng', 'long')
            zip_idx = find_col('zipcode', 'zip', 'pin', 'pincode', 'postal')

            if link_idx is not None and header_idx + 1 < len(rows):
                test_row = rows[header_idx + 1]
                if link_idx < len(test_row) and 'http' not in test_row[link_idx]:
                    for i, cell in enumerate(test_row):
                        if any(domain in cell.lower() for domain in ('makemytrip.com', 'booking.com', 'goibibo.com', 'agoda.com', 'expedia.com')):
                            link_idx = i
                            break

            self.original_headers = headers

            for row in rows[header_idx + 1:]:
                if len(row) <= (name_idx or 0):
                    continue
                name = row[name_idx].strip() if name_idx is not None and name_idx < len(row) else ''
                city = row[city_idx].strip() if city_idx is not None and city_idx < len(row) else ''
                url = row[link_idx].strip() if link_idx is not None and link_idx < len(row) else ''
                hotel_id = row[id_idx].strip() if id_idx is not None and id_idx < len(row) else ''
                mmt_id = row[mmt_id_idx].strip() if mmt_id_idx is not None and mmt_id_idx < len(row) else ''
                bcom_id = row[bcom_id_idx].strip() if bcom_id_idx is not None and bcom_id_idx < len(row) else ''
                address = row[address_idx].strip() if address_idx is not None and address_idx < len(row) else ''
                latitude = row[lat_idx].strip() if lat_idx is not None and lat_idx < len(row) else ''
                longitude = row[lon_idx].strip() if lon_idx is not None and lon_idx < len(row) else ''
                zipcode = row[zip_idx].strip() if zip_idx is not None and zip_idx < len(row) else ''
                raw_line = ' '.join(row).strip()

                # Ensure we have at least one valid identifier (name, url, or hotel_id)
                if not name and not url and not hotel_id:
                    continue

                self.original_rows.append(row)

                item = {
                    'name': name,
                    'city': city,
                    'url': url,
                    'source': 'search',
                    'hotel_id': hotel_id,
                    'mmt_id': mmt_id,
                    'bcom_id': bcom_id,
                    'address': address,
                    'latitude': latitude,
                    'longitude': longitude,
                    'zipcode': zipcode
                }

                # Prioritization logic:
                # 1. Hotel Name is present
                if name and name.lower() not in ('name', 'hotel name', 'hotel'):
                    # Check if the URL column contains a specific platform link
                    if url and 'booking.com' in url.lower():
                        item['source'] = 'booking'
                    elif url and 'makemytrip.com' in url.lower():
                        m = re.search(r'hotelId=(\w+)', url)
                        if m:
                            item['hotel_id'] = m.group(1)
                        item['source'] = 'mmt'
                    elif url and 'goibibo.com' in url.lower():
                        item['source'] = 'goibibo'
                    elif url and 'agoda.com' in url.lower():
                        item['source'] = 'agoda'
                    elif url and 'expedia.com' in url.lower():
                        item['source'] = 'expedia'
                    else:
                        item['source'] = 'search'
                    self.items.append(item)

                # 2. Name is NOT present, but URL is present
                elif url:
                    detected = detect_input_type(url)
                    extracted_name = detected.get('name') or url[:50]
                    item['name'] = extracted_name
                    if detected['platform'] == 'mmt':
                        item['hotel_id'] = detected.get('hotel_id') or hotel_id
                        item['source'] = 'mmt'
                    elif detected['platform'] in ('booking', 'agoda', 'expedia', 'goibibo'):
                        item['source'] = detected['platform']
                    else:
                        item['source'] = 'search'
                    self.items.append(item)

                # 3. Only ID is present
                elif hotel_id:
                    clean_id = hotel_id.replace('#', '').strip()
                    item['name'] = clean_id
                    item['source'] = 'mmt'
                    item['hotel_id'] = clean_id
                    self.items.append(item)

            has_links = sum(1 for i in self.items if i['url'] and 'http' in i['url'])
            names_only = len(self.items) - has_links
            self.bulk_input.setPlainText(
                f"Loaded: {Path(path).name} — {len(self.items)} hotels "
                f"({has_links} links, {names_only} name-only)"
            )
            self.log.append(f"Loaded {len(self.items)} hotels from {Path(path).name}")
        except Exception as e:
            self.log.append(f"ERROR loading CSV: {e}")

    # ── Scraping Lifecycle ──────────────────────────────────

    def start_scraping(self):
        # Parse paste area if CSV wasn't loaded
        bulk_text = self.bulk_input.toPlainText().strip()
        if bulk_text and not self.items:
            lines = [l.strip() for l in bulk_text.split('\n') if l.strip()]
            self.items = []
            for line in lines:
                parts = [p.strip() for p in line.split('|')]
                url_part = parts[0]
                hid_part = parts[1] if len(parts) > 1 else ''
                
                detected = detect_input_type(url_part)
                if not detected:
                    continue
                if detected['platform'] == 'mmt':
                    self.items.append({
                        'name': detected.get('name', url_part[:50]),
                        'city': detected.get('city', ''),
                        'url': url_part if 'http' in url_part else '',
                        'source': 'mmt',
                        'hotel_id': detected.get('hotel_id') or hid_part
                    })
                elif detected['platform'] in ('booking', 'goibibo', 'agoda', 'expedia'):
                    self.items.append({
                        'name': detected.get('name', url_part[:50]),
                        'city': detected.get('city', ''),
                        'url': url_part if 'http' in url_part else '',
                        'source': detected['platform'],
                        'hotel_id': hid_part
                    })
                else:
                    self.items.append({
                        'name': url_part, 'city': '', 'url': '', 'source': 'search', 'hotel_id': hid_part
                    })

        if not self.items:
            self.log.append("Nothing to scrape. Paste links/names above or load a CSV.")
            return

        # Filter and adapt by active platform (if not 'quick')
        if self._active_platform != 'quick':
            kept_indices = []
            for i, item in enumerate(self.items):
                # Respect specific platform URLs/sources already detected or loaded
                if item.get('source') in ('mmt', 'goibibo', 'booking', 'agoda', 'expedia'):
                    # The source has already been correctly resolved via CSV headers or URLs
                    kept_indices.append(i)
                    continue

                if item.get('url') and any(domain in item['url'].lower() for domain in ('booking.com', 'makemytrip.com', 'goibibo.com', 'agoda.com', 'expedia.com')):
                    detected = detect_input_type(item['url'])
                    if detected['platform']:
                        item['source'] = detected['platform']
                        if detected.get('hotel_id'):
                            item['hotel_id'] = detected['hotel_id']
                    kept_indices.append(i)
                    continue

                # 1. Booking.com Tab
                if self._active_platform == 'booking':
                    if item.get('url') and 'booking.com' in item['url']:
                        item['source'] = 'booking'
                        kept_indices.append(i)
                    elif item.get('name'):
                        item['source'] = 'search'
                        kept_indices.append(i)

                # 2. MakeMyTrip Tab
                elif self._active_platform == 'mmt':
                    if item.get('hotel_id'):
                        item['source'] = 'mmt'
                        kept_indices.append(i)
                    elif item.get('url') and 'makemytrip' in item['url']:
                        m = re.search(r'hotelId=(\w+)', item['url'])
                        if m:
                            item['hotel_id'] = m.group(1)
                            item['source'] = 'mmt'
                            kept_indices.append(i)

                # 3. Goibibo, Agoda or Expedia Tabs
                elif self._active_platform in ('goibibo', 'agoda', 'expedia'):
                    if item.get('url') and self._active_platform in item['url']:
                        item['source'] = self._active_platform
                        kept_indices.append(i)
                    elif item.get('name'):
                        item['source'] = self._active_platform
                        kept_indices.append(i)

            skipped = len(self.items) - len(kept_indices)
            if skipped:
                self.log.append(
                    f"Platform filter [{PLATFORM_DESCRIPTIONS[self._active_platform]['title']}]: "
                    f"skipping {skipped} items that cannot be scraped on this platform"
                )
            self.items = [self.items[i] for i in kept_indices]
            if self.original_rows:
                self.original_rows = [self.original_rows[i] for i in kept_indices]

        if not self.items:
            self.log.append("Nothing matches the selected platform filter. Change platform or load different data.")
            self.start_btn.setEnabled(True)
            return

        # Check for checkpoint resume
        existing_results = None
        resume_output_path = None
        resume_total = 0
        if self.csv_path:
            checkpoint = load_checkpoint(self.csv_path)
            if checkpoint is not None:
                chk_results, chk_output, chk_processed, chk_total = checkpoint
                if chk_total == len(self.items) and chk_processed < chk_total:
                    existing_results = chk_results
                    resume_output_path = chk_output
                    resume_total = chk_total

        self.start_btn.setEnabled(False)
        self.browse_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(self.items))

        if existing_results is not None:
            done = sum(1 for r in existing_results if r is not None)
            self.progress.setValue(done)
            self.log.append(
                f"\nAuto-resuming from checkpoint — {done}/{resume_total} already scraped..."
            )
        else:
            self.progress.setValue(0)
            self.log.append(
                f"\nStarting scrape of {len(self.items)} hotels "
                f"with {self.worker_spin.value()} workers..."
            )

        self.worker = ScrapeWorker(
            self.items, self.worker_spin.value(),
            self.original_rows, self.original_headers,
            existing_results=existing_results,
            resume_output_path=resume_output_path,
            input_file=self.csv_path
        )
        self.worker.progress.connect(self.on_progress)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()
        self.pause_btn.setEnabled(True)
        self.stop_btn.setEnabled(True)

    def pause_scraping(self):
        if hasattr(self, 'worker') and self.worker and self.worker.isRunning():
            if self.worker._pause:
                self.worker.resume()
                self.pause_btn.setText("Pause")
                self.log.append("Resumed")
            else:
                self.worker.pause()
                self.pause_btn.setText("Resume")
                self.log.append("Paused")

    def stop_scraping(self):
        if hasattr(self, 'worker') and self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log.append("Stopping...")

    def download_csv(self):
        if self.output_path and os.path.exists(self.output_path):
            os.startfile(self.output_path)

    def clear_all(self):
        if self.csv_path:
            clear_checkpoint(self.csv_path)
        self.log.clear()
        self.bulk_input.clear()
        self.items = []
        self.original_rows = []
        self.original_headers = []
        self.csv_path = None
        self.output_path = None
        self.download_btn.setEnabled(False)
        self.progress.setVisible(False)
        self.progress.setValue(0)

    def on_progress(self, current, total, status):
        self.progress.setValue(current)
        self.log.append(f"[{current}/{total}] {status}")
        scrollbar = self.log.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def on_finished(self, output_path, success, total):
        self.progress.setValue(self.progress.maximum())
        self.log.append(f"\n{'=' * 40}")
        self.log.append(f"DONE! {success}/{total} hotels scraped successfully")
        self.log.append(f"Output: {output_path}")
        self.start_btn.setEnabled(True)
        self.browse_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.stop_btn.setEnabled(False)
        self.pause_btn.setText("Pause")
        self.output_path = output_path
        self.download_btn.setEnabled(True)
        self.bulk_input.setPlainText(
            f"Done! {success}/{total} scraped -> {Path(output_path).name}"
        )
        threading.Thread(target=notify_complete, args=(
            f"Scraping complete! {success}/{total} hotels done.\nClick Downloads to open.",
        ), daemon=True).start()

    def do_mmt_login(self):
        self.log.append("\nOpening Chrome — log in to MMT, then CLOSE Chrome...")
        self.mmt_login_btn.setEnabled(False)

        def login_thread():
            try:
                mmt_login()
                self.log_signal.emit("MMT session saved! You can now scrape MMT hotels.")
                self.ui_signal.emit(lambda: self.mmt_label.setText("MMT session active"))
            except Exception as e:
                self.log_signal.emit(f"MMT login failed: {e}")
            self.ui_signal.emit(lambda: self.mmt_login_btn.setEnabled(True))

        threading.Thread(target=login_thread, daemon=True).start()

    def show_settings(self):
        from settings_dialog import SettingsDialog
        dialog = SettingsDialog(self, on_resume_callback=self.resume_ratings_run)
        dialog.exec()

    def resume_ratings_run(self, run_data):
        input_file = run_data.get("input_file")
        current_index = run_data.get("current_index", 0)
        total_items = run_data.get("total_items", 0)
        output_file = run_data.get("output_file")
        
        if not input_file or not os.path.exists(input_file):
            self.log.append("Unable to resume: input file not found or is a manual text paste.")
            return
            
        self.load_csv(input_file)
        
        existing_results = [None] * len(self.items)
        if output_file and os.path.exists(output_file):
            try:
                with open(output_file, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    
                    # find matching headers dynamically
                    rating_idx = headers.index('Scraped_Rating') if 'Scraped_Rating' in headers else -1
                    reviews_idx = headers.index('Scraped_Reviews') if 'Scraped_Reviews' in headers else -1
                    source_idx = headers.index('Scraped_Source') if 'Scraped_Source' in headers else -1
                    fail_idx = headers.index('Scraped_Fail_Reason') if 'Scraped_Fail_Reason' in headers else -1
                    
                    if rating_idx != -1 and reviews_idx != -1 and source_idx != -1:
                        rows = list(reader)
                        for idx in range(min(len(rows), len(self.items))):
                            r_row = rows[idx]
                            if idx < len(r_row):
                                rating = r_row[rating_idx]
                                reviews = r_row[reviews_idx]
                                src = r_row[source_idx]
                                fail_reason = r_row[fail_idx] if fail_idx != -1 and fail_idx < len(r_row) else ''
                                if rating:
                                    existing_results[idx] = {
                                        'rating': rating,
                                        'review_count': reviews,
                                        'source': src,
                                        'fail_reason': fail_reason
                                    }
            except Exception as e:
                self.log.append(f"Failed to load existing results from output file for resume: {e}")
                
        self.start_btn.setEnabled(False)
        self.browse_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(self.items))
        done = sum(1 for r in existing_results if r is not None)
        self.progress.setValue(done)
        self.log.append(f"\nResuming run from index {current_index}...")
        
        self.worker = ScrapeWorker(
            self.items, self.worker_spin.value(),
            self.original_rows, self.original_headers,
            existing_results=existing_results,
            resume_output_path=output_file,
            input_file=self.csv_path or input_file
        )
        self.worker.progress.connect(self.on_progress)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()
        self.pause_btn.setEnabled(True)
        self.stop_btn.setEnabled(True)


# ── ScrapeWorker (same as app.py, imported for reuse) ────

class ScrapeWorker(QThread):
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(str, int, int)

    def __init__(self, items, num_workers, original_rows=None, original_headers=None,
                 existing_results=None, resume_output_path=None, input_file=None):
        super().__init__()
        self.items = items
        self.num_workers = num_workers
        self.original_rows = original_rows or []
        self.original_headers = original_headers or []
        self.existing_results = existing_results
        self.resume_output_path = resume_output_path
        self.input_file = input_file
        self._stop = False
        self._pause = False

    def stop(self):
        self._stop = True

    def pause(self):
        self._pause = True

    def resume(self):
        self._pause = False

    def _format_eta(self, elapsed, processed, total):
        remaining = total - processed
        if processed == 0 or remaining <= 0:
            return ""
        avg_secs = elapsed / processed
        eta_secs = avg_secs * remaining
        if eta_secs < 60:
            return f"~{int(eta_secs)}s remaining"
        elif eta_secs < 3600:
            mins = int(eta_secs // 60)
            secs = int(eta_secs % 60)
            if secs >= 30:
                mins += 1
            return f"~{mins}m remaining"
        else:
            hours = int(eta_secs // 3600)
            mins = int((eta_secs % 3600) // 60)
            return f"~{hours}h {mins}m remaining"

    def run(self):
        import time
        import os
        import csv
        from pathlib import Path
        import asyncio
        from async_api_scraper import AsyncScraperEngine
        import db_cache

        total = len(self.items)
        SAVE_INTERVAL = 50
        start_time = time.time()

        if self.existing_results is not None and self.resume_output_path:
            results = list(self.existing_results)
            processed_count = [sum(1 for r in results if r is not None)]
            output_path = self.resume_output_path
            remaining = {i for i in range(total) if results[i] is None}
            self.progress.emit(processed_count[0], total,
                               f"Resuming — {processed_count[0]}/{total} already done, {len(remaining)} remaining")
        else:
            results = [None] * total
            processed_count = [0]
            if self.input_file:
                path_obj = Path(self.input_file)
                stem = path_obj.stem
                ext = path_obj.suffix.lower()
                if ext in ['.xlsx', '.xls']:
                    output_path = str(Path.home() / "Downloads" / f"{stem}_scraped.xlsx")
                else:
                    output_path = str(Path.home() / "Downloads" / f"{stem}_scraped.csv")
            else:
                output_path = str(Path.home() / "Downloads" / f"ratings_output_{int(time.time())}.csv")
            remaining = set(range(total))
            
            if not output_path.endswith('.xlsx'):
                if self.original_rows and self.original_headers:
                    out_headers = self.original_headers + ['Scraped_Rating', 'Scraped_Reviews', 'Scraped_Source', 'Scraped_Fail_Reason']
                    with open(output_path, 'w', newline='', encoding='utf-8') as f:
                        csv.writer(f).writerow(out_headers)
                else:
                    with open(output_path, 'w', newline='', encoding='utf-8') as f:
                        w = csv.DictWriter(f, fieldnames=['name', 'city', 'url', 'rating', 'review_count', 'source', 'fail_reason'])
                        w.writeheader()

        def save_incremental():
            is_excel = output_path.endswith('.xlsx')
            if is_excel:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                if self.original_rows and self.original_headers:
                    new_cols = ['Scraped_Rating', 'Scraped_Reviews', 'Scraped_Source', 'Scraped_Fail_Reason']
                    out_headers = list(self.original_headers)
                    lower_orig = [str(h).lower() for h in self.original_headers]
                    col_indices = {}
                    for col in new_cols:
                        try:
                            col_indices[col] = lower_orig.index(col.lower())
                        except ValueError:
                            col_indices[col] = None
                            out_headers.append(col)
                    ws.append(out_headers)
                    for idx, orig_row in enumerate(self.original_rows):
                        r = results[idx] if idx < len(results) and results[idx] else {
                            'rating': '', 'review_count': '', 'source': '', 'fail_reason': ''
                        }
                        out_row = list(orig_row)
                        for col in new_cols:
                            key = 'review_count' if col == 'Scraped_Reviews' else col.replace('Scraped_', '').lower()
                            val = r.get(key, '')
                            if col_indices[col] is not None:
                                out_row[col_indices[col]] = val
                            else:
                                out_row.append(val)
                        ws.append(out_row)
                else:
                    ws.append(['name', 'city', 'url', 'rating', 'review_count', 'source', 'fail_reason'])
                    for idx, item in enumerate(self.items):
                        r = results[idx] if idx < len(results) and results[idx] else {
                            'rating': 'N/A', 'review_count': 'N/A', 'source': '', 'fail_reason': ''
                        }
                        ws.append([
                            item.get('name', ''), item.get('city', ''), item.get('url', ''),
                            r.get('rating', 'N/A'), r.get('review_count', 'N/A'), r.get('source', ''), r.get('fail_reason', '')
                        ])
                wb.save(output_path + '.tmp')
            else:
                with open(output_path + '.tmp', 'w', newline='', encoding='utf-8') as f:
                    if self.original_rows and self.original_headers:
                        new_cols = ['Scraped_Rating', 'Scraped_Reviews', 'Scraped_Source', 'Scraped_Fail_Reason']
                        out_headers = list(self.original_headers)
                        lower_orig = [str(h).lower() for h in self.original_headers]
                        col_indices = {}
                        for col in new_cols:
                            try:
                                col_indices[col] = lower_orig.index(col.lower())
                            except ValueError:
                                col_indices[col] = None
                                out_headers.append(col)
                        writer = csv.writer(f)
                        writer.writerow(out_headers)
                        for idx, orig_row in enumerate(self.original_rows):
                            r = results[idx] if idx < len(results) and results[idx] else {
                                'rating': '', 'review_count': '', 'source': '', 'fail_reason': ''
                            }
                            out_row = list(orig_row)
                            for col in new_cols:
                                key = 'review_count' if col == 'Scraped_Reviews' else col.replace('Scraped_', '').lower()
                                val = r.get(key, '')
                                if col_indices[col] is not None:
                                    out_row[col_indices[col]] = val
                                else:
                                    out_row.append(val)
                            writer.writerow(out_row)
                    else:
                        writer = csv.DictWriter(f, fieldnames=[
                            'name', 'city', 'url', 'rating', 'review_count', 'source', 'fail_reason'
                        ])
                        writer.writeheader()
                        for idx, item in enumerate(self.items):
                            r = results[idx] if idx < len(results) and results[idx] else {
                                'rating': 'N/A', 'review_count': 'N/A', 'source': '', 'fail_reason': ''
                            }
                            writer.writerow({
                                'name': item.get('name', ''), 'city': item.get('city', ''),
                                'url': item.get('url', ''), **r
                            })
            if os.path.exists(output_path):
                os.remove(output_path)
            os.rename(output_path + '.tmp', output_path)

        def save_and_checkpoint():
            save_incremental()
            if self.input_file:
                from db_cache import save_checkpoint
                save_checkpoint(self.input_file, output_path, results, total, processed_count[0])

        def emit_status(i, result):
            item = self.items[i]
            eta_str = self._format_eta(time.time() - start_time, processed_count[0], total)
            rating_str = result.get('rating', 'N/A') if result else 'N/A'
            reviews_str = result.get('review_count', 'N/A') if result else 'N/A'
            fail_reason = result.get('fail_reason') if result else None
            if str(rating_str).strip() in ('N/A', '') and fail_reason:
                status = f"{item.get('name', '')[:35]} -> [{fail_reason}]"
            else:
                status = f"{item.get('name', '')[:35]} -> Rating: {rating_str}, Reviews: {reviews_str}"
            if eta_str:
                status += f" • {eta_str}"
            self.progress.emit(processed_count[0], total, status)

        try:
            self.progress.emit(processed_count[0], total, f"Initiating high-speed API engine...")
            scraper = AsyncScraperEngine(concurrency_limit=50)
            
            async_items = []
            for i in list(remaining):
                if self._stop:
                    break
                item = dict(self.items[i])
                url = item.get('url', '').strip()
                name = item.get('name', '')
                city = item.get('city', '')
                source = item.get('source', '')
                identifier = url if url else f"{name}:{city}"
                
                cached = db_cache.get_cached_rating(source or 'booking', identifier)
                if cached:
                    final_res = {
                        'rating': cached['rating'] or 'N/A',
                        'review_count': cached['review_count'] or 'N/A',
                        'source': cached['scraped_source'],
                        'fail_reason': 'Cached'
                    }
                    results[i] = final_res
                    remaining.discard(i)
                    processed_count[0] += 1
                    emit_status(i, final_res)
                    continue
                
                item['idx'] = i
                item['identifier'] = identifier
                async_items.append(item)
                
            if async_items and not self._stop:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                batch_results = loop.run_until_complete(scraper.scrape_batch(async_items))
                loop.close()
                
                for res in batch_results:
                    if self._stop:
                        break
                    i = res.get('item_idx')
                    if i is not None:
                        src_label = self.items[i].get('source', '').capitalize()
                        if 'bcom' in src_label.lower() or 'booking' in src_label.lower(): src_label = 'Booking.com'
                        elif 'mmt' in src_label.lower() or 'makemytrip' in src_label.lower(): src_label = 'MMT'
                        elif 'goibibo' in src_label.lower(): src_label = 'Goibibo'
                        elif 'agoda' in src_label.lower(): src_label = 'Agoda'
                        elif 'expedia' in src_label.lower(): src_label = 'Expedia'
                        
                        rating_val = res.get('rating')
                        if rating_val is None or str(rating_val).strip() == '':
                            rating_val = 'N/A'
                        rc_val = res.get('review_count')
                        if rc_val is None or str(rc_val).strip() == '':
                            rc_val = 'N/A'
                            
                        final_res = {
                            'rating': rating_val,
                            'review_count': rc_val,
                            'source': src_label,
                            'fail_reason': res.get('reason') if str(rating_val) == 'N/A' else None
                        }
                        
                        try:
                            identifier = self.items[i].get('identifier')
                            db_cache.set_cached_rating(self.items[i].get('source') or 'booking', identifier, res.get('rating'), res.get('review_count'), src_label)
                            db_cache.update_batch_run('ratings_batch', self.input_file or 'bulk_input', processed_count[0], total, 'RUNNING', output_path)
                        except Exception as e:
                            pass
                            
                        results[i] = final_res
                        remaining.discard(i)
                        processed_count[0] += 1
                        emit_status(i, final_res)
                        
                        if processed_count[0] % SAVE_INTERVAL == 0:
                            save_and_checkpoint()

        except Exception as e:
            print(f"Async Scrape Error: {e}")

        save_and_checkpoint()
        if not self._stop:
            try:
                db_cache.update_batch_run('ratings_batch', self.input_file or 'bulk_input', total, total, 'COMPLETED', output_path)
            except Exception:
                pass
            self.finished.emit(output_path, processed_count[0], total)

